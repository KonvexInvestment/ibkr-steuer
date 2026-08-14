"""Regression guard for the form-first UI result overview.

app.py is a Streamlit script and cannot be imported without executing the UI.
The small pure rendering helpers are therefore extracted through ``ast``.
The structural test asserts the state architecture (PLAN.md Revision 6)
instead of source positions: two-state model, view-model builder, single
active renderer, no CSS-hide section hack.
"""

import ast
import html
import io
import os
from collections import defaultdict

from openpyxl import load_workbook


APP_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app.py",
)


def load_summary_renderer():
    with open(APP_PATH, encoding="utf-8") as app_file:
        tree = ast.parse(app_file.read())
    wanted = {"fmt", "esc", "kap_row", "build_tax_result_summary_html"}
    nodes = [
        node for node in tree.body
        if isinstance(node, ast.FunctionDef) and node.name in wanted
    ]
    assert {node.name for node in nodes} == wanted
    namespace = {"html": html}
    exec(compile(ast.Module(body=nodes, type_ignores=[]), APP_PATH, "exec"),
         namespace)
    return namespace["build_tax_result_summary_html"]


def load_excel_builder():
    """Load the pure Excel builder without executing the Streamlit script."""
    with open(APP_PATH, encoding="utf-8") as app_file:
        tree = ast.parse(app_file.read())
    wanted_functions = {
        "fmt_de", "_format_instrument", "_get_group_key", "_build_excel",
    }
    wanted_assignments = {
        "topf_readable", "EXPORT_TOPF_ORDER", "cat_labels",
    }
    nodes = []
    for node in tree.body:
        if isinstance(node, ast.FunctionDef) and node.name in wanted_functions:
            nodes.append(node)
        elif isinstance(node, ast.Assign) and any(
                isinstance(target, ast.Name)
                and target.id in wanted_assignments
                for target in node.targets):
            nodes.append(node)
    namespace = {
        "steuerjahr": 2025,
        "defaultdict": defaultdict,
    }
    exec(compile(ast.Module(body=nodes, type_ignores=[]), APP_PATH, "exec"),
         namespace)
    return namespace["_build_excel"]


def test_summary_keeps_every_form_value_and_status():
    render = load_summary_renderer()
    final = {
        "zeile_7": 17.25,
        "zeile_19": 23152.16,
        "zeile_20": 846.82,
        "zeile_22": 27853.77,
        "zeile_23": 397.82,
        "zeile_37": 4.31,
        "zeile_38": 0.24,
        "quellensteuer": 449.75,
    }
    kap_inv_lines = [
        {
            "line": 8,
            "kind": "distribution",
            "fund_type": "Sonstige Investmentfonds",
            "amount_raw_eur": 4823.88,
        },
        {
            "line": 14,
            "kind": "sale",
            "fund_type": "Aktienfonds",
            "amount_raw_eur": -3442.84,
        },
        {
            "line": 26,
            "kind": "sale",
            "fund_type": "Sonstige Investmentfonds",
            "amount_raw_eur": -210.78,
        },
    ]
    rendered = render(
        2024,
        final,
        kap_inv_lines=kap_inv_lines,
        kap_inv_enabled=True,
        kap_status="Berechnet",
        kap_inv_status="Vorläufig / prüfen",
        kap_inv_status_tone="warning",
        review_items=["Vorabpauschalen ergänzen."],
        so_rows=[{
            "line": "SO",
            "label": "Steuerpflichtiger Gewinn/Verlust (≤ 1 Jahr)",
            "value": 123.45,
            "highlight": True,
        }],
    )

    expected_values = [
        "17,25 €",
        "4,31 €",
        "0,24 €",
        "23.152,16 €",
        "846,82 €",
        "27.853,77 €",
        "397,82 €",
        "449,75 €",
        "4.823,88 €",
        "-3.442,84 €",
        "-210,78 €",
        "123,45 €",
    ]
    for value in expected_values:
        assert value in rendered, f"Formularwert fehlt in Übersicht: {value}"

    for line in ("Z. 7", "Z. 19", "Z. 20", "Z. 22", "Z. 23",
                 "Z. 37", "Z. 38", "Z. 41", "Z. 8", "Z. 14", "Z. 26"):
        assert line in rendered, f"Formularzeile fehlt in Übersicht: {line}"

    assert "KAP-INV · Vorläufig / prüfen" in rendered
    assert "vorläufig" in rendered
    assert "Vorabpauschalen ergänzen." in rendered
    assert 'class="kap-value"' in rendered


def test_summary_escapes_dynamic_labels_and_review_text():
    render = load_summary_renderer()
    rendered = render(
        2025,
        {
            "zeile_7": 0,
            "zeile_19": 1,
            "zeile_20": 2,
            "zeile_22": 3,
            "zeile_23": 4,
            "quellensteuer": 5,
        },
        kap_inv_enabled=True,
        kap_inv_lines=[{
            "line": 14,
            "kind": "sale",
            "fund_type": "<unsafe>",
            "amount_raw_eur": 6,
        }],
        review_items=["<script>alert(1)</script>"],
    )
    assert "<unsafe>" not in rendered
    assert "&lt;unsafe&gt;" in rendered
    assert "<script>" not in rendered
    assert "&lt;script&gt;" in rendered


def test_state_architecture_invariants():
    """Die Zustandsarchitektur aus PLAN.md Revision 6 bleibt erhalten."""
    with open(APP_PATH, encoding="utf-8") as app_file:
        source = app_file.read()
    ui_model_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "ui_model.py",
    )
    with open(ui_model_path, encoding="utf-8") as ui_model_file:
        ui_source = ui_model_file.read()

    # Zwei-Zustands-Modell: Start-Screen endet mit st.stop(), bevor der
    # Berichtsmodus beginnt.
    assert 'st-key-start_upload_card' in source
    # Fenster gross genug fuer den Hero-Block inkl. Ablauf-SVG.
    start_pos = source.index('class="start-title"')
    assert 'st.stop()' in source[start_pos:start_pos + 8000]

    # Upload-Snapshot: XML/CSV haben getrennte Callbacks und getrennte
    # Start-/Sidebar-Widget-Keys. Ein leer neu montiertes Schwester-Widget
    # darf die bereits geladenen Daten nicht als vermeintliche Loeschung
    # ueberschreiben.
    assert 'on_change=_capture_xml_uploads' in source
    assert 'on_change=_capture_csv_upload' in source
    assert "_render_uploaders('start')" in source
    assert "_render_uploaders('sidebar')" in source
    assert 'ui_model.update_upload_dataset' in source
    assert 'ui_model.append_xml_uploads' in source
    assert 'ui_model.file_digest' in source
    assert 'def update_upload_dataset' in ui_source
    assert 'build_dataset_id(files)' in ui_source

    # Compute-Cache mit atomarem Commit und Schema-Version.
    assert 'ui_model.build_input_key' in source
    assert 'ui_model.should_commit_snapshot' in source
    assert 'ui_model.snapshot_is_current' in source

    # stdout-Hygiene unter dem prozessweiten Lock aus ui_model.
    assert 'ui_model.REDIRECT_LOCK' in source
    assert 'contextlib.redirect_stdout' in source

    # View-Model-Builder als einzige Ableitungsquelle.
    assert 'ui_model.build_view_model' in source

    # Navigation: stabile Seiten-IDs, Normalisierung, genau EIN aktiver
    # Renderer — kein CSS-Verstecken aller Sektionen mehr.
    assert 'ui_model.normalize_nav' in source
    assert '_PAGE_RENDERERS[_nav_current]()' in source
    assert 'detail_section_' not in source
    assert "st.container(key=\"detail_section" not in source

    # Domain-State getrennt vom Widget-State (dataset-genamespaced _ui_-Keys).
    assert "_ui_tg_variante_b_" in source
    assert "_ui_etf_conf_" in source
    assert "_ui_so_overrides_" in source
    assert "'etf_overrides'" in source

    # Export-Cache haengt am view_key.
    assert "cache.get('view_key') == vm['view_key']" in source

    # Alle sieben Renderer existieren.
    for renderer in ("def render_overview", "def render_kap",
                     "def render_kap_inv", "def render_anlage_so",
                     "def render_prueffaelle", "def render_rechenwege",
                     "def render_export"):
        assert renderer in source, f"Renderer fehlt: {renderer}"

    # Die Übersicht dupliziert keine Detailtexte aus dem eigenen
    # Prüffälle-Bereich. Sie zeigt nur Anzahl und klare Handlungsaufforderung.
    overview = source[
        source.index('def render_overview'):
        source.index('def render_kap')
    ]
    assert 'offene Prüffälle – bitte vor der Abgabe' in overview
    assert 'review_items=review_items' in overview
    assert 'render_notices(kritisch_notices' not in overview
    assert 'notice_html(notice)' not in overview


def test_excel_contains_blocked_partnership_details_and_summary():
    build_excel = load_excel_builder()
    final = {
        'stocks_gain': 0.0, 'stocks_loss': 0.0, 'topf_1': 0.0,
        'dividends': 0.0, 'interest': 0.0, 'options_gain': 0.0,
        'options_loss': 0.0, 'topf_2': 0.0, 'zeile_7': 0.0,
        'zeile_19': 0.0, 'zeile_20': 0.0, 'zeile_22': 0.0,
        'zeile_23': 0.0, 'zeile_37': 0.0, 'zeile_38': 0.0,
        'quellensteuer': 0.0, 'etf_wht': 0.0,
    }
    row = {
        'reportDate': '2025-06-01', 'dateTime': '2025-06-01 10:00:00',
        'symbol': 'USO', 'description': 'United States Oil Fund LP',
        'isin': 'US91232N2071', 'assetCategory': 'STK',
        'subCategory': 'ETF', 'buySell': 'SELL', 'quantity': '-10',
        'tradePrice': 70.0, 'cost': -600.0, 'proceeds': 700.0,
        'fifoPnlRealized': 100.0, 'fxRateToBase': 1.0,
        'pnl_eur': 100.0, 'topf': 'Personengesellschaft',
        'source': 'trades',
    }
    partnership_item = {
        'ticker': 'USO', 'observed_trade_pnl_eur': 100.0,
        'observed_tageskurs_delta_eur': 4.0,
        'observed_distributions_eur': 20.0,
        'observed_other_cash_eur': 5.0,
        'observed_withholding_tax_eur': -3.0,
    }
    content = build_excel(
        [row], {'Personengesellschaft': [row]}, {
            'final': final, 'base_currency': 'EUR',
            'created_at': '09.08.2026 10:00', 'has_etf_data': False,
            'invstg_aktiv': True, 'kap_inv_form': {},
            'no_invstg_summary': {},
            'partnership_tax_items': {'US91232N2071': partnership_item},
            'has_so_data': False, 'so_taxable': 0.0, 'so_free': 0.0,
        })
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    detail_values = {
        cell.value for row_cells in workbook['Trade-Details 2025'].iter_rows()
        for cell in row_cells if cell.value is not None
    }
    summary_values = {
        cell.value for row_cells in workbook['Zusammenfassung'].iter_rows()
        for cell in row_cells if cell.value is not None
    }
    assert 'US91232N2071' in detail_values
    assert 'USO (United States Oil Fund LP)' in detail_values
    assert any(
        isinstance(value, str) and value.startswith('Personengesellschaft -')
        for value in detail_values)
    assert 'Personengesellschaft · blockiert' in summary_values
    assert 'USO (US91232N2071) · beobachtete Brokerwerte' in summary_values


if __name__ == "__main__":
    test_summary_keeps_every_form_value_and_status()
    test_summary_escapes_dynamic_labels_and_review_text()
    test_state_architecture_invariants()
    test_excel_contains_blocked_partnership_details_and_summary()
    print("OK: Formularzentrierte Übersicht enthält alle Steuerzeilen")

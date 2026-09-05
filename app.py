"""IBKR Steuerbericht — Streamlit-UI.

Architektur (siehe PLAN.md, Revision 6):
  Upload-Snapshot (SHA-256, validiert)
    -> kontoweiser Compute-Snapshot (unveraenderlich, Cache-Protokoll)
    -> reiner View-Model-Builder (ui_model.build_view_model)
    -> genau ein aktiver Renderer (Sidebar-Navigation)

Fachliche Eingaben (Toggles, Fondsbestaetigungen, Anlage-SO-Overrides) leben
als Domain-State in st.session_state, genamespaced per dataset_id — Widgets
tragen nur _ui_*-Keys und kopieren per Callback in den Domain-State, weil
Streamlit den State nicht gerenderter Widgets am Run-Ende loescht.

Die Datei laeuft unveraendert unter lokalem Streamlit UND stlite/Pyodide
(index.html) — keine Fremdkomponenten, keine externen Fonts, kein Netz.
"""

import contextlib
import copy
import html
import io
import os
import tempfile
import time
from collections import defaultdict
from datetime import datetime as _dt

import streamlit as st

import extract_ibkr_data
import calculate_tax_report
import ui_model
from etf_classification import (
    classification_catalog_to_csv,
    get_classification,
    get_classification_catalog,
    get_etf_info,
    is_anlage_so,
)
from ibkr_dates import is_supported_ibkr_date, normalize_ibkr_date

st.set_page_config(
    page_title="IBKR Steuerbericht",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Designsystem ─────────────────────────────────────────────────────────────
# Apple-Dark im Konvex-Stil, Werkzeug-Variante: neutrale Grautoene, genau ein
# Akzentblau, semantische Statusfarben, System-Font-Stack (Offline-Paritaet
# lokal/stlite). Signatur: die ELSTER-Formularzeile als Objekt (Zeilen-Chip,
# gepunktete Fuehrungslinie, tabellarischer Betrag).

st.markdown("""
<style>
    :root {
        --bg: #0c0e12;
        --bg-raised: #10131a;
        --card: #151924;
        --card-soft: rgba(255,255,255,0.025);
        --line: rgba(255,255,255,0.07);
        --line-soft: rgba(255,255,255,0.045);
        --accent: #7aa5f8;
        --accent-dim: rgba(122,165,248,0.13);
        --accent-border: rgba(122,165,248,0.28);
        --ok: #3ecf8e;
        --ok-dim: rgba(62,207,142,0.10);
        --warn: #e8b04b;
        --warn-dim: rgba(232,176,75,0.10);
        --crit: #f0716c;
        --crit-dim: rgba(240,113,108,0.10);
        --ink: #eef1f6;
        --ink-2: #9aa3b2;
        --ink-3: #667082;
        --radius-xs: 6px;
        --radius-sm: 10px;
        --radius: 12px;
        --radius-lg: 16px;
    }

    html, body, [class*="css"] {
        font-family: -apple-system, BlinkMacSystemFont, "SF Pro Text",
                     "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    }
    .stApp { background: var(--bg); color: var(--ink-2); }
    .block-container { padding: 1.4rem 2rem 3.5rem 2rem; max-width: 1080px; }
    #MainMenu, footer { visibility: hidden; }
    /* The Streamlit header owns the control that reopens a collapsed sidebar. */
    [data-testid="stHeader"] { background: transparent; }
    [data-testid="stAppDeployButton"] { display: none; }

    /* ── Typo-Rollen ── */
    .page-title {
        color: var(--ink);
        font-size: 1.45rem;
        font-weight: 700;
        letter-spacing: -0.02em;
        margin: 0 0 0.15rem 0;
    }
    .page-sub { color: var(--ink-3); font-size: 0.85rem; margin: 0 0 1.5rem 0; }
    .eyebrow {
        color: var(--ink-3);
        font-size: 0.68rem;
        font-weight: 700;
        letter-spacing: 0.09em;
        text-transform: uppercase;
        margin: 0 0 0.5rem 0;
    }
    .section-title {
        color: var(--ink);
        font-size: 0.95rem;
        font-weight: 650;
        letter-spacing: -0.01em;
        margin: 2rem 0 0.7rem 0;
        padding-bottom: 0.45rem;
        border-bottom: 1px solid var(--line-soft);
    }
    .meta-row {
        display: flex; flex-wrap: wrap; gap: 0.4rem 1.4rem;
        color: var(--ink-3); font-size: 0.78rem; margin-bottom: 1.1rem;
    }
    .meta-row strong { color: var(--ink-2); font-weight: 600; }

    /* ── Karten ── */
    .card {
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: var(--radius-lg);
        padding: 1.1rem 1.2rem;
        margin-bottom: 1rem;
    }
    .card-title {
        color: var(--ink); font-size: 1rem; font-weight: 650;
        letter-spacing: -0.01em; margin-bottom: 0.15rem;
    }
    .card-sub { color: var(--ink-3); font-size: 0.76rem; line-height: 1.5; }

    /* ── Formularzeilen (Signatur) ── */
    .kap-row {
        display: flex; align-items: baseline; gap: 0.7rem;
        padding: 0.62rem 0.2rem;
        border-bottom: 1px solid var(--line-soft);
    }
    .kap-row:last-child { border-bottom: none; }
    .kap-left { display: flex; align-items: baseline; gap: 0.7rem; min-width: 0; }
    .kap-badge {
        font-size: 0.68rem; font-weight: 700;
        font-variant-numeric: tabular-nums;
        color: var(--accent);
        background: var(--accent-dim);
        border: 1px solid var(--accent-border);
        border-radius: var(--radius-xs);
        padding: 0.14rem 0.42rem;
        white-space: nowrap; flex-shrink: 0;
    }
    .kap-desc {
        font-size: 0.86rem; color: var(--ink-2); font-weight: 500;
        min-width: 0;
    }
    .kap-leader {
        flex: 1; min-width: 1.2rem;
        border-bottom: 1px dotted rgba(255,255,255,0.14);
        transform: translateY(-0.28em);
    }
    .kap-value {
        font-size: 0.95rem; font-weight: 650;
        font-variant-numeric: tabular-nums;
        text-align: right; white-space: nowrap; flex-shrink: 0;
    }
    .kap-row.highlight .kap-desc { color: var(--ink); font-weight: 600; }
    .kap-row.highlight .kap-value { font-size: 1.05rem; font-weight: 700; }
    .kap-row.highlight .kap-badge {
        background: var(--accent); color: #0c0e12; border-color: var(--accent);
    }

    /* ── Status-Chips ── */
    .status-chip {
        display: inline-flex; align-items: center; gap: 0.3rem;
        border-radius: 999px;
        padding: 0.24rem 0.6rem;
        font-size: 0.68rem; font-weight: 700; white-space: nowrap;
        border: 1px solid var(--line);
        color: var(--ink-2); background: var(--card-soft);
    }
    .status-chip.ok { color: var(--ok); border-color: rgba(62,207,142,0.3); background: var(--ok-dim); }
    .status-chip.warning { color: var(--warn); border-color: rgba(232,176,75,0.3); background: var(--warn-dim); }
    .status-chip.crit { color: var(--crit); border-color: rgba(240,113,108,0.3); background: var(--crit-dim); }
    .kap-row-status { margin-left: 0.5rem; transform: translateY(-0.1em); }

    /* ── Metric-Grid ── */
    .metric-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
        gap: 0.7rem; margin-bottom: 0.8rem;
    }
    .metric-card {
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: var(--radius);
        padding: 0.85rem 1rem; min-width: 0;
    }
    .metric-label { font-size: 0.74rem; color: var(--ink-3); margin-bottom: 0.3rem; font-weight: 550; }
    .metric-value {
        font-size: clamp(0.95rem, 3vw, 1.2rem); font-weight: 650;
        font-variant-numeric: tabular-nums;
        word-break: break-word; line-height: 1.25;
    }
    .metric-value.green { color: var(--ok); }
    .metric-value.red { color: var(--crit); }
    .metric-value.white { color: var(--ink); }
    .metric-card.saldo { border-color: var(--accent-border); background: var(--accent-dim); }
    .metric-card.saldo .metric-label { color: var(--accent); }

    /* ── Hinweis-System (drei Klassen, je EIN Stil) ── */
    .notice {
        border: 1px solid var(--line);
        border-left: 3px solid var(--ink-3);
        border-radius: 0 var(--radius) var(--radius) 0;
        background: var(--card-soft);
        padding: 0.8rem 1rem;
        margin-bottom: 0.7rem;
        font-size: 0.8rem; color: var(--ink-2); line-height: 1.55;
    }
    .notice-title { font-weight: 650; color: var(--ink); margin-bottom: 0.15rem; }
    .notice.prueffall { border-left-color: var(--warn); }
    .notice.prueffall .notice-title { color: var(--warn); }
    .notice.kritisch { border-left-color: var(--crit); background: var(--crit-dim); }
    .notice.kritisch .notice-title { color: var(--crit); }
    .notice.transparenz { border-left-color: var(--accent); }
    .notice.transparenz .notice-title { color: var(--accent); }
    .notice.fehler { border-left-color: var(--crit); background: var(--crit-dim); }
    .notice.fehler .notice-title { color: var(--crit); }
    .notice-target { color: var(--ink-3); font-size: 0.72rem; margin-top: 0.3rem; }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {
        background: var(--bg-raised);
        border-right: 1px solid var(--line-soft);
    }
    [data-testid="stSidebar"][aria-expanded="true"] {
        min-width: 300px !important; max-width: 300px !important;
    }
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] { padding-top: 0.9rem; }
    .sidebar-brand {
        color: var(--ink); font-size: 0.95rem; font-weight: 700;
        letter-spacing: -0.01em; margin-bottom: 0.1rem;
    }
    .sidebar-brand-sub { color: var(--ink-3); font-size: 0.72rem; margin-bottom: 0.9rem; }
    .data-card {
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: var(--radius);
        padding: 0.7rem 0.8rem; margin-bottom: 0.6rem;
        font-size: 0.75rem; color: var(--ink-2); line-height: 1.5;
    }
    .data-card-title { color: var(--ink); font-weight: 650; font-size: 0.78rem; margin-bottom: 0.2rem; }
    .data-card-files { color: var(--ink-3); font-size: 0.7rem; overflow-wrap: anywhere; }

    /* Navigation (st.button-Liste im Container mit key="nav_buttons") */
    .st-key-nav_buttons [data-testid="stButton"] { margin-bottom: 0.12rem; }
    .st-key-nav_buttons [data-testid="stButton"] button {
        border: 1px solid transparent;
        justify-content: flex-start;
        min-height: 2.2rem;
        text-align: left;
        font-size: 0.85rem;
        color: var(--ink-2);
        background: transparent;
        border-radius: var(--radius-sm);
        width: 100%;
    }
    .st-key-nav_buttons [data-testid="stButton"] button:hover {
        background: var(--card-soft);
        border-color: var(--line-soft);
        color: var(--ink);
    }
    .st-key-nav_buttons [data-testid="stButton"] button[kind="primary"] {
        background: var(--accent-dim);
        border-color: var(--accent-border);
        color: var(--ink);
        font-weight: 600;
    }

    /* ── Start-Screen ── */
    .stApp:has(.start-title) .block-container {
        max-width: 1080px;
        padding-top: 7vh;
    }
    .stApp:has(.start-title) {
        background:
            radial-gradient(circle at 18% 20%, rgba(122,165,248,0.10), transparent 28rem),
            radial-gradient(circle at 82% 72%, rgba(62,207,142,0.045), transparent 24rem),
            var(--bg);
    }
    .start-title {
        color: var(--ink); font-size: clamp(2rem, 4vw, 3.15rem);
        font-weight: 750; line-height: 1.05; letter-spacing: -0.045em;
        margin-bottom: 0.9rem; max-width: 20ch;
    }
    .start-sub {
        color: var(--ink-2); font-size: 0.95rem; line-height: 1.65;
        margin-bottom: 1.35rem; max-width: 34rem;
    }
    .start-trust {
        display: flex; flex-wrap: wrap; gap: 0.45rem;
        margin-bottom: 1.7rem;
    }
    .start-trust-chip {
        display: inline-flex; align-items: center;
        color: var(--ok); background: var(--ok-dim);
        border: 1px solid rgba(62,207,142,0.25); border-radius: 999px;
        padding: 0.28rem 0.62rem; font-size: 0.7rem; font-weight: 650;
    }
    /* Ablauf-Grafik: dieselbe Formular-Sprache wie die kap-row
       (Zeilen-Badges, gepunktete Fuehrungslinien, tabellarische Werte). */
    .start-visual { max-width: 33rem; margin-top: 1.1rem; }
    .start-visual svg { width: 100%; height: auto; display: block; }
    .start-visual text { font-family: inherit; }
    .sv-card { fill: var(--card); stroke: var(--line); }
    .sv-code { fill: rgba(255,255,255,0.10); }
    .sv-glyph {
        fill: var(--accent); font-size: 13px; font-weight: 650;
        font-family: ui-monospace, "SF Mono", Menlo, Consolas, monospace;
    }
    .sv-flow {
        stroke: var(--accent-border); stroke-width: 1.6;
        stroke-dasharray: 5 7; fill: none;
        animation: sv-flow 2.8s linear infinite;
    }
    .sv-arrow { fill: var(--accent-border); }
    .sv-check {
        stroke: var(--ok); stroke-width: 3; fill: none;
        stroke-linecap: round; stroke-linejoin: round;
    }
    .sv-badge { fill: var(--accent-dim); stroke: var(--accent-border); }
    .sv-badge.fill { fill: var(--accent); stroke: var(--accent); }
    .sv-badge-text { fill: var(--accent); font-size: 9.5px; font-weight: 700; }
    .sv-badge-text.fill { fill: var(--bg); }
    .sv-leader {
        stroke: rgba(255,255,255,0.14); stroke-width: 1.4;
        stroke-dasharray: 1.5 3.6;
    }
    .sv-value {
        fill: var(--ink); font-size: 10.5px; font-weight: 650;
        font-variant-numeric: tabular-nums;
    }
    .sv-label { fill: var(--ink-2); font-size: 12.5px; font-weight: 600; }
    .sv-label tspan { fill: var(--accent); font-weight: 700; }
    @keyframes sv-flow { to { stroke-dashoffset: -24; } }
    @media (prefers-reduced-motion: reduce) {
        .sv-flow { animation: none; }
    }
    .st-key-start_upload_card {
        background: rgba(21,25,36,0.92);
        border: 1px solid rgba(122,165,248,0.18);
        border-radius: var(--radius-lg); padding: 1.15rem 1.2rem 1rem 1.2rem;
        box-shadow: 0 24px 70px rgba(0,0,0,0.24);
    }
    .start-upload-kicker {
        color: var(--accent); font-size: 0.68rem; font-weight: 700;
        letter-spacing: 0.08em; text-transform: uppercase;
        margin-bottom: 0.3rem;
    }
    .start-upload-title {
        color: var(--ink); font-size: 1.12rem; font-weight: 700;
        letter-spacing: -0.02em; margin-bottom: 0.25rem;
    }
    .start-upload-copy {
        color: var(--ink-3); font-size: 0.76rem; line-height: 1.5;
        margin-bottom: 0.8rem;
    }
    .stApp:has(.start-title) [data-testid="stFileUploader"] {
        background: rgba(122,165,248,0.055);
        border-color: rgba(122,165,248,0.28);
        transition: border-color 160ms ease, background 160ms ease,
                    transform 160ms ease;
    }
    .stApp:has(.start-title) [data-testid="stFileUploader"]:hover {
        border-color: rgba(122,165,248,0.55);
        background: rgba(122,165,248,0.09);
        transform: translateY(-1px);
    }
    .start-foot {
        color: var(--ink-3); background: var(--card-soft);
        border: 1px solid var(--line-soft); border-left: 3px solid var(--warn);
        border-radius: 0 var(--radius) var(--radius) 0; padding: 0.65rem 0.8rem;
        font-size: 0.7rem; line-height: 1.5; margin-top: 1.35rem;
        text-align: left;
    }
    .start-foot strong { color: var(--ink-2); font-weight: 700; }

    /* ── Streamlit-Widgets angleichen ── */
    [data-testid="stFileUploader"] {
        background: var(--card-soft);
        border: 1px dashed rgba(255,255,255,0.14);
        border-radius: var(--radius-lg);
        padding: 0.4rem;
    }
    [data-testid="stFileUploaderDropzoneInstructions"],
    [data-testid="stFileUploaderDropzoneInstructions"] * { font-size: 0 !important; }
    [data-testid="stFileUploaderDropzoneInstructions"] svg { height: 2.2rem; width: 2.2rem; }
    [data-testid="stFileUploaderDropzoneInstructions"]::after {
        content: "Dateien hierher ziehen";
        font-size: 0.85rem; color: var(--ink-2);
    }
    [data-testid="stFileUploaderDropzone"] button { font-size: 0 !important; }
    [data-testid="stFileUploaderDropzone"] button::after {
        content: "Dateien auswählen"; font-size: 0.85rem;
    }
    [data-testid="stExpander"] {
        border: 1px solid var(--line-soft) !important;
        border-radius: var(--radius) !important;
        background: var(--card-soft) !important;
    }
    /* Streamlit-Alerts ins Designsystem einpassen (ein Stil pro Klasse) */
    [data-testid="stAlert"] {
        border-radius: 0 var(--radius) var(--radius) 0;
        border: 1px solid var(--line);
        font-size: 0.82rem;
        overflow: hidden;
    }
    /* Streamlits Default-Alert-Flaeche neutralisieren; die Farbe kommt
       ausschliesslich aus den :has()-Regeln unten (ein Stil pro Klasse). */
    [data-testid="stAlert"] > div,
    [data-testid="stAlertContainer"] { background: transparent !important; }
    [data-testid="stAlertContentWarning"] { color: var(--ink-2); }
    [data-testid="stAlertContentInfo"] { color: var(--ink-2); }
    [data-testid="stAlertContentSuccess"] { color: var(--ink-2); }
    [data-testid="stAlertContentError"] { color: var(--ink-2); }
    div[data-testid="stAlert"]:has([data-testid="stAlertContentWarning"]) {
        background: var(--warn-dim); border-left: 3px solid var(--warn);
    }
    div[data-testid="stAlert"]:has([data-testid="stAlertContentInfo"]) {
        background: var(--accent-dim); border-left: 3px solid var(--accent);
    }
    div[data-testid="stAlert"]:has([data-testid="stAlertContentSuccess"]) {
        background: var(--ok-dim); border-left: 3px solid var(--ok);
    }
    div[data-testid="stAlert"]:has([data-testid="stAlertContentError"]) {
        background: var(--crit-dim); border-left: 3px solid var(--crit);
    }
    [data-testid="stDownloadButton"] > button {
        width: 100%;
        background: var(--accent-dim);
        border: 1px solid var(--accent-border);
        color: var(--accent);
        border-radius: var(--radius-sm); font-weight: 600; padding: 0.6rem;
    }
    [data-testid="stDownloadButton"] > button:hover { background: rgba(122,165,248,0.22); }
    /* Sekundaere Buttons (ausserhalb der Navigation) in die Karten-Sprache */
    [data-testid="stButton"] button[kind="secondary"] {
        background: var(--card-soft);
        border: 1px solid var(--line);
        color: var(--ink-2);
        border-radius: var(--radius-sm);
        font-size: 0.83rem; font-weight: 600;
    }
    [data-testid="stButton"] button[kind="secondary"]:hover {
        border-color: var(--accent-border);
        color: var(--ink);
        background: var(--card);
    }
    [data-testid="stCheckbox"] label p,
    [data-testid="stSelectbox"] label p,
    [data-testid="stMultiSelect"] label p,
    .stSelectbox label, .stCheckbox label { color: var(--ink-2) !important; }
    /* Select-/Multiselect-Flaechen in die Karten-Sprache */
    [data-baseweb="select"] > div {
        background: var(--card);
        border-color: var(--line);
    }
    [data-testid="stTabs"] [data-baseweb="tab"] { color: var(--ink-2); }
    [data-testid="stTabs"] [aria-selected="true"] { color: var(--ink); }

    /* ── Tabellen (Markdown) in der Formular-Sprache ── */
    [data-testid="stMarkdownContainer"] table {
        border-collapse: collapse;
        font-size: 0.8rem;
        margin: 0.3rem 0 0.9rem 0;
    }
    [data-testid="stMarkdownContainer"] table th {
        color: var(--ink-3);
        font-size: 0.66rem; font-weight: 700;
        letter-spacing: 0.07em; text-transform: uppercase;
        text-align: left;
        border: none; border-bottom: 1px solid var(--line);
        padding: 0.35rem 0.85rem 0.35rem 0;
        background: transparent;
    }
    [data-testid="stMarkdownContainer"] table td {
        color: var(--ink-2);
        font-variant-numeric: tabular-nums;
        border: none; border-bottom: 1px solid var(--line-soft);
        padding: 0.42rem 0.85rem 0.42rem 0;
        background: transparent;
    }
    [data-testid="stMarkdownContainer"] table tr:last-child td { border-bottom: none; }
    [data-testid="stMarkdownContainer"] table tbody tr:hover td { background: var(--card-soft); }
    [data-testid="stMarkdownContainer"] table td strong,
    [data-testid="stMarkdownContainer"] table th strong { color: var(--ink); }

    /* ── Dataframes einrahmen ── */
    [data-testid="stDataFrame"] {
        border: 1px solid var(--line);
        border-radius: var(--radius);
        overflow: hidden;
    }

    /* ── Markdown-Header in die Typo-Hierarchie einpassen ── */
    [data-testid="stMarkdownContainer"] h3 {
        color: var(--ink);
        font-size: 1rem; font-weight: 650;
        letter-spacing: -0.01em;
        margin: 1.2rem 0 0.4rem 0; padding: 0;
    }
    [data-testid="stMarkdownContainer"] h4 {
        color: var(--ink-2);
        font-size: 0.86rem; font-weight: 650;
        margin: 0.9rem 0 0.3rem 0; padding: 0;
    }
    [data-testid="stTabs"] [data-baseweb="tab-highlight"] {
        background-color: var(--accent);
    }

    /* ── Doku-Fliesstext in Expandern ── */
    [data-testid="stExpander"] [data-testid="stMarkdownContainer"] p,
    [data-testid="stExpander"] [data-testid="stMarkdownContainer"] li {
        font-size: 0.82rem; color: var(--ink-2); line-height: 1.6;
    }
    [data-testid="stExpander"] [data-testid="stMarkdownContainer"] code {
        background: var(--card-soft);
        border: 1px solid var(--line-soft);
        border-radius: var(--radius-xs);
        padding: 0.05rem 0.3rem;
        font-size: 0.74rem;
    }

    .badge-beta {
        display: inline-block;
        border: 1px solid var(--warn);
        color: var(--warn);
        border-radius: 6px;
        font-size: 0.66rem; font-weight: 700; letter-spacing: 0.06em;
        padding: 0.12rem 0.45rem;
        text-transform: uppercase;
    }

    @media (max-width: 900px) {
        [data-testid="stSidebar"][aria-expanded="true"] {
            min-width: 252px !important;
            max-width: 252px !important;
        }
        .block-container { padding-left: 1rem; padding-right: 1rem; }
        [data-testid="stMarkdownContainer"] {
            max-width: 100%;
            overflow-x: auto;
        }
        [data-testid="stMarkdownContainer"] table {
            font-size: 0.76rem;
        }
        [data-testid="stMarkdownContainer"] table:not(:has(tr > :nth-child(5))) {
            width: 100%;
            min-width: 100%;
        }
        [data-testid="stMarkdownContainer"] table:has(tr > :nth-child(5)) {
            min-width: max-content;
        }
        [data-testid="stMarkdownContainer"] th { white-space: normal; }
        [data-testid="stMarkdownContainer"] th,
        [data-testid="stMarkdownContainer"] td {
            padding: 0.4rem 0.55rem;
        }
    }

    @media (max-width: 640px) {
        .block-container { padding: 1rem 0.8rem 2rem 0.8rem; }
        .kap-row { flex-wrap: wrap; }
        .kap-leader { display: none; }
        .kap-value { margin-left: auto; }
        .stApp:has(.start-title) .block-container { padding-top: 2.5vh; }
        .start-title { font-size: 2.15rem; max-width: none; }
        .start-sub { font-size: 0.88rem; }
        .start-visual { max-width: none; margin-bottom: 0.8rem; }
        .st-key-start_upload_card { padding: 1rem; }
    }
</style>
""", unsafe_allow_html=True)


# ── Format- und Baustein-Helper ──────────────────────────────────────────────

def fmt(value: float, decimals: int = 2) -> str:
    s = f"{value:,.{decimals}f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.') + " €"


def fmt_de(value: float, decimals: int = 2) -> str:
    s = f"{value:,.{decimals}f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def color_class(value: float) -> str:
    if value > 0:
        return "green"
    elif value < 0:
        return "red"
    return "white"


def esc(value) -> str:
    return html.escape(str(value))


def section_title(text: str):
    st.markdown(f'<div class="section-title">{esc(text)}</div>',
                unsafe_allow_html=True)


def eyebrow(text: str):
    st.markdown(f'<div class="eyebrow">{esc(text)}</div>',
                unsafe_allow_html=True)


def metric_card(label: str, value: float, variant: str = "auto") -> str:
    """variant: auto | gain | loss | saldo | info"""
    if variant == "auto":
        variant = "gain" if value > 0 else ("loss" if value < 0 else "info")
    # Informations-/Kontrollwerte bleiben neutral: eine negative Korrektur ist
    # kein Fehlerzustand. Gruen/Rot nur fuer echte Gewinn-/Verlustkarten.
    val_color = "white" if variant in ("info", "saldo") else color_class(value)
    css_variant = "saldo" if variant == "saldo" else ""
    return (
        f'<div class="metric-card {css_variant}">'
        f'<div class="metric-label">{esc(label)}</div>'
        f'<div class="metric-value {val_color}">{fmt(value)}</div>'
        f'</div>'
    )


def metric_grid(*cards) -> str:
    return '<div class="metric-grid">' + ''.join(c for c in cards if c) + '</div>'


def kap_row(zeile: str, label: str, value: float, highlight: bool = False,
            force_positive: bool = False, status: str = "",
            status_tone: str = "ok") -> str:
    """Eine Formularzeile: Zeilen-Chip, Label, Fuehrungslinie, Betrag."""
    display_val = abs(value) if force_positive else value
    hl = "highlight" if highlight else ""
    status_html = ""
    if status:
        safe_tone = status_tone if status_tone in {"ok", "warning", "crit"} else ""
        status_html = (
            f'<span class="status-chip {safe_tone} kap-row-status">'
            f'{esc(status)}</span>'
        )
    # Ein zusammenhaengendes HTML-Fragment: eine Leerzeile in rohem HTML laesst
    # Streamlits Markdown-Renderer die Folge-Tags als Text anzeigen.
    return (
        f'<div class="kap-row {hl}">'
        f'<div class="kap-left">'
        f'<span class="kap-badge">{esc(zeile)}</span>'
        f'<span class="kap-desc">{esc(label)}</span>'
        f'{status_html}'
        f'</div>'
        f'<span class="kap-leader"></span>'
        f'<span class="kap-value" style="color:#eef1f6">{fmt(display_val)}</span>'
        f'</div>'
    )


def notice_html(notice: dict, show_target: bool = True) -> str:
    """Einheitliche Darstellung eines Hinweises aus ui_model.collect_notices."""
    cls = notice.get('class', 'transparenz')
    severity = notice.get('severity', 'normal')
    css = cls if cls in ('prueffall', 'transparenz', 'fehler') else 'transparenz'
    if cls == 'prueffall' and severity == 'kritisch':
        css += ' kritisch'
    target_html = ''
    if show_target and notice.get('target'):
        target_label = ui_model.page_label(notice['target'])
        target_html = (
            f'<div class="notice-target">Bereich: {esc(target_label)}</div>'
        )
    return (
        f'<div class="notice {css}">'
        f'<div class="notice-title">{esc(notice.get("title", ""))}</div>'
        f'{esc(notice.get("body", ""))}'
        f'{target_html}'
        f'</div>'
    )


def render_notices(notices, show_target: bool = True):
    for notice in notices:
        st.markdown(notice_html(notice, show_target=show_target),
                    unsafe_allow_html=True)


def build_tax_result_summary_html(
        steuerjahr: int,
        final: dict,
        kap_inv_lines=None,
        kap_inv_enabled: bool = False,
        kap_status: str = "Berechnet",
        kap_status_tone: str = "ok",
        kap_inv_status: str = "",
        kap_inv_status_tone: str = "ok",
        review_items=None,
        so_rows=None) -> str:
    """Formularzentrierte Eintragungsuebersicht — keine Zeile faellt weg."""
    kap_inv_lines = list(kap_inv_lines or [])
    review_items = list(review_items or [])
    so_rows = list(so_rows or [])

    status_html = (
        f'<span class="status-chip {kap_status_tone}">KAP · '
        f'{html.escape(kap_status)}</span>'
    )
    if kap_inv_enabled:
        status_html += (
            f'<span class="status-chip {kap_inv_status_tone}">KAP-INV · '
            f'{html.escape(kap_inv_status or "Berechnet")}</span>'
        )

    rows_html = '<div class="eyebrow" style="margin-top:0.4rem;">Anlage KAP</div>'
    if abs(final.get('zeile_7', 0)) > 0.01:
        rows_html += (
            kap_row("Z. 7", "Kapitalerträge mit inländischem Steuerabzug",
                    final['zeile_7'], highlight=True)
            + kap_row("Z. 37", "Kapitalertragsteuer", final['zeile_37'])
            + kap_row("Z. 38", "Solidaritätszuschlag", final['zeile_38'])
        )
    rows_html += (
        kap_row("Z. 19", "Ausländische Kapitalerträge (Netto)",
                final['zeile_19'], highlight=True)
        + kap_row("Z. 20", "Davon: Aktiengewinne", final['zeile_20'])
        + kap_row("Z. 22", "Verluste ohne Aktien", final['zeile_22'],
                  force_positive=True)
        + kap_row("Z. 23", "Aktienverluste", final['zeile_23'],
                  force_positive=True)
        + kap_row("Z. 41", "Anrechenbare ausländische Quellensteuer",
                  final['quellensteuer'])
    )

    if kap_inv_enabled:
        rows_html += (
            '<div class="eyebrow" style="margin-top:1.1rem;">Anlage KAP-INV</div>'
        )
        if kap_inv_lines:
            for line in kap_inv_lines:
                is_sale = line.get('kind') == 'sale'
                kind_label = (
                    "Ausschüttungen" if not is_sale
                    else "Veräußerungsergebnis vor Abzug von Vorabpauschalen"
                )
                rows_html += kap_row(
                    f"Z. {line['line']}",
                    f"{kind_label} · {line['fund_type']} (vor TFS)",
                    line['amount_raw_eur'],
                    highlight=True,
                    status="vorläufig" if is_sale else "berechnet",
                    status_tone="warning" if is_sale else "ok",
                )
        else:
            rows_html += (
                '<div class="card-sub">Keine KAP-INV-Formularzeile erzeugt. '
                'Offene Produktzuordnungen stehen bei den Prüffällen.</div>'
            )

    if so_rows:
        rows_html += (
            '<div class="eyebrow" style="margin-top:1.1rem;">Anlage SO (§23 EStG)</div>'
        )
        for row in so_rows:
            rows_html += kap_row(
                row['line'], row['label'], row['value'],
                highlight=row.get('highlight', False),
            )

    review_html = ""
    if review_items:
        items = "".join(
            f"<li>{html.escape(str(item))}</li>" for item in review_items
        )
        review_html = (
            '<div class="notice prueffall" style="margin-top:0.9rem;">'
            '<div class="notice-title">Noch zu prüfen</div>'
            f'<ul style="margin:0.3rem 0 0 1.1rem;padding:0;">{items}</ul></div>'
        )

    return f"""
    <div class="card">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:0.75rem;flex-wrap:wrap;margin-bottom:0.6rem;">
            <div>
                <div class="card-title">Eintragungsübersicht {steuerjahr}</div>
                <div class="card-sub">Die Formularwerte zum Übertragen; Rechenwege und Nachweise stehen in den Bereichen links.</div>
            </div>
            <div style="display:flex;gap:0.4rem;flex-wrap:wrap;">{status_html}</div>
        </div>
        {rows_html}
        {review_html}
    </div>
    """

def classify_xmls(xml_files):
    """Parse XMLs to extract accountId and date range, group by account.
    Returns accounts, multi-statement files and invalid XML diagnostics.

    Latest XML per account = tax year, older = history.  Every selected file
    must be an IBKR Flex export; callers fail closed when a file is invalid or
    bundles several FlexStatements, because either case would otherwise create
    a plausible-looking partial tax report.
    """
    import xml.etree.ElementTree as ET
    accounts = {}
    multi_stmt_files = []
    invalid_files = []
    for xml_file in xml_files:
        try:
            content = xml_file.getvalue()
            root = ET.fromstring(content)
            all_stmts = root.findall('.//FlexStatement')
            if not all_stmts:
                invalid_files.append({
                    'name': xml_file.name,
                    'reason': 'kein FlexStatement gefunden',
                })
                continue
            if len(all_stmts) > 1:
                multi_stmt_files.append({
                    'name': xml_file.name,
                    'account_ids': [s.get('accountId', '?') for s in all_stmts],
                })
                continue
            stmt = all_stmts[0]
            acct = stmt.find('.//AccountInformation')
            account_id = stmt.get('accountId', 'unknown')
            raw_from_date = stmt.get('fromDate', '')
            raw_to_date = stmt.get('toDate', '')
            invalid_period_fields = [
                field for field, value in (
                    ('fromDate', raw_from_date), ('toDate', raw_to_date)
                )
                if value and not is_supported_ibkr_date(value)
            ]
            if invalid_period_fields:
                invalid_files.append({
                    'name': xml_file.name,
                    'reason': (
                        'nicht unterstütztes Datumsformat in '
                        + ', '.join(invalid_period_fields)
                    ),
                })
                continue
            entry = {
                'file': xml_file,
                'from_date': normalize_ibkr_date(raw_from_date),
                'to_date': normalize_ibkr_date(raw_to_date),
                'name': xml_file.name,
                'account_name': acct.get('name', '') if acct is not None else '',
                'currency': acct.get('currency', 'EUR') if acct is not None else 'EUR',
            }
            accounts.setdefault(account_id, []).append(entry)
        except Exception as exc:
            invalid_files.append({
                'name': xml_file.name,
                'reason': f'XML nicht lesbar: {exc}',
            })
    # Sort each account's XMLs by to_date (latest = tax year)
    for account_id in accounts:
        accounts[account_id].sort(key=lambda x: x['to_date'])
    # Detect quarterly mode: multiple XMLs for same account, all within the same year
    for account_id in accounts:
        xmls = accounts[account_id]
        if len(xmls) > 1:
            years = set()
            for x in xmls:
                if x['from_date']:
                    years.add(x['from_date'][:4])
                if x['to_date']:
                    years.add(x['to_date'][:4])
            if len(years) == 1:
                for x in xmls:
                    x['is_quarterly'] = True
    return accounts, multi_stmt_files, invalid_files

def merge_report_data(reports):
    """Merge multiple report_data dicts (one per account) by summing numeric fields."""
    if not reports:
        return {}
    if len(reports) == 1:
        return reports[0]

    merged = {}

    # Simple sum fields
    for field in ['stocks_gain_eur', 'stocks_loss_eur', 'dividends_eur', 'interest_eur',
                  'domestic_taxed_dividends_eur', 'domestic_withholding_tax_eur',
                  'debit_interest_eur', 'other_fees_eur', 'options_gain_eur', 'options_loss_eur',
                  'withholding_tax_eur', 'fx_total_gain', 'fx_total_loss', 'fx_translation',
                  'fx_correction_total', 'fx_correction_kap_inv_taxable']:
        merged[field] = sum(r.get(field, 0) for r in reports)

    # Recalculated fields
    merged['stocks_net_eur'] = merged['stocks_gain_eur'] + merged['stocks_loss_eur']
    merged['options_net_eur'] = merged['options_gain_eur'] + merged['options_loss_eur']
    merged['topf_1_aktien_netto'] = sum(r.get('topf_1_aktien_netto', 0) for r in reports)
    merged['topf_2_sonstiges_netto'] = sum(r.get('topf_2_sonstiges_netto', 0) for r in reports)
    merged['zeile_19_netto_eur'] = merged['topf_1_aktien_netto'] + merged['topf_2_sonstiges_netto']
    merged['zeile_20_stock_gains_eur'] = sum(r.get('zeile_20_stock_gains_eur', 0) for r in reports)
    merged['zeile_22_other_losses_eur'] = sum(r.get('zeile_22_other_losses_eur', 0) for r in reports)
    merged['zeile_23_stock_losses_eur'] = sum(r.get('zeile_23_stock_losses_eur', 0) for r in reports)
    merged['zeile_7_kapitalertraege_mit_inlaendischem_steuerabzug_eur'] = merged['domestic_taxed_dividends_eur']
    merged['zeile_37_kapitalertragsteuer_eur'] = sum(r.get('zeile_37_kapitalertragsteuer_eur', 0) for r in reports)
    merged['zeile_38_solidaritaetszuschlag_eur'] = sum(r.get('zeile_38_solidaritaetszuschlag_eur', 0) for r in reports)
    merged['zeile_41_withholding_tax_eur'] = sum(
        r.get('zeile_41_withholding_tax_eur', r.get('withholding_tax_eur', 0))
        for r in reports
    )

    # Scalars (from first report)
    merged['base_currency'] = reports[0].get('base_currency', 'EUR')
    merged['tax_year'] = reports[0].get('tax_year', 2025)

    # Booleans
    merged['has_trade_price'] = all(r.get('has_trade_price', False) for r in reports)
    merged['xml_has_fx_data'] = all(r.get('xml_has_fx_data', True) for r in reports)
    merged['fx_has_prior_data'] = all(r.get('fx_has_prior_data', False) for r in reports)
    merged['fx_has_negative_balance'] = any(r.get('fx_has_negative_balance', False) for r in reports)
    merged['fx_margin_correction_enabled'] = all(
        r.get('fx_margin_correction_enabled', True) for r in reports
    )
    merged['dba_wht_beta_enabled'] = all(
        r.get('dba_wht_beta_enabled', False) for r in reports
    )

    # FX source
    sources = set(r.get('fx_source', 'none') for r in reports)
    merged['fx_source'] = sources.pop() if len(sources) == 1 else 'mixed'

    # fx_results (by currency)
    merged_fx = {}
    fx_sum_keys = (
        'gain', 'loss', 'net',
        'raw_gain', 'raw_loss', 'raw_net',
        'corrected_gain', 'corrected_loss', 'corrected_net',
        'disposals_count', 'raw_disposals_count', 'corrected_disposals_count',
    )
    for r in reports:
        for curr, data in r.get('fx_results', {}).items():
            if curr not in merged_fx:
                merged_fx[curr] = {k: 0 for k in fx_sum_keys}
                merged_fx[curr].update({
                    'lots_remaining': 0,
                    'days_negative': 0,
                    'final_balance': 0,
                    'starting_balance': 0,
                })
            for k in fx_sum_keys:
                merged_fx[curr][k] += data.get(k, 0)
            merged_fx[curr]['lots_remaining'] += data.get('lots_remaining', 0)
            merged_fx[curr]['days_negative'] = max(
                merged_fx[curr].get('days_negative', 0),
                data.get('days_negative', 0)
            )
            merged_fx[curr]['final_balance'] += data.get('final_balance', 0)
            merged_fx[curr]['starting_balance'] += data.get('starting_balance', 0)
    merged['fx_results'] = merged_fx

    # Issue #59/#84 metadata for multi-account UI.
    meta_sum_keys = ('debt_repayments', 'debt_repayment_pnl')
    merged_meta = {k: 0 for k in meta_sum_keys}
    merged_meta.update({
        'has_negative_balance': merged['fx_has_negative_balance'],
        'correction_enabled': merged['fx_margin_correction_enabled'],
        'corrected_total': 0.0,
        'raw_total': 0.0,
        'csv_raw_only': False,
        'open_rows_with_pnl': [],
    })
    for r in reports:
        meta = r.get('fx_option_a_meta', {}) or {}
        for k in meta_sum_keys:
            merged_meta[k] += meta.get(k, 0)
        merged_meta['open_rows_with_pnl'].extend(meta.get('open_rows_with_pnl', []) or [])
        merged_meta['corrected_total'] += meta.get(
            'corrected_total',
            sum(d.get('corrected_net', d.get('net', 0.0)) for d in r.get('fx_results', {}).values())
        )
        merged_meta['raw_total'] += meta.get(
            'raw_total',
            sum(d.get('raw_net', d.get('net', 0.0)) for d in r.get('fx_results', {}).values())
        )
        merged_meta['csv_raw_only'] = merged_meta['csv_raw_only'] or meta.get('csv_raw_only', False)
    merged['fx_option_a_meta'] = merged_meta

    # fx_mtm
    merged_mtm = {}
    for r in reports:
        for curr, val in r.get('fx_mtm', {}).items():
            merged_mtm[curr] = merged_mtm.get(curr, 0) + (val or 0)
    merged['fx_mtm'] = merged_mtm

    # Dict sum fields
    for dict_field in ['fx_correction_by_topf', 'fx_corr_gain_adj', 'fx_corr_loss_adj']:
        merged_dict = {}
        for r in reports:
            for k, v in r.get(dict_field, {}).items():
                merged_dict[k] = merged_dict.get(k, 0) + v
        merged[dict_field] = merged_dict

    merged_kap_inv_fx = {}
    for r in reports:
        for isin, data in r.get('fx_correction_kap_inv_by_isin', {}).items():
            if isin not in merged_kap_inv_fx:
                merged_kap_inv_fx[isin] = dict(data)
            else:
                existing = merged_kap_inv_fx[isin]
                existing['raw_delta'] = existing.get('raw_delta', 0) + data.get('raw_delta', 0)
                existing['taxable_delta'] = existing.get('taxable_delta', 0) + data.get('taxable_delta', 0)
    merged['fx_correction_kap_inv_by_isin'] = merged_kap_inv_fx

    # FX correction details (list concat)
    merged['fx_correction_details'] = []
    for r in reports:
        merged['fx_correction_details'].extend(r.get('fx_correction_details', []))

    # csv_category_totals (nested dict)
    merged_csv = {}
    for r in reports:
        for cat, data in r.get('csv_category_totals', {}).items():
            if cat not in merged_csv:
                merged_csv[cat] = {}
            for k, v in data.items():
                merged_csv[cat][k] = merged_csv[cat].get(k, 0) + v
    merged['csv_category_totals'] = merged_csv

    # csv_income_totals
    merged_income = {}
    for r in reports:
        for k, v in r.get('csv_income_totals', {}).items():
            merged_income[k] = merged_income.get(k, 0) + v
    merged['csv_income_totals'] = merged_income

    # KAP-INV merge
    merged_kap = {
        'etf_gain_raw_eur': 0, 'etf_loss_raw_eur': 0,
        'etf_gain_taxable_eur': 0, 'etf_loss_taxable_eur': 0,
        'etf_dividends_raw_eur': 0, 'etf_dividends_taxable_eur': 0,
        'etf_wht_eur': 0, 'etf_wht_anrechenbar_eur': 0, 'etf_net_taxable_eur': 0,
        'etf_by_isin': {}, 'etf_unknown_isins': [],
        'wht_events': [], 'wht_review_items': [],
        'etf_stillhalter_premium_eur': 0,
    }
    for r in reports:
        ki = r.get('kap_inv', {})
        for k in ['etf_gain_raw_eur', 'etf_loss_raw_eur', 'etf_gain_taxable_eur',
                   'etf_loss_taxable_eur', 'etf_dividends_raw_eur', 'etf_dividends_taxable_eur',
                   'etf_wht_eur', 'etf_wht_anrechenbar_eur', 'etf_net_taxable_eur',
                   'etf_stillhalter_premium_eur']:
            if k != 'etf_wht_anrechenbar_eur':
                merged_kap[k] += ki.get(k, 0)
        for isin, data in ki.get('etf_by_isin', {}).items():
            if isin not in merged_kap['etf_by_isin']:
                merged_kap['etf_by_isin'][isin] = dict(data)
                merged_kap['etf_by_isin'][isin]['wht_events'] = list(
                    data.get('wht_events', [])
                )
            else:
                existing = merged_kap['etf_by_isin'][isin]
                for nk in ['gain', 'loss', 'div', 'div_received', 'div_paid',
                           'wht', 'wht_anrechenbar',
                           'gain_taxable', 'loss_taxable', 'div_taxable']:
                    existing[nk] = existing.get(nk, 0) + data.get(nk, 0)
                existing.setdefault('wht_events', []).extend(data.get('wht_events', []))
        merged_kap['wht_events'].extend(ki.get('wht_events', []))
        merged_kap['wht_review_items'].extend(ki.get('wht_review_items', []))
        for isin in ki.get('etf_unknown_isins', []):
            if isin not in merged_kap['etf_unknown_isins']:
                merged_kap['etf_unknown_isins'].append(isin)
    merged_kap['etf_wht_anrechenbar_eur'] = (
        calculate_tax_report.merge_kap_inv_wht_for_reporting(
            [r.get('kap_inv', {}) for r in reports]
        )
    )
    merged['kap_inv'] = merged_kap
    merged['kap_inv_form'] = calculate_tax_report.build_kap_inv_form(
        merged_kap['etf_by_isin'],
        merged.get('fx_correction_kap_inv_by_isin', {}),
        merged_kap['etf_unknown_isins'],
        include_tageskurs=True,
    )
    merged['kap_inv_form']['kap_line_41_creditable_tax_eur'] = merged_kap[
        'etf_wht_anrechenbar_eur'
    ]

    # no_invstg income merge (Ausschüttungen/Quellensteuer je ISIN)
    merged_no_invstg_income = {}
    for r in reports:
        for isin, data in (r.get('no_invstg_income_by_isin') or {}).items():
            if isin not in merged_no_invstg_income:
                merged_no_invstg_income[isin] = dict(data)
            else:
                existing = merged_no_invstg_income[isin]
                existing['div'] = existing.get('div', 0) + data.get('div', 0)
                existing['wht'] = existing.get('wht', 0) + data.get('wht', 0)
    merged['no_invstg_income_by_isin'] = merged_no_invstg_income

    # Auslaendische Personengesellschaften: nur beobachtete Brokerwerte
    # zusammenfuehren. Sie bleiben aus allen Steuerzeilen ausgeschlossen, bis
    # die fehlende Jahresallokation (K-1/K-3) vorliegt.
    merged_partnership_items = {}
    partnership_numeric_fields = (
        'observed_trade_pnl_eur', 'observed_distributions_eur',
        'observed_withholding_tax_eur', 'observed_other_cash_eur',
        'observed_tageskurs_delta_eur', 'observed_transactions',
    )
    for r in reports:
        for isin, data in (r.get('partnership_tax_items') or {}).items():
            if isin not in merged_partnership_items:
                merged_partnership_items[isin] = copy.deepcopy(data)
                continue
            existing = merged_partnership_items[isin]
            for field in partnership_numeric_fields:
                existing[field] = existing.get(field, 0) + data.get(field, 0)
            for field in ('required_documents', 'sources'):
                existing[field] = list(dict.fromkeys(
                    list(existing.get(field) or []) + list(data.get(field) or [])
                ))
    merged['partnership_tax_items'] = merged_partnership_items

    # Anlage SO merge
    merged_so = {
        'total_gain': 0, 'total_loss': 0,
        'taxable_gain': 0, 'taxable_loss': 0,
        'tax_free_gain': 0, 'tax_free_loss': 0,
        'unknown_gain': 0, 'unknown_loss': 0,
        'details': [], 'by_isin': {},
    }
    for r in reports:
        so = r.get('anlage_so', {})
        for k in ['total_gain', 'total_loss', 'taxable_gain', 'taxable_loss',
                   'tax_free_gain', 'tax_free_loss', 'unknown_gain', 'unknown_loss']:
            merged_so[k] += so.get(k, 0)
        merged_so['details'].extend(so.get('details', []))
        for isin, data in so.get('by_isin', {}).items():
            if isin not in merged_so['by_isin']:
                merged_so['by_isin'][isin] = dict(data)
            else:
                existing = merged_so['by_isin'][isin]
                for nk in ['taxable', 'tax_free', 'total']:
                    existing[nk] = existing.get(nk, 0) + data.get(nk, 0)
    merged['anlage_so'] = merged_so

    # all_traded_etf_isins: Union aller Accounts (Issue #51 - für Override-UI)
    merged_all_etfs = set()
    for r in reports:
        merged_all_etfs.update(r.get('all_traded_etf_isins', []))
    merged['all_traded_etf_isins'] = sorted(merged_all_etfs)
    merged['anlage_so_overrides_applied'] = list(reports[0].get('anlage_so_overrides_applied', []))
    merged_review_items = {}
    for r in reports:
        for item in r.get('classification_review_items', []) or []:
            isin = item.get('isin', '')
            if isin:
                merged_review_items[isin] = dict(item)
    merged['classification_review_items'] = list(merged_review_items.values())

    # Audit merge
    merged_audit = {
        'funds_processed': sum(r.get('audit', {}).get('funds_processed', 0) for r in reports),
        'funds_skipped_year': sum(r.get('audit', {}).get('funds_skipped_year', 0) for r in reports),
        'raw_div_base': sum(r.get('audit', {}).get('raw_div_base', 0) for r in reports),
        'raw_tax_base': sum(r.get('audit', {}).get('raw_tax_base', 0) for r in reports),
        'added_from_summary': sum(r.get('audit', {}).get('added_from_summary', 0) for r in reports),
        'usd_to_eur_rates_count': max(r.get('audit', {}).get('usd_to_eur_rates_count', 0) for r in reports),
        'ecb_rates_used': any(r.get('audit', {}).get('ecb_rates_used', False) for r in reports),
        'stillhalter_count': sum(r.get('audit', {}).get('stillhalter_count', 0) for r in reports),
        'stillhalter_premium_eur': sum(r.get('audit', {}).get('stillhalter_premium_eur', 0) for r in reports),
        'stillhalter_unmatched': [],
        'stillhalter_details': [],
        'cross_year_premium_eur': sum(r.get('audit', {}).get('cross_year_premium_eur', 0) for r in reports),
        'cross_year_by_year': {},
        'cross_year_put_corrections': [],
        'invstg_put_basis_adjustments': [],
        'cross_year_put_total': sum(r.get('audit', {}).get('cross_year_put_total', 0) for r in reports),
        'no_invstg_gain': sum(r.get('audit', {}).get('no_invstg_gain', 0) for r in reports),
        'no_invstg_loss': sum(r.get('audit', {}).get('no_invstg_loss', 0) for r in reports),
        'partnership_tax_items_count': len(merged_partnership_items),
        'zufluss_premium_eur': sum(r.get('audit', {}).get('zufluss_premium_eur', 0) for r in reports),
        'zufluss_count': sum(r.get('audit', {}).get('zufluss_count', 0) for r in reports),
        'zufluss_details': [],
        'prior_zufluss_correction_eur': sum(r.get('audit', {}).get('prior_zufluss_correction_eur', 0) for r in reports),
        'prior_zufluss_details': [],
        'zufluss_unmatched': [],
        'occ_rename_matches': [],
        'stillhalter_corrections_dropped': [],
        'future_assignment_corrections': [],
        'stillhalter_open_short': [],
        'stk_correction_cy': sum(r.get('audit', {}).get('stk_correction_cy', 0) for r in reports),
        'etf_correction_cy': sum(r.get('audit', {}).get('etf_correction_cy', 0) for r in reports),
        'put_nosell_premium_eur': sum(r.get('audit', {}).get('put_nosell_premium_eur', 0) for r in reports),
        'underlying_symbol_aliases': {},
        'unrouted_asset_categories': [],
        'cfd_interest_income_eur': sum(r.get('audit', {}).get('cfd_interest_income_eur', 0) for r in reports),
        'cfd_financing_cost_eur': sum(r.get('audit', {}).get('cfd_financing_cost_eur', 0) for r in reports),
        'fee_by_activity_code': {},
        'transaction_tax': {
            key: sum(
                r.get('audit', {}).get('transaction_tax', {}).get(key, 0)
                for r in reports
            )
            for key in (
                'found_count', 'applied_count', 'applied_eur',
                'deferred_count', 'deferred_eur',
                'already_in_trade_count', 'historical_count',
                'unmatched_count',
            )
        },
        'unhandled_activity_codes': [],
        'fx_rate_parse_failures': {
            'funds': sum(r.get('audit', {}).get('fx_rate_parse_failures', {}).get('funds', 0) for r in reports),
            'trades': sum(r.get('audit', {}).get('fx_rate_parse_failures', {}).get('trades', 0) for r in reports),
        },
    }
    _fees_by_code = {}
    _unhandled_by_code = {}
    for r in reports:
        for _c, _v in (r.get('audit', {}).get('fee_by_activity_code') or {}).items():
            _fees_by_code[_c] = _fees_by_code.get(_c, 0.0) + _v
        for _e in r.get('audit', {}).get('unhandled_activity_codes', []):
            _code = _e.get('code', '(leer)')
            _agg = _unhandled_by_code.setdefault(
                _code, {'code': _code, 'count': 0, 'amount_eur': 0.0, 'descriptions': []}
            )
            _agg['count'] += _e.get('count', 0)
            _agg['amount_eur'] += _e.get('amount_eur', 0.0)
            for _d in _e.get('descriptions', []):
                if _d not in _agg['descriptions'] and len(_agg['descriptions']) < 3:
                    _agg['descriptions'].append(_d)
    merged_audit['fee_by_activity_code'] = dict(sorted(_fees_by_code.items()))
    merged_audit['unhandled_activity_codes'] = sorted(
        _unhandled_by_code.values(), key=lambda e: e['code']
    )
    _unrouted_by_cat = {}
    for r in reports:
        for _e in r.get('audit', {}).get('unrouted_asset_categories', []):
            _cat = _e.get('category', '(leer)')
            _agg = _unrouted_by_cat.setdefault(
                _cat, {'category': _cat, 'count': 0, 'pnl_eur': 0.0, 'symbols': [], 'sources': []}
            )
            _agg['count'] += _e.get('count', 0)
            _agg['pnl_eur'] += _e.get('pnl_eur', 0.0)
            for _s in _e.get('symbols', []):
                if _s not in _agg['symbols']:
                    _agg['symbols'].append(_s)
            for _src in _e.get('sources', []):
                if _src not in _agg['sources']:
                    _agg['sources'].append(_src)
    merged_audit['unrouted_asset_categories'] = sorted(
        _unrouted_by_cat.values(), key=lambda e: e['category']
    )
    for r in reports:
        a = r.get('audit', {})
        for _canon, _members in a.get('underlying_symbol_aliases', {}).items():
            _existing = merged_audit['underlying_symbol_aliases'].setdefault(_canon, [])
            for _m in _members:
                if _m not in _existing:
                    _existing.append(_m)
        merged_audit['stillhalter_unmatched'].extend(a.get('stillhalter_unmatched', []))
        merged_audit['stillhalter_details'].extend(a.get('stillhalter_details', []))
        merged_audit['cross_year_put_corrections'].extend(a.get('cross_year_put_corrections', []))
        merged_audit['invstg_put_basis_adjustments'].extend(
            a.get('invstg_put_basis_adjustments', [])
        )
        merged_audit['zufluss_details'].extend(a.get('zufluss_details', []))
        merged_audit['prior_zufluss_details'].extend(a.get('prior_zufluss_details', []))
        merged_audit['zufluss_unmatched'].extend(a.get('zufluss_unmatched', []))
        merged_audit['occ_rename_matches'].extend(a.get('occ_rename_matches', []))
        merged_audit['stillhalter_corrections_dropped'].extend(a.get('stillhalter_corrections_dropped', []))
        merged_audit['future_assignment_corrections'].extend(
            a.get('future_assignment_corrections', []))
        merged_audit['stillhalter_open_short'].extend(a.get('stillhalter_open_short', []))
        merged_audit['transaction_tax'].setdefault('details', []).extend(
            a.get('transaction_tax', {}).get('details', [])
        )
        for year, val in a.get('cross_year_by_year', {}).items():
            merged_audit['cross_year_by_year'][year] = merged_audit['cross_year_by_year'].get(year, 0) + val
    merged['audit'] = merged_audit

    # Trade details merge (concatenate lists from all accounts)
    merged['trade_details'] = []
    for r in reports:
        merged['trade_details'].extend(r.get('trade_details', []))
    merged['trade_details'].sort(key=lambda r: r.get('dateTime', '') or r.get('reportDate', '') or 'zzzz')

    # Aufschluesselung Topf 2 (Label -> gain/loss): per Kategorie summieren.
    # Ohne diese Regel fiel das Feld beim Multi-Account-Merge still heraus und
    # der Expander "Aufschluesselung Topf 2" verschwand komplett.
    merged_topf2_cats = {}
    for r in reports:
        for label, vals in (r.get('topf2_by_category') or {}).items():
            entry = merged_topf2_cats.setdefault(label, {'gain': 0.0, 'loss': 0.0})
            entry['gain'] += vals.get('gain', 0.0)
            entry['loss'] += vals.get('loss', 0.0)
    merged['topf2_by_category'] = merged_topf2_cats

    return merged

def render_classification_catalog(
        catalog_rows, key_prefix, show_filters=True, offer_download=False):
    """Render the shared product-classification transparency view."""
    catalog_rows = list(catalog_rows)
    filtered_rows = catalog_rows

    if show_filters:
        search = st.text_input(
            "Ticker, Name oder ISIN suchen",
            key=f"{key_prefix}_search",
            placeholder="z. B. GLD, JEPI oder US78463V1070",
        ).strip().casefold()
        classifications = sorted({
            row['classification_label'] for row in catalog_rows
        })
        evidence_labels = sorted({
            row['evidence_label'] for row in catalog_rows
        })
        filter_col_1, filter_col_2 = st.columns(2)
        selected_classifications = filter_col_1.multiselect(
            "Zuordnung",
            classifications,
            key=f"{key_prefix}_classification",
            placeholder="Alle",
        )
        selected_evidence = filter_col_2.multiselect(
            "Nachweisstatus",
            evidence_labels,
            key=f"{key_prefix}_evidence",
            placeholder="Alle",
        )

        if search:
            filtered_rows = [
                row for row in filtered_rows
                if search in ' '.join((
                    row['ticker'], row['name'], row['isin'],
                    row['classification_label'], row['decision_reason'],
                )).casefold()
            ]
        if selected_classifications:
            filtered_rows = [
                row for row in filtered_rows
                if row['classification_label'] in selected_classifications
            ]
        if selected_evidence:
            filtered_rows = [
                row for row in filtered_rows
                if row['evidence_label'] in selected_evidence
            ]

    status_counts = {}
    for row in catalog_rows:
        label = row['evidence_label']
        status_counts[label] = status_counts.get(label, 0) + 1
    status_summary = ' · '.join(
        f"{count}× {label}"
        for label, count in sorted(status_counts.items())
    )
    st.caption(
        f"{len(filtered_rows)} von {len(catalog_rows)} Zuordnungen · "
        f"{status_summary}"
    )
    st.caption(
        "„Katalogzuordnung · aktiv“ ist eine fest angewandte Zuordnung und "
        "kein offener Prüffall. „Produktindividuell geprüft“ bedeutet "
        "zusätzlich, dass für dieses Produkt ein eigener Primärbeleg im "
        "Katalog verknüpft ist."
    )
    if not filtered_rows:
        st.info("Für die gewählten Filter wurden keine Produkte gefunden.")
        return

    table_rows = [{
        'Ticker': row['ticker'],
        'Name': row['name'],
        'ISIN': row['isin'],
        'Zuordnung': row['classification_label'],
        'TFS': row['tfs_label'],
        'Steuerpfad': row['tax_route'],
        'Nachweis': row['evidence_label'],
        'Kurzbegründung': row['decision_reason'],
    } for row in filtered_rows]
    st.dataframe(table_rows, use_container_width=True, hide_index=True)

    row_by_isin = {row['isin']: row for row in filtered_rows}
    selected_isin = st.selectbox(
        "Begründung und Quellen im Detail",
        options=list(row_by_isin),
        format_func=lambda isin: (
            f"{row_by_isin[isin]['ticker']} · "
            f"{row_by_isin[isin]['name'] or 'Unbekanntes Produkt'} · {isin}"
        ),
        key=f"{key_prefix}_detail",
    )
    selected = row_by_isin[selected_isin]
    st.write(
        f"{selected['ticker']} · "
        f"{selected['name'] or 'Produktname nicht bekannt'} ({selected['isin']})"
    )
    st.caption(
        f"{selected['classification_label']} · TFS {selected['tfs_label']} · "
        f"{selected['tax_route']} · Stand {selected['as_of']}"
    )
    st.markdown("**Begründung**")
    st.write(selected['decision_reason'])
    st.markdown("**Rechtsform und Rechtsgrundlage**")
    st.write(f"{selected['legal_form']} · {selected['legal_basis']}")

    if selected['product_sources']:
        product_links = ' · '.join(
            f"[{label}]({url})"
            for label, url in selected['product_sources']
        )
        st.markdown(f"**Produktbeleg:** {product_links}")
    else:
        st.caption(
            "Feste, aktiv berechnete Katalogzuordnung – kein Prüffall und keine "
            "Quarantäne. Für dieses Produkt ist nur kein eigener "
            "produktindividueller Primärbeleg verknüpft."
        )
    if selected['legal_sources']:
        legal_links = ' · '.join(
            f"[{label}]({url})" for label, url in selected['legal_sources']
        )
        st.markdown(f"**Rechtsquellen:** {legal_links}")

    if offer_download:
        st.download_button(
            "Gesamtkatalog als CSV herunterladen",
            data=classification_catalog_to_csv(catalog_rows).encode('utf-8-sig'),
            file_name="etf_fonds_klassifikationen.csv",
            mime="text/csv",
            key=f"{key_prefix}_download",
        )



# ── Upload-Snapshot ──────────────────────────────────────────────────────────

class _MemUpload:
    """Minimales File-Objekt ueber gesnapshotteten Bytes (classify_xmls und
    die Extraktion arbeiten nur mit .name/.getvalue()/.getbuffer())."""

    def __init__(self, name, data):
        self.name = name
        self._data = data

    def getvalue(self):
        return self._data

    def getbuffer(self):
        return self._data


def _upload_widget_keys(scope):
    """Separate widget identities for start screen and persistent sidebar.

    Streamlit unmounts the start-screen uploaders once XML data exists.  If
    the sidebar reused those keys, the newly mounted empty widgets could be
    mistaken for an explicit deletion on the next interaction.
    """
    if scope not in {'start', 'sidebar'}:
        raise ValueError(f"Unbekannter Upload-Bereich: {scope}")
    epoch = st.session_state.get('upload_epoch', 0)
    return (
        f"_ui_{scope}_xml_uploads_{epoch}",
        f"_ui_{scope}_csv_upload_{epoch}",
    )


def _capture_xml_uploads(widget_key, append_existing=False):
    """XML-Callback: Start ersetzt, Sidebar ergänzt den XML-Datensatz."""
    raw_entries = []
    for up in st.session_state.get(widget_key) or []:
        data = bytes(up.getvalue())
        raw_entries.append({
            'name': up.name,
            'digest': ui_model.file_digest(data),
            'kind': 'xml',
            'data': data,
        })
    existing = st.session_state.get('dataset')
    if append_existing:
        st.session_state['dataset'] = ui_model.append_xml_uploads(
            existing, raw_entries,
        )
    else:
        st.session_state['dataset'] = ui_model.update_upload_dataset(
            existing, xml_entries=raw_entries,
        )


def _capture_csv_upload(widget_key):
    """CSV-Callback: nur den CSV-Teil aktualisieren, XML unverändert lassen."""
    csv_entry = None
    csv_up = st.session_state.get(widget_key)
    if csv_up is not None:
        data = bytes(csv_up.getvalue())
        csv_entry = {
            'name': csv_up.name,
            'digest': ui_model.file_digest(data),
            'kind': 'csv',
            'data': data,
        }
    st.session_state['dataset'] = ui_model.update_upload_dataset(
        st.session_state.get('dataset'), csv_entry=csv_entry,
    )


def _reset_all_data():
    """Expliziter Reset: neuer Widget-Key (upload_epoch), Domain-State und
    Snapshot verwerfen."""
    st.session_state['upload_epoch'] = st.session_state.get('upload_epoch', 0) + 1
    for key in ('dataset', 'snapshot', 'export_cache', 'domain', 'nav'):
        st.session_state.pop(key, None)


def _render_uploaders(scope):
    xml_key, csv_key = _upload_widget_keys(scope)
    append_existing = scope == 'sidebar'
    st.file_uploader(
        ("Weitere Flex Query XMLs hinzufügen" if append_existing
         else "Flex Query XMLs hochladen"),
        type="xml",
        accept_multiple_files=True,
        key=xml_key,
        on_change=_capture_xml_uploads,
        args=(xml_key, append_existing),
        help=(
            (("Die ausgewählten XMLs werden zum bestehenden Datensatz "
              "hinzugefügt. Für einen vollständigen Austausch zuerst "
              "'Alle Daten entfernen' verwenden. ") if append_existing else "")
            + "Steuerjahr-XML hochladen; bei jahresübergreifenden Optionen "
              "zusätzlich die Vorjahres-XMLs. Mehrere Konten werden separat "
              "berechnet und addiert. Quartals-XMLs desselben Jahres werden "
              "automatisch zusammengeführt."
        ),
    )
    if scope == 'start':
        with st.expander("Flex-Query-XML in IBKR erstellen", expanded=False):
            st.caption(
                "IBKR → Performance & Berichte → Flex-Abfragen → XML "
                "exportieren. Bei jahresübergreifenden Optionen zusätzlich "
                "die Vorjahres-XMLs gemeinsam auswählen."
            )
        csv_context = st.expander(
            "Erweiterte Prüfung mit IBKR-CSV · optional", expanded=False,
        )
    else:
        csv_context = contextlib.nullcontext()

    with csv_context:
        if scope == 'start':
            st.caption(
                "Ergänzt einen Plausibilitätsvergleich und kann als "
                "FX-Fallback dienen. Für den Steuerbericht ist die CSV nicht "
                "erforderlich."
            )
        st.file_uploader(
            ("IBKR-Standardbericht (CSV)" if scope == 'start'
             else "Optional: IBKR-Standardbericht (CSV)"),
            type="csv",
            key=csv_key,
            on_change=_capture_csv_upload,
            args=(csv_key,),
            help=(
                "Plausibilitätscheck gegen IBKRs eigene Kategoriesummen und "
                "FX-Fallback für EUR-Basiskonten ohne FxTransactions-Sektion. "
                "Erstellen: IBKR → Performance & Berichte → Kontoauszüge → "
                "Übersicht: realisierter G&V → CSV. Nur bei einem einzelnen "
                "Konto aktiv."
            ),
        )


# ── Domain-State (fachliche Eingaben, per dataset_id genamespaced) ───────────

def _domain(dataset_id):
    domains = st.session_state.setdefault('domain', {})
    return domains.setdefault(dataset_id, {
        'toggles': ui_model.default_toggles(),
        'etf_overrides': {},
        'anlage_so_overrides': [],
    })


def _bind_toggle(dom, toggle_key, widget_key, label, help_text=None,
                 impact_text=None):
    """Checkbox mit _ui_*-Widget-Key, Domain-State als Quelle der Wahrheit."""

    def _sync():
        dom['toggles'][toggle_key] = bool(st.session_state[widget_key])

    st.checkbox(
        label,
        value=bool(dom['toggles'].get(toggle_key)),
        key=widget_key,
        on_change=_sync,
        help=help_text,
    )
    if impact_text:
        st.caption(f"**Auswirkung:** {impact_text}")


# ── Start-Zustand ────────────────────────────────────────────────────────────

_dataset = st.session_state.get('dataset') or {}
_dataset_files = _dataset.get('files') or []

if not _dataset_files:
    intro_col, upload_col = st.columns([0.94, 1.06], gap="large")
    with intro_col:
        st.markdown(
            '<div class="start-title">IBKR-Steuerbericht. Vom Flex Query '
            'zur Anlage KAP.</div>'
            '<div class="start-sub">Flex Query hochladen, automatisch prüfen '
            'lassen und die berechneten KAP-, KAP-INV- und SO-Zeilen '
            'übernehmen. Die Verarbeitung erfolgt vollständig lokal auf dem '
            'eigenen Rechner.</div>'
            '<div class="start-trust">'
            '<span class="start-trust-chip">✓ 100 % lokal</span>'
            '<span class="start-trust-chip">✓ Keine Datenübertragung</span>'
            '<span class="start-trust-chip">✓ Open Source</span>'
            '</div>'
            '<div class="start-visual">'
            '<svg viewBox="0 0 520 186" role="img" '
            'aria-label="Ablauf: Flex Query laden, automatisch prüfen, '
            'Formularzeilen übernehmen">'
            # Station 1: Flex-Query-Dokument
            '<rect class="sv-card" x="18" y="24" width="78" height="100" rx="10"/>'
            '<text class="sv-glyph" x="57" y="56" text-anchor="middle">&lt;XML/&gt;</text>'
            '<rect class="sv-code" x="32" y="72" width="48" height="4" rx="2"/>'
            '<rect class="sv-code" x="32" y="84" width="34" height="4" rx="2"/>'
            '<rect class="sv-code" x="32" y="96" width="42" height="4" rx="2"/>'
            # Flusslinie 1
            '<line class="sv-flow" x1="102" y1="74" x2="186" y2="74"/>'
            '<path class="sv-arrow" d="M196 74 l-8 -4.5 v9 z"/>'
            # Station 2: Pruefung
            '<circle class="sv-card" cx="240" cy="74" r="32"/>'
            '<path class="sv-check" d="M227 75 l9 9 l18 -20"/>'
            # Flusslinie 2
            '<line class="sv-flow" x1="278" y1="74" x2="342" y2="74"/>'
            '<path class="sv-arrow" d="M352 74 l-8 -4.5 v9 z"/>'
            # Station 3: Mini-Formular mit Zeilen-Badges und Fuehrungslinien
            '<rect class="sv-card" x="358" y="18" width="148" height="112" rx="10"/>'
            '<rect class="sv-badge fill" x="370" y="32" width="26" height="16" rx="5"/>'
            '<text class="sv-badge-text fill" x="383" y="44" text-anchor="middle">19</text>'
            '<line class="sv-leader" x1="402" y1="40" x2="450" y2="40"/>'
            '<text class="sv-value" x="494" y="44" text-anchor="end">1.234,56</text>'
            '<rect class="sv-badge" x="370" y="62" width="26" height="16" rx="5"/>'
            '<text class="sv-badge-text" x="383" y="74" text-anchor="middle">20</text>'
            '<line class="sv-leader" x1="402" y1="70" x2="458" y2="70"/>'
            '<text class="sv-value" x="494" y="74" text-anchor="end">876,00</text>'
            '<rect class="sv-badge" x="370" y="92" width="26" height="16" rx="5"/>'
            '<text class="sv-badge-text" x="383" y="104" text-anchor="middle">41</text>'
            '<line class="sv-leader" x1="402" y1="100" x2="458" y2="100"/>'
            '<text class="sv-value" x="494" y="104" text-anchor="end">312,40</text>'
            # Beschriftungen
            '<text class="sv-label" x="57" y="162" text-anchor="middle">'
            '<tspan>1</tspan> · Flex Query</text>'
            '<text class="sv-label" x="240" y="162" text-anchor="middle">'
            '<tspan>2</tspan> · Automatische Prüfung</text>'
            '<text class="sv-label" x="432" y="162" text-anchor="middle">'
            '<tspan>3</tspan> · Anlage KAP</text>'
            '</svg>'
            '</div>',
            unsafe_allow_html=True,
        )
    with upload_col:
        with st.container(key="start_upload_card"):
            st.markdown(
                '<div class="start-upload-kicker">Schritt 1</div>'
                '<div class="start-upload-title">Flex Query hochladen</div>'
                '<div class="start-upload-copy">Erforderlich für die '
                'Berechnung. Mehrere XMLs können gemeinsam ausgewählt oder '
                'hierher gezogen werden.</div>',
                unsafe_allow_html=True,
            )
            _render_uploaders('start')
    st.markdown(
        '<div class="start-foot"><strong>Keine Steuerberatung · '
        'Haftungsbeschränkung:</strong> '
        'Nutzung und Prüfung der Ergebnisse erfolgen eigenverantwortlich. '
        'Soweit gesetzlich zulässig, wird für Schäden aus der Nutzung sowie '
        'für fehlerhafte oder unvollständige Berechnungen keine Haftung '
        'übernommen.</div>',
        unsafe_allow_html=True,
    )
    st.stop()

_dataset_id = _dataset.get('dataset_id', '')
_dom = _domain(_dataset_id)

# ── Sidebar-Skelett ──────────────────────────────────────────────────────────
# Reihenfolge: Marke, Datengrundlage, Navigation, Berechnungsoptionen.
# Nav und Optionen brauchen das View-Model und werden nachtraeglich in ihre
# Container gefuellt (Streamlit erlaubt Out-of-order-Rendering).

with st.sidebar:
    st.markdown('<div class="sidebar-brand">IBKR Steuerbericht</div>',
                unsafe_allow_html=True)
    st.markdown(
        '<div class="sidebar-brand-sub">Anlage KAP · Flex Query · lokal</div>',
        unsafe_allow_html=True,
    )
    _data_card_slot = st.container()
    with st.expander("Daten ändern", expanded=False):
        _render_uploaders('sidebar')
        st.button(
            "Alle Daten entfernen",
            key="_ui_reset_all",
            on_click=_reset_all_data,
            use_container_width=True,
        )
    _nav_slot = st.container()
    st.markdown('<div class="eyebrow" style="margin-top:1.1rem;">Berechnung</div>',
                unsafe_allow_html=True)
    _toggle_slot = st.container()

# ── Compute-Snapshot ─────────────────────────────────────────────────────────

_csv_entry = _dataset.get('csv')
_requested_key = ui_model.build_input_key(
    _dataset_id,
    _csv_entry['digest'] if _csv_entry else None,
    _dom['toggles'].get('fx_margin', True),
    _dom['toggles'].get('dba_beta', False),
    _dom.get('anlage_so_overrides', []),
)


class UploadValidationError(Exception):
    pass


def _run_compute(dataset, csv_entry, dom, requested_key, generation):
    """Extraktion + calculate_tax pro Konto. Laeuft mit umgeleitetem
    stdout/stderr (Core-prints enthalten Kontonummern und Betraege) unter dem
    prozessweiten Lock aus ui_model; gespeichert wird nur eine
    Allowlist-Zusammenfassung."""
    started = time.time()
    logbuf = io.StringIO()
    mem_files = [_MemUpload(f['name'], f['data']) for f in dataset['files']]

    with ui_model.REDIRECT_LOCK:
        with contextlib.redirect_stdout(logbuf), \
                contextlib.redirect_stderr(logbuf):
            accounts, multi_stmt_files, invalid_files = classify_xmls(mem_files)
            if invalid_files:
                details = "; ".join(
                    f"{item['name']}: {item['reason']}"
                    for item in invalid_files
                )
                raise UploadValidationError(
                    "Ungültige Upload-Datei(en). Jede ausgewählte XML muss "
                    f"ein IBKR-Flex-Query-Export sein: {details}."
                )
            if multi_stmt_files:
                details = "; ".join(
                    f"{item['name']} ({', '.join(item['account_ids'])})"
                    for item in multi_stmt_files
                )
                raise UploadValidationError(
                    "Mehrere Konten innerhalb derselben XML werden nicht "
                    f"teilweise verarbeitet: {details}. Bitte pro Konto eine "
                    "eigene Flex Query exportieren und alle Dateien gemeinsam "
                    "hochladen."
                )
            if not accounts:
                raise UploadValidationError(
                    "Keine der Dateien enthält ein lesbares FlexStatement."
                )

            overlaps = ui_model.find_period_overlaps(accounts)
            if overlaps:
                lines = "; ".join(
                    f"Konto {o['account_id']} ({o['year']}): "
                    f"{o['files'][0]} und {o['files'][1]} überschneiden sich "
                    f"({o['periods'][0]} / {o['periods'][1]})"
                    for o in overlaps
                )
                raise UploadValidationError(
                    "Überlappende Berichtszeiträume desselben Kontos im "
                    f"selben Jahr: {lines}. Der Quartalsmodus akzeptiert nur "
                    "disjunkte Zeiträume, sonst würden Ergebnisse doppelt "
                    "gezählt. Bitte die Exporte korrigieren."
                )

            all_to_dates = [x['to_date'] for xs in accounts.values() for x in xs]
            global_tax_year = max(all_to_dates)[:4] if all_to_dates else '2025'

            accounts_to_process = {}
            accounts_skipped = []
            for acct_id, xmls in accounts.items():
                if xmls[-1]['to_date'][:4] == global_tax_year:
                    accounts_to_process[acct_id] = xmls
                else:
                    label = xmls[-1]['account_name'] or acct_id
                    accounts_skipped.append(
                        f"{label} ({acct_id}, nur bis {xmls[-1]['to_date'][:4]})"
                    )
            if not accounts_to_process:
                raise UploadValidationError(
                    "Keine XML für das Steuerjahr gefunden."
                )

            if len(accounts_to_process) > 1:
                currencies = {
                    xs[-1]['currency'] for xs in accounts_to_process.values()
                }
                if len(currencies) > 1:
                    raise UploadValidationError(
                        "Unterschiedliche Basiswährungen erkannt: "
                        f"{', '.join(sorted(currencies))}. Alle Konten müssen "
                        "dieselbe Basiswährung haben."
                    )

            csv_enabled = (
                csv_entry is not None and len(accounts_to_process) == 1
            )

            reports = []
            account_names = []
            quarterly_infos = []
            for acct_id, xmls in sorted(accounts_to_process.items()):
                main_xml = xmls[-1]
                history_xmls = xmls[:-1]
                is_quarterly = (
                    all(x.get('is_quarterly') for x in xmls) and len(xmls) > 1
                )
                acct_label = main_xml['account_name'] or acct_id
                if is_quarterly:
                    periods = [
                        f"{x['from_date'][5:7]}-{x['to_date'][5:7]}"
                        for x in xmls if x['from_date'] and x['to_date']
                    ]
                    quarterly_infos.append({
                        'account': acct_label,
                        'count': len(xmls),
                        'periods': periods,
                    })

                with tempfile.TemporaryDirectory() as tmp:
                    csv_report_path = None
                    if csv_enabled:
                        csv_report_path = os.path.join(tmp, "ibkr_report.csv")
                        with open(csv_report_path, "wb") as fh:
                            fh.write(csv_entry['data'])

                    if is_quarterly:
                        xml_paths = []
                        for i, qxml in enumerate(xmls):
                            qp = os.path.join(tmp, f"quarter_{i}.xml")
                            with open(qp, "wb") as fh:
                                fh.write(qxml['file'].getbuffer())
                            xml_paths.append(qp)
                        extract_ibkr_data.extract_quarterly_xmls(xml_paths, tmp)
                    else:
                        xml_path = os.path.join(tmp, "input.xml")
                        with open(xml_path, "wb") as fh:
                            fh.write(main_xml['file'].getbuffer())
                        history_paths = []
                        for i, hxml in enumerate(history_xmls):
                            hp = os.path.join(tmp, f"history_{i}.xml")
                            with open(hp, "wb") as fh:
                                fh.write(hxml['file'].getbuffer())
                            history_paths.append(hp)
                        if history_paths:
                            extract_ibkr_data.extract_fx_multi_xml(
                                sorted(history_paths) + [xml_path], tmp
                            )
                        else:
                            extract_ibkr_data.parse_ibkr_xml(xml_path, tmp)

                    d_acct = calculate_tax_report.calculate_tax(
                        tmp,
                        fx_csv_path=csv_report_path,
                        anlage_so_overrides=dom.get('anlage_so_overrides', []),
                        fx_margin_correction_enabled=dom['toggles'].get(
                            'fx_margin', True),
                        dba_wht_beta_enabled=dom['toggles'].get(
                            'dba_beta', False),
                    )
                    if reports and d_acct.get('base_currency') != \
                            reports[0].get('base_currency'):
                        raise UploadValidationError(
                            f"Konto {acct_label} hat Basiswährung "
                            f"{d_acct.get('base_currency')}, erwartet "
                            f"{reports[0].get('base_currency')}."
                        )
                    reports.append(d_acct)
                    account_names.append(acct_label)

            per_account_wht_pools = [
                copy.deepcopy(r.get('kap_inv', {}).get('etf_by_isin', {}) or {})
                for r in reports
            ]
            merged = merge_report_data(reports)
            if len(reports) == 1:
                # merge_report_data gibt bei Einzelkonto das Original zurueck —
                # die Snapshot-Grenze kapselt defensiv per Deep-Copy, damit
                # kein UI-Pfad den eingefrorenen Report mutieren kann.
                merged = copy.deepcopy(merged)

    return {
        'input_key': requested_key,
        'schema_version': ui_model.SCHEMA_VERSION,
        'generation': generation,
        'status': 'ok',
        'computed_at': _dt.now().strftime('%d.%m.%Y %H:%M'),
        'duration_s': round(time.time() - started, 2),
        'suppressed_log_lines': logbuf.getvalue().count('\n'),
        'payload': {
            'reports': reports,
            'account_names': account_names,
            'merged': merged,
            'per_account_wht_pools': per_account_wht_pools,
            'accounts_skipped': accounts_skipped,
            'multi_stmt_files': multi_stmt_files,
            'quarterly_infos': quarterly_infos,
            'csv_enabled': csv_enabled,
            'csv_present': csv_entry is not None,
            'n_accounts': len(reports),
        },
    }


_snapshot = st.session_state.get('snapshot')
_cache_hit = ui_model.snapshot_is_current(_snapshot, _requested_key)
if not _cache_hit:
    _generation = st.session_state.get('compute_generation', 0) + 1
    st.session_state['compute_generation'] = _generation
    try:
        with st.spinner("Berechne Steuerbericht…"):
            _computed = _run_compute(
                _dataset, _csv_entry, _dom, _requested_key, _generation,
            )
    except (UploadValidationError, calculate_tax_report.FxCurrencyError) as exc:
        st.markdown(notice_html({
            'class': 'fehler', 'severity': 'kritisch',
            'title': 'Berechnung nicht möglich', 'body': str(exc),
            'target': None,
        }), unsafe_allow_html=True)
        st.stop()
    except Exception as exc:  # noqa: BLE001 — Fehler sichtbar machen
        st.markdown(notice_html({
            'class': 'fehler', 'severity': 'kritisch',
            'title': 'Fehler bei der Verarbeitung', 'body': str(exc),
            'target': None,
        }), unsafe_allow_html=True)
        st.exception(exc)
        st.stop()

    # Atomarer Commit: ein alter Lauf darf keinen inzwischen neu
    # angeforderten Key ueberschreiben (fastReruns).
    _current_generation = st.session_state.get('compute_generation')
    if ui_model.should_commit_snapshot(
            _requested_key, _current_generation,
            _computed['input_key'], _computed['generation']):
        st.session_state['snapshot'] = _computed
        _snapshot = _computed
    else:
        st.stop()

_payload = _snapshot['payload']
d = _payload['merged']
reports = _payload['reports']
account_names = _payload['account_names']
n_accounts = _payload['n_accounts']
per_account_wht_pools = _payload['per_account_wht_pools']
steuerjahr = d.get('tax_year', 2025)
created_at = _snapshot['computed_at']

# ── View-Model ───────────────────────────────────────────────────────────────

_availability = ui_model.toggle_availability(d)
_eff_toggles = ui_model.effective_toggles(_dom['toggles'], _availability)


def build_plausibility(d, toggles):
    """Plausibilitaetscheck gegen IBKRs eigene CSV-Kategoriesummen (rein)."""
    csv_cats = d.get('csv_category_totals', {})
    if not csv_cats:
        return None
    audit = d.get('audit', {}) or {}
    kap_inv_data = d.get('kap_inv', {}) or {}
    anlage_so = d.get('anlage_so', {}) or {}
    has_etf = bool(kap_inv_data.get('etf_by_isin'))
    has_so = bool(anlage_so.get('by_isin'))
    invstg_on = toggles.get('invstg', False)

    cross_put = audit.get('cross_year_put_total', 0)
    no_invstg_gain = audit.get('no_invstg_gain', 0)
    no_invstg_loss = audit.get('no_invstg_loss', 0)
    stillhalter_addback = (audit.get('stk_correction_cy', 0)
                           + audit.get('etf_correction_cy', 0))
    our_stk_gain = (d.get('stocks_gain_eur', 0) + stillhalter_addback
                    + cross_put + no_invstg_gain)
    our_stk_loss = d.get('stocks_loss_eur', 0) + no_invstg_loss
    if has_etf and invstg_on:
        our_stk_gain += kap_inv_data.get('etf_gain_raw_eur', 0)
        our_stk_loss += kap_inv_data.get('etf_loss_raw_eur', 0)
    if has_so:
        our_stk_gain += anlage_so.get('total_gain', 0)
        our_stk_loss += anlage_so.get('total_loss', 0)

    ibkr_topf2_cats = ["Aktien- und Indexoptionen", "Futures",
                       "Optionen auf Futures (Future-Style)",
                       "Optionen auf Futures", "Anleihen", "Treasury Bills"]
    zufluss_adj = (audit.get('zufluss_premium_eur', 0)
                   - audit.get('prior_zufluss_correction_eur', 0))
    our_topf2_gain = (d.get('options_gain_eur', 0)
                      - audit.get('stillhalter_premium_eur', 0)
                      - d.get('fx_total_gain', 0) - no_invstg_gain
                      - zufluss_adj)
    our_topf2_loss = (d.get('options_loss_eur', 0)
                      - d.get('fx_total_loss', 0) - no_invstg_loss)
    ibkr_topf2_gain = sum(
        csv_cats.get(c, {}).get('gain', 0) for c in ibkr_topf2_cats)
    ibkr_topf2_loss = sum(
        csv_cats.get(c, {}).get('loss', 0) for c in ibkr_topf2_cats)
    ibkr_stk = csv_cats.get('Aktien', {})
    ibkr_fx = csv_cats.get('Devisen', {})
    fx_total_gain = d.get('fx_total_gain', 0)
    fx_total_loss = d.get('fx_total_loss', 0)

    our_div = d.get('dividends_eur', 0)
    our_wht = d.get('withholding_tax_eur', 0)
    if has_etf and invstg_on:
        our_div += kap_inv_data.get('etf_dividends_raw_eur', 0)
        our_wht += kap_inv_data.get('etf_wht_eur', 0)

    rows = [
        ("Aktien (Topf 1) Netto", ibkr_stk.get('net', 0),
         our_stk_gain + our_stk_loss),
        ("Sonstiges (Topf 2) Netto", ibkr_topf2_gain + ibkr_topf2_loss,
         our_topf2_gain + our_topf2_loss),
        ("FX (Devisen) Netto", ibkr_fx.get('net', 0),
         fx_total_gain + fx_total_loss),
    ]
    csv_income = d.get('csv_income_totals', {})
    if 'dividends_eur' in csv_income:
        rows.append(("Dividenden", csv_income['dividends_eur'], our_div))
    if 'interest_eur' in csv_income:
        rows.append(("Zinsen", csv_income['interest_eur'],
                     d.get('interest_eur', 0) + d.get('debit_interest_eur', 0)))
    if 'withholding_tax_eur' in csv_income:
        rows.append((
            "Quellensteuer",
            calculate_tax_report.get_withholding_tax_for_reporting(
                csv_income['withholding_tax_eur']),
            our_wht,
        ))

    fx_meta = d.get('fx_option_a_meta', {}) or {}
    fx_corr_active = d.get('fx_margin_correction_enabled', True)
    fx_margin_diff = (fx_meta.get('corrected_total', 0.0)
                      - fx_meta.get('raw_total', 0.0))
    fx_margin_explains = fx_corr_active and abs(fx_margin_diff) > 0.01

    result_rows = []
    all_match = True
    zinsen_fx_diff = False
    fx_saldo_diff = False
    for label, ibkr_val, our_val in rows:
        diff = our_val - ibkr_val
        match = abs(diff) < 1.0
        is_fx_saldo = (label == "FX (Devisen) Netto" and fx_margin_explains
                       and abs(diff - fx_margin_diff) < 1.0)
        if not match:
            if label == "Zinsen":
                zinsen_fx_diff = True
            elif is_fx_saldo:
                fx_saldo_diff = True
            else:
                all_match = False
        result_rows.append({
            'label': label, 'ibkr': ibkr_val, 'ours': our_val, 'diff': diff,
            'match': match, 'is_fx_saldo': is_fx_saldo,
        })
    return {
        'rows': result_rows,
        'all_match': all_match,
        'zinsen_fx_diff': zinsen_fx_diff,
        'fx_saldo_diff': fx_saldo_diff,
        'zufluss_adj': zufluss_adj,
        'fx_margin_diff': fx_margin_diff,
        'fx_margin_relevant': fx_margin_explains,
    }


plaus = build_plausibility(d, _eff_toggles)

_vm_context = {
    'input_key': _requested_key,
    'multi_stmt_files': _payload['multi_stmt_files'],
    'accounts_skipped': _payload['accounts_skipped'],
    'dropped_duplicates': _dataset.get('dropped_duplicates', []),
    'csv_disabled_multi_account': (
        _payload['csv_present'] and not _payload['csv_enabled']
    ),
    'plausibility_mismatch': bool(plaus) and not plaus['all_match'],
}
vm = ui_model.build_view_model(
    d, per_account_wht_pools,
    {
        'toggles': _dom['toggles'],
        'etf_overrides': _dom['etf_overrides'],
        'dba_beta_enabled': _dom['toggles'].get('dba_beta', False),
    },
    _vm_context,
)
final = vm['final']
per_account_finals = ui_model.build_per_account_finals(
    reports,
    per_account_wht_pools,
    {
        'toggles': _dom['toggles'],
        'etf_overrides': _dom['etf_overrides'],
        'dba_beta_enabled': _dom['toggles'].get('dba_beta', False),
    },
) if n_accounts > 1 else []
toggles = vm['toggles']
audit = d.get('audit', {}) or {}

# ── Sidebar fuellen: Datengrundlage, Navigation, Berechnungsoptionen ─────────

with _data_card_slot:
    file_names = ", ".join(esc(f['name']) for f in _dataset_files)
    acct_line = esc(", ".join(account_names)) if account_names else "-"
    st.markdown(
        f'<div class="data-card">'
        f'<div class="data-card-title">Steuerjahr {esc(steuerjahr)} · '
        f'{esc(d.get("base_currency", "EUR"))}-Konto'
        f'{"" if n_accounts == 1 else f" · {n_accounts} Konten"}</div>'
        f'{acct_line}'
        f'<div class="data-card-files">{file_names}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    for info in _payload['quarterly_infos']:
        st.markdown(
            f'<div class="data-card">'
            f'<div class="data-card-title">{info["count"]} Quartals-XMLs '
            f'zusammengeführt</div>'
            f'{esc(info["account"])} · {esc(", ".join(info["periods"]))}'
            f'</div>',
            unsafe_allow_html=True,
        )

_nav_current = ui_model.normalize_nav(
    st.session_state.get('nav', 'overview'), vm['visible_pages'],
)
st.session_state['nav'] = _nav_current


def _select_nav(page_id):
    st.session_state['nav'] = page_id


_nav_icons = {
    'overview': ':material/grid_view:',
    'kap': ':material/description:',
    'kap_inv': ':material/account_balance:',
    'anlage_so': ':material/paid:',
    'prueffaelle': ':material/rule:',
    'rechenwege': ':material/calculate:',
    'export': ':material/download:',
}

with _nav_slot:
    st.markdown('<div class="eyebrow" style="margin-top:0.9rem;">Bericht</div>',
                unsafe_allow_html=True)
    with st.container(key="nav_buttons"):
        for _page_id in vm['visible_pages']:
            _label = ui_model.page_label(_page_id)
            if _page_id == 'prueffaelle':
                _n = vm['notice_counts']['prueffaelle']
                if _n:
                    _label = f"{_label} · {_n}"
            st.button(
                _label,
                key=f"_ui_nav_{_page_id}",
                icon=_nav_icons.get(_page_id),
                type="primary" if _page_id == _nav_current else "tertiary",
                use_container_width=True,
                on_click=_select_nav,
                args=(_page_id,),
            )

with _toggle_slot:
    _avail = vm['toggle_availability']
    st.caption(
        "Diese Schalter ändern die Berechnung. Die konkrete Wirkung steht "
        "jeweils direkt darunter."
    )
    if _avail.get('zufluss'):
        _bind_toggle(
            _dom, 'zufluss', f"_ui_tg_zufluss_{_dataset_id[:12]}",
            "Zuflussprinzip (BMF Rn. 25, 33)",
            "Verschiebt Assignment-Prämien aus Vorjahren aus dem aktuellen "
            "Steuerjahr heraus; sie gehören in die Erklärung des "
            "Zuflussjahres. Offene Positionen und Vorjahres-Korrekturen sind "
            "bereits automatisch berechnet.",
            "ordnet Optionsprämien dem richtigen Zuflussjahr zu und "
            "verhindert eine doppelte Erfassung; KAP-Zeilen können sich "
            "dadurch ändern.",
        )
    if _avail.get('invstg'):
        _bind_toggle(
            _dom, 'invstg', f"_ui_tg_invstg_{_dataset_id[:12]}",
            "InvStG-Klassifizierung (KAP-INV)",
            "ETFs als Investmentfonds nach InvStG: separate Meldung auf "
            "Anlage KAP-INV mit Teilfreistellung je Fondsart. Deaktivieren "
            "behandelt alle ETFs wie normale Aktien auf Anlage KAP.",
            "meldet Fondswerte auf Anlage KAP-INV und berücksichtigt die "
            "hinterlegte Teilfreistellung; ohne die Methode laufen die "
            "Werte über Anlage KAP.",
        )
    if _avail.get('tageskurs'):
        _fx_corr_total = d.get('fx_correction_total', 0)
        _bind_toggle(
            _dom, 'tageskurs', f"_ui_tg_tageskurs_{_dataset_id[:12]}",
            "Tageskurs-Methode "
            f"({'+' if _fx_corr_total >= 0 else ''}{fmt_de(_fx_corr_total)} EUR)",
            "Rechnet Erlöse und Anschaffungskosten jeweils zum FX-Kurs ihres "
            "eigenen Datums um (§20 Abs. 4 S. 1 EStG) statt den Netto-PnL "
            "zum Schlusskurs. Benötigt CLOSED_LOT-Daten der Flex Query.",
            "ersetzt für betroffene Geschäfte die IBKR-Schlusskursumrechnung; "
            f"erkannte Gesamtdifferenz in diesem Bericht: "
            f"{'+' if _fx_corr_total >= 0 else ''}"
            f"{fmt_de(_fx_corr_total)} EUR. Die betroffenen Steuerzeilen "
            "werden neu berechnet.",
        )
    if _avail.get('fx_margin'):
        _bind_toggle(
            _dom, 'fx_margin', f"_ui_tg_fx_margin_{_dataset_id[:12]}",
            "FX-Saldo-Korrektur",
            "Negative Fremdwährungssalden werden als Margin-Schuld behandelt: "
            "die Tilgung einer Schuld veräußert kein Fremdwährungsguthaben "
            "(BMF Rn. 131). Deaktivieren übernimmt die IBKR-Rohwerte. "
            "Ändert die Berechnung und löst einen Neulauf aus.",
            "entfernt Tilgungen von Fremdwährungsschulden aus dem "
            "steuerpflichtigen FX-Ergebnis; insbesondere Zeile 19 kann sich "
            "ändern.",
        )
    with st.expander("DBA-Prüfung Fonds-Quellensteuer", expanded=False):
        st.markdown('<span class="badge-beta">Beta</span>',
                    unsafe_allow_html=True)
        st.caption(
            "Standard: Rohsteuer × (1 − Teilfreistellung). Die Beta prüft "
            "jedes Ereignis einzeln: Ausschüttung, Einbehalt und Erstattung "
            "werden gematcht, hinterlegte DBA-Höchstsätze (z. B. 15 % nach "
            "Art. 10 Abs. 2 DBA-USA) und der deutsche 25%-Höchstbetrag "
            "(BMF Rn. 148) begrenzen Zeile 41; Auffälligkeiten werden "
            "Prüffälle."
        )
        _bind_toggle(
            _dom, 'dba_beta', f"_ui_tg_dba_beta_{_dataset_id[:12]}",
            "Beta aktivieren",
            "Ändert die Berechnung und löst einen Neulauf aus. Werte vor "
            "Übernahme in die Steuererklärung manuell prüfen.",
            "kann Zeile 41 und die zugehörigen Prüffälle verändern; andere "
            "Formularzeilen bleiben unverändert.",
        )
        if _dom['toggles'].get('dba_beta'):
            st.caption(
                "Beta aktiv: unbekannte oder zeitversetzte Vorgänge bleiben "
                "Prüffälle im Bereich Anlage KAP-INV."
            )


# ── Gemeinsame Ableitungen fuer die Renderer (nur Lesezugriff auf vm/d) ──────

kap_inv = vm['kap_inv']
etf_by_isin = vm['etf_by_isin']
has_etf_data = vm['has_etf_data']
kap_inv_form = vm['kap_inv_form']
no_invstg_summary = vm['no_invstg_summary']
anlage_so = vm['anlage_so']
has_so_data = vm['has_so_data']

invstg_aktiv = toggles['invstg']
tageskurs_aktiv = toggles['tageskurs']
zuflussprinzip_aktiv = toggles['zufluss']
de_kest_variante_b = toggles['variante_b']
dba_wht_beta_enabled = vm['dba_beta_enabled']

cross_year_premium = audit.get('cross_year_premium_eur', 0)
cross_year_by_year = audit.get('cross_year_by_year', {})
cross_year_details = [
    det for det in audit.get('stillhalter_details', [])
    if det.get('is_cross_year')
]
zufluss_details = audit.get('zufluss_details', [])
zufluss_premium = audit.get('zufluss_premium_eur', 0)
prior_zufluss_details = audit.get('prior_zufluss_details', [])
prior_zufluss_correction = audit.get('prior_zufluss_correction_eur', 0)

fx_corr_total = d.get('fx_correction_total', 0)
tk_gain_adj = d.get('fx_corr_gain_adj', {}) or {}
tk_loss_adj = d.get('fx_corr_loss_adj', {}) or {}
tageskurs_kapinv_corr = final.get('tageskurs_kapinv_corr', 0)
tageskurs_kapinv_corr_raw = final.get('tageskurs_kapinv_corr_raw', 0)
adj_cross = final.get('adj_cross', 0)

fx_results = d.get('fx_results', {}) or {}
fx_total_gain = d.get('fx_total_gain', 0)
fx_total_loss = d.get('fx_total_loss', 0)
fx_mtm = d.get('fx_mtm', {}) or {}
fx_source = d.get('fx_source', 'none')

so_taxable = anlage_so.get('taxable_gain', 0) + anlage_so.get('taxable_loss', 0)
so_free = anlage_so.get('tax_free_gain', 0) + anlage_so.get('tax_free_loss', 0)
so_total = anlage_so.get('total_gain', 0) + anlage_so.get('total_loss', 0)

topf2_cats = d.get('topf2_by_category', {}) or {}
topf2_breakdown = None
if topf2_cats:
    topf2_breakdown = calculate_tax_report.build_topf2_breakdown(
        topf2_cats,
        final['dividends'],
        final['interest'],
        tageskurs_gain_adjustment=(
            tk_gain_adj.get('Topf2', 0) if tageskurs_aktiv else 0
        ),
        tageskurs_loss_adjustment=(
            tk_loss_adj.get('Topf2', 0) if tageskurs_aktiv else 0
        ),
        zufluss_adjustment=(-adj_cross if zuflussprinzip_aktiv else 0),
    )

prueffall_notices = [
    n for n in vm['notices'] if n['class'] == 'prueffall'
]
kritisch_notices = [
    n for n in prueffall_notices if n['severity'] == 'kritisch'
]
transparenz_notices = [
    n for n in vm['notices'] if n['class'] == 'transparenz'
]


def _inline_marker(page_id):
    """Schmaler Hinweis in Fachbereichen statt voller Warnboxen."""
    related = [n for n in prueffall_notices if n.get('target') == page_id]
    if related:
        if len(related) == 1:
            st.caption(
                "1 Prüffall betrifft diesen Bereich; Details im Bereich "
                "Prüffälle."
            )
        else:
            st.caption(
                f"{len(related)} Prüffälle betreffen diesen Bereich; "
                "Details im Bereich Prüffälle."
            )


# ── Renderer: Übersicht ──────────────────────────────────────────────────────

def render_overview():
    st.markdown('<p class="page-title">Steuerbericht ' + esc(steuerjahr) +
                '</p>', unsafe_allow_html=True)
    base_curr = d.get('base_currency', 'EUR')
    curr_hint = (
        "StmtFunds-Beträge bereits in EUR" if base_curr == "EUR"
        else "USD-Beträge über Tageskurse in EUR umgerechnet"
    )
    accounts_label = (
        esc(", ".join(account_names)) if n_accounts == 1
        else f"{n_accounts} Konten, separat berechnet und addiert"
    )
    st.markdown(
        f'<div class="meta-row">'
        f'<span>Konto: <strong>{accounts_label}</strong></span>'
        f'<span>Basiswährung: <strong>{esc(base_curr)}</strong> · {esc(curr_hint)}</span>'
        f'<span>Stand: <strong>{esc(created_at)}</strong></span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    kap_inv_lines_for_summary = (
        kap_inv_form.get('lines', []) if (has_etf_data and invstg_aktiv) else []
    )
    summary_so_rows = []
    if has_so_data:
        summary_so_rows.append({
            'line': 'SO',
            'label': 'Steuerpflichtiger Gewinn/Verlust (bis 1 Jahr)',
            'value': so_taxable,
            'highlight': True,
        })
        if abs(so_free) > 0.01:
            summary_so_rows.append({
                'line': 'SO',
                'label': 'Steuerfrei (über 1 Jahr Haltedauer)',
                'value': so_free,
            })

    n_prueffaelle = vm['notice_counts']['prueffaelle']
    if n_prueffaelle == 1:
        review_items = [
            "1 offener Prüffall – bitte vor der Abgabe im Bereich Prüffälle "
            "ansehen."
        ]
    elif n_prueffaelle > 1:
        review_items = [
            f"{n_prueffaelle} offene Prüffälle – bitte vor der Abgabe im "
            "Bereich Prüffälle ansehen."
        ]
    else:
        review_items = []
    kap_inv_has_review = any(
        n.get('target') == 'kap_inv' for n in prueffall_notices
    )

    st.markdown(
        build_tax_result_summary_html(
            steuerjahr,
            final,
            kap_inv_lines=kap_inv_lines_for_summary,
            kap_inv_enabled=has_etf_data and invstg_aktiv,
            kap_status=(
                "Prüffälle vorhanden" if n_prueffaelle else "Berechnet"
            ),
            kap_status_tone="warning" if n_prueffaelle else "ok",
            kap_inv_status=(
                "Vorläufig / prüfen" if kap_inv_has_review else "Berechnet"
            ),
            kap_inv_status_tone="warning" if kap_inv_has_review else "ok",
            review_items=review_items,
            so_rows=summary_so_rows,
        ),
        unsafe_allow_html=True,
    )

    if de_kest_variante_b and abs(final.get('zeile_7', 0)) < 0.01:
        z7_raw = d.get(
            'zeile_7_kapitalertraege_mit_inlaendischem_steuerabzug_eur', 0)
        z_kest = (d.get('zeile_37_kapitalertragsteuer_eur', 0)
                  + d.get('zeile_38_solidaritaetszuschlag_eur', 0))
        st.caption(
            f"Variante B aktiv: {fmt_de(z7_raw)} EUR Bruttodividende auf "
            f"DE-ISINs zu Zeile 19 addiert, {fmt_de(z_kest)} EUR DE-KESt+Soli "
            "zu Zeile 41. Umschalten im Bereich Anlage KAP."
        )

    # Status-Streifen
    chips = []
    if plaus:
        if plaus['all_match']:
            chips.append('<span class="status-chip ok">Plausibilitätscheck ok</span>')
        else:
            chips.append('<span class="status-chip warning">Plausibilitätscheck abweichend</span>')
    if n_prueffaelle:
        chips.append(
            f'<span class="status-chip warning">{n_prueffaelle} Prüffälle</span>'
        )
    else:
        chips.append('<span class="status-chip ok">Keine Prüffälle</span>')
    if has_etf_data and invstg_aktiv:
        chips.append('<span class="status-chip">KAP-INV vorhanden</span>')
    if has_so_data:
        chips.append('<span class="status-chip">Anlage SO vorhanden</span>')
    if zuflussprinzip_aktiv and cross_year_details:
        chips.append('<span class="status-chip">Zuflussprinzip aktiv</span>')
    if tageskurs_aktiv:
        chips.append('<span class="status-chip">Tageskurs-Methode aktiv</span>')
    st.markdown(
        '<div style="display:flex;gap:0.4rem;flex-wrap:wrap;margin:0.2rem 0 1rem 0;">'
        + ''.join(chips) + '</div>',
        unsafe_allow_html=True,
    )

    # Sekundärkarten
    secondary = [
        metric_card("Topf 1 · Aktien", final['topf_1'], "saldo"),
        metric_card("Topf 2 · Sonstiges", final['topf_2'], "saldo"),
    ]
    if has_etf_data and invstg_aktiv:
        secondary.append(metric_card(
            "KAP-INV steuerpflichtig (Kontrollwert)",
            final['etf_net_taxable'], "info",
        ))
    st.markdown(metric_grid(*secondary), unsafe_allow_html=True)
    st.caption(
        "Zeile 19 = Topf 1 + Topf 2. Herleitung und Zwischenwerte stehen im "
        "Bereich Anlage KAP."
    )


# ── Renderer: Anlage KAP ─────────────────────────────────────────────────────

def render_kap():
    st.markdown('<p class="page-title">Anlage KAP</p>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="meta-row"><span>Zeile 19 ergibt sich aus '
        f'Topf 1 (<strong>{fmt(final["topf_1"])}</strong>) und '
        f'Topf 2 (<strong>{fmt(final["topf_2"])}</strong>).</span></div>',
        unsafe_allow_html=True,
    )
    _inline_marker('kap')

    topf_1_label = (
        "Topf 1 · Aktien ohne ETF-Fonds (§20 Abs. 6 S. 4 EStG)"
        if (has_etf_data and invstg_aktiv)
        else "Topf 1 · Aktien (§20 Abs. 6 S. 4 EStG)"
    )
    section_title(topf_1_label)
    st.markdown(metric_grid(
        metric_card("Aktiengewinne", final['stocks_gain'], "gain"),
        metric_card("Aktienverluste", final['stocks_loss'], "loss"),
        metric_card("Saldo Aktien", final['topf_1'], "saldo"),
    ), unsafe_allow_html=True)

    cross_put_corrections = audit.get('cross_year_put_corrections', [])
    cross_put_total = audit.get('cross_year_put_total', 0)
    if cross_put_corrections:
        has_invstg_basis_extra = any(
            c.get('invstg_basis_extra_per_share_raw', 0) > 0
            for c in cross_put_corrections
        )
        cross_put_explanation = (
            f"Die Kostenbasis-Korrektur ({fmt_de(cross_put_total)} EUR) "
            "stellt den Einstandskurs auf den Put-Ausübungspreis. Sie "
            "umfasst die bereits im Assignment-Jahr versteuerte Prämie und "
            "eine zusätzliche, für KAP-INV nicht übernommene ausländische "
            "Basisreduktion (z. B. ROC)."
            if has_invstg_basis_extra else
            f"Die Prämie ({fmt_de(cross_put_total)} EUR) wurde bereits im "
            "Assignment-Jahr versteuert und wird hier vom Aktien-PnL "
            "abgezogen (Einstandskurs = Strike, nicht Strike minus "
            "Prämie)."
        )
        st.markdown(notice_html({
            'class': 'transparenz', 'severity': 'normal',
            'title': 'Put-Assignment-Korrektur (BMF Rn. 33)',
            'body': (
                f"{len(cross_put_corrections)} Aktienverkäufe stammen aus "
                f"Put-Assignments früherer Jahre. {cross_put_explanation}"
            ),
            'target': None,
        }, show_target=False), unsafe_allow_html=True)
        with st.expander(
                f"Details: {len(cross_put_corrections)} Cross-Year "
                "Put-Korrekturen"):
            put_table = ("| Symbol | Shares | Strike | Korrektur | Aus Jahr |\n"
                         "|--------|--------|--------|-----------|----------|\n")
            for c in cross_put_corrections:
                put_table += (
                    f"| {c['symbol']} | {c['shares']} | {c['strike']} | "
                    f"{fmt_de(c['correction_eur'])} EUR | "
                    f"{c['assignment_year']} |\n"
                )
            put_table += (
                f"| **Gesamt** | | | **{fmt_de(cross_put_total)} EUR** | |\n"
            )
            st.markdown(put_table)

    section_title("Topf 2 · Sonstiges (Termingeschäfte, Dividenden, Zinsen)")
    st.markdown(metric_grid(
        metric_card("Dividenden", final['dividends']),
        metric_card("Zinsen (netto)", final['interest']),
        (metric_card("Sollzinsen (n. abzf.)", d.get('debit_interest_eur', 0),
                     "info")
         if abs(d.get('debit_interest_eur', 0)) > 0.01 else ''),
        (metric_card("Gebühren (nachrichtl.)", d.get('other_fees_eur', 0),
                     "info")
         if abs(d.get('other_fees_eur', 0)) > 0.01 else ''),
        metric_card("Sonstige Gewinne", final['options_gain'], "gain"),
        metric_card("Sonstige Verluste", final['options_loss'], "loss"),
        metric_card("Saldo Sonstiges", final['topf_2'], "saldo"),
    ), unsafe_allow_html=True)

    if topf2_breakdown:
        section_title("Aufschlüsselung Topf 2")
        cat_table = ("| Gattung | Gewinne | Verluste | Netto |\n"
                     "|---------|--------:|---------:|------:|\n")
        for row in topf2_breakdown['rows']:
            cat_table += (
                f"| {row['label']} | {fmt_de(row['gain'])} EUR | "
                f"{fmt_de(row['loss'])} EUR | {fmt_de(row['net'])} EUR |\n"
            )
        cat_table += (
            f"| **Saldo Topf 2** | **{fmt_de(topf2_breakdown['total_gain'])} EUR** | "
            f"**{fmt_de(topf2_breakdown['total_loss'])} EUR** | "
            f"**{fmt_de(topf2_breakdown['net'])} EUR** |\n"
        )
        st.markdown(cat_table)
        if any(row['is_adjustment'] for row in topf2_breakdown['rows']):
            st.caption(
                "Anpassungszeilen verändern Gewinn- und Verlustspalten "
                "mit Vorzeichen; beide Spalten sind bis zur Summenzeile "
                "addierbar."
            )
        if abs(topf2_breakdown['net'] - final['topf_2']) > 0.01:
            st.error(
                "Interner Abstimmungsfehler: Die Topf-2-Detailzeilen "
                "stimmen nicht mit dem ausgewiesenen Saldo überein."
            )

    _render_kap_fx_section()
    _render_kap_sonderprodukte()
    _render_kap_variante_b()
    _render_kap_multi_account()


def _render_kap_fx_section():
    if not fx_results:
        return
    src_label = {
        'csv': 'IBKR-Bericht', 'xml': 'XML FxTransactions',
    }.get(fx_source, 'FIFO-Approximation')
    section_title(f"Fremdwährungs-Gewinne/Verluste ({src_label})")
    st.markdown(metric_grid(
        metric_card("FX Gewinne", fx_total_gain, "gain"),
        metric_card("FX Verluste", fx_total_loss, "loss"),
        metric_card("FX Netto", fx_total_gain + fx_total_loss, "saldo"),
    ), unsafe_allow_html=True)

    fx_has_neg = d.get('fx_has_negative_balance', False)
    fx_opt_a_meta = d.get('fx_option_a_meta', {}) or {}
    correction_enabled = d.get(
        'fx_margin_correction_enabled',
        fx_opt_a_meta.get('correction_enabled', True),
    )
    has_raw_data = any('raw_net' in data for data in fx_results.values())
    has_corrected_data = any(
        'corrected_net' in data for data in fx_results.values())
    active_total = fx_total_gain + fx_total_loss
    raw_total = fx_opt_a_meta.get('raw_total', sum(
        data.get('raw_net', data.get('net', 0.0))
        for data in fx_results.values()))
    corrected_total = fx_opt_a_meta.get('corrected_total', sum(
        data.get('corrected_net', data.get('net', 0.0))
        for data in fx_results.values()))
    diff_corr_raw = corrected_total - raw_total
    total_neg_days = max(
        (data.get('days_negative', 0) for data in fx_results.values()),
        default=0)
    debt_repayments = fx_opt_a_meta.get('debt_repayments', 0)
    debt_repayment_pnl = fx_opt_a_meta.get('debt_repayment_pnl', 0.0)
    corrected_gain_total = sum(
        data.get('corrected_gain', data.get('gain', 0.0))
        for data in fx_results.values())
    raw_gain_total = sum(
        data.get('raw_gain', data.get('gain', 0.0))
        for data in fx_results.values())
    corrected_loss_total = sum(
        data.get('corrected_loss', data.get('loss', 0.0))
        for data in fx_results.values())
    raw_loss_total = sum(
        data.get('raw_loss', data.get('loss', 0.0))
        for data in fx_results.values())
    diff_gain = corrected_gain_total - raw_gain_total
    diff_loss = corrected_loss_total - raw_loss_total

    if (has_raw_data or has_corrected_data) and (
            fx_has_neg or debt_repayments or abs(diff_corr_raw) > 0.01
            or abs(diff_gain) > 0.01 or abs(diff_loss) > 0.01):
        comparison_label = ("gegenüber IBKR" if fx_source == 'xml'
                            else "gegenüber dem Rohwert")
        raw_label = ("IBKR-Rohwert vor Saldo-Prüfung" if fx_source == 'xml'
                     else "Rohwert vor Saldo-Prüfung")
        unchanged_label = "IBKR-Wert" if fx_source == 'xml' else "Rohwert"
        neg_currs = [c for c, dt in fx_results.items()
                     if dt.get('days_negative', 0) > 0]
        if len(neg_currs) == 1:
            curr_label = f"{neg_currs[0]}-Konto"
        elif neg_currs:
            curr_label = f"{' / '.join(neg_currs)}-Konten"
        else:
            curr_label = "Fremdwährungskonto"

        if abs(diff_corr_raw) > 0.005:
            effect_word = "erhöht" if diff_corr_raw > 0 else "reduziert"
            if correction_enabled:
                correction_text = (
                    f"Die Saldo-Prüfung {effect_word} den Topf-2-FX-Wert "
                    f"{comparison_label} um {fmt_de(abs(diff_corr_raw))} EUR."
                )
            else:
                correction_text = (
                    f"Bei aktivierter Saldo-Prüfung läge der Topf-2-FX-Wert "
                    f"{comparison_label} um {fmt_de(abs(diff_corr_raw))} EUR "
                    f"{'höher' if diff_corr_raw > 0 else 'niedriger'}."
                )
        elif abs(diff_gain) > 0.005 or abs(diff_loss) > 0.005:
            correction_text = (
                f"Der Netto-FX-Wert bleibt gleich, die Bruttowerte nicht: "
                f"{fmt_de(abs(diff_gain))} EUR Gewinne und "
                f"{fmt_de(abs(diff_loss))} EUR Verluste fallen weg. Das "
                "wirkt auf Zeile 22, nicht auf Zeile 19."
            )
        else:
            correction_text = (
                f"Die Saldo-Prüfung ändert den {unchanged_label} rechnerisch "
                "nicht."
            )

        detail_parts = []
        if debt_repayments:
            if correction_enabled:
                detail_parts.append(
                    f"{debt_repayments} Buchungen haben eine "
                    f"Fremdwährungs-Schuld getilgt statt Guthaben zu "
                    f"veräußern; ihr IBKR-Ergebnis von "
                    f"{fmt_de(debt_repayment_pnl)} EUR bleibt außen vor."
                )
            else:
                detail_parts.append(
                    f"{debt_repayments} Buchungen haben eine "
                    f"Fremdwährungs-Schuld getilgt; bei aktivierter Korrektur "
                    f"bliebe ihr IBKR-Ergebnis von "
                    f"{fmt_de(debt_repayment_pnl)} EUR außen vor."
                )
            detail_parts.append(
                "Abflüsse werden ungekürzt übernommen: IBKR weist bei nur "
                "teilweise gedeckten Buchungen bereits allein das Ergebnis "
                "des gedeckten Teils aus."
            )

        situation = (
            f"An {total_neg_days} Tagen im Steuerjahr war der "
            "Fremdwährungssaldo negativ. "
            if total_neg_days else
            f"IBKR weist {debt_repayments} Buchungen aus, die eine "
            "Fremdwährungs-Schuld geschlossen haben. "
        )
        stance = (
            "Deshalb rechnet das Tool Tilgungen einer Fremdwährungs-Schuld "
            "heraus." if correction_enabled else
            "Die Korrektur ist deaktiviert; Topf 2 übernimmt den Rohwert."
        )
        compare_line = (
            f"{raw_label}: {fmt_de(raw_total)} EUR."
            if correction_enabled else
            f"Wert mit Saldo-Prüfung: {fmt_de(corrected_total)} EUR."
        )
        body = (
            f"{situation}Das ist steuerlich Margin-Schuld, kein "
            "Fremdwährungsguthaben (§20 Abs. 2 S. 1 Nr. 7 i.V.m. Abs. 4 S. 1 "
            f"EStG, BMF Rn. 131). {stance} In Topf 2 übernommen: "
            f"{fmt_de(active_total)} EUR. {compare_line} {correction_text} "
            + " ".join(detail_parts)
        )
        st.markdown(notice_html({
            'class': 'transparenz', 'severity': 'normal',
            'title': (
                f"FX-Saldo-Korrektur: {curr_label} war zeitweise im Minus"
                if total_neg_days else
                "FX-Saldo-Korrektur: Fremdwährungs-Schuld getilgt"
            ) + ("" if correction_enabled else " (deaktiviert)"),
            'body': body, 'target': None,
        }, show_target=False), unsafe_allow_html=True)

    fx_open_anomalies = fx_opt_a_meta.get('open_rows_with_pnl', []) or []
    if fx_open_anomalies:
        st.warning(
            f"{len(fx_open_anomalies)} Devisen-Buchungen tragen ein "
            "realisiertes Ergebnis, obwohl IBKR sie als Eröffnung ausweist. "
            "Sie wurden als steuerbar behandelt; bitte prüfen."
        )
        with st.expander(f"Betroffene Buchungen ({len(fx_open_anomalies)})"):
            anom_table = (
                "| Datum | Währung | Code | Menge | Ergebnis EUR | Beschreibung |\n"
                "|-------|---------|------|------:|-------------:|--------------|\n")
            for a in fx_open_anomalies[:50]:
                anom_table += (
                    f"| {esc(a.get('date', ''))} "
                    f"| {esc(a.get('currency', ''))} "
                    f"| {esc(a.get('code', ''))} "
                    f"| {fmt_de(a.get('quantity', 0))} "
                    f"| {fmt_de(a.get('realized_pnl', 0))} "
                    f"| {esc(str(a.get('description', ''))[:60])} |\n"
                )
            st.markdown(anom_table)

    with st.expander("Details pro Währung"):
        has_raw_col = any('raw_net' in data for data in fx_results.values())
        if has_raw_col:
            raw_col = "IBKR vor Prüfung" if fx_source == 'xml' else "Vor Prüfung"
            fx_table = (
                f"| Währung | Steuerlich verwendet | Nach Saldo-Prüfung | "
                f"{raw_col} | Korrektur | Minustage | IBKR-MTM |\n"
                "|---------|----------:|-----------:|---------:|----------:|"
                "----------:|----:|\n")
            for curr, data in sorted(fx_results.items()):
                mtm_val = fx_mtm.get(curr)
                mtm_str = f"{fmt_de(mtm_val)}" if mtm_val is not None else "-"
                used_net = data.get('net', 0)
                corr_net = data.get('corrected_net', used_net)
                raw_net = data.get('raw_net', corr_net)
                fx_table += (
                    f"| {curr} | {fmt_de(used_net)} | {fmt_de(corr_net)} | "
                    f"{fmt_de(raw_net)} | {fmt_de(corr_net - raw_net)} | "
                    f"{data.get('days_negative', 0)} | {mtm_str} |\n")
        else:
            fx_table = (
                "| Währung | Gewinn | Verlust | Netto | MTM (Vergleich) |\n"
                "|---------|--------|---------|-------|----------------|\n")
            for curr, data in sorted(fx_results.items()):
                mtm_val = fx_mtm.get(curr)
                mtm_str = (f"{fmt_de(mtm_val)} EUR"
                           if mtm_val is not None else "-")
                fx_table += (
                    f"| {curr} | {fmt_de(data['gain'])} | "
                    f"{fmt_de(data['loss'])} | {fmt_de(data['net'])} | "
                    f"{mtm_str} |\n")
        st.markdown(fx_table)
        st.caption(
            "Alle Werte in EUR. 'Steuerlich verwendet' folgt dem gewählten "
            "Saldo-Modus; 'Korrektur' ist die Differenz zum IBKR-Rohwert. "
            "Minustage zeigen Tage mit Fremdwährungsschuld. IBKR-MTM dient "
            "nur als Plausibilitätsreferenz."
        )
        fx_tgl = d.get('fx_translation', 0)
        if fx_tgl != 0:
            st.markdown(
                f"**IBKR Referenz (fxTranslationGainLoss):** "
                f"{fmt_de(fx_tgl)} EUR")
        if fx_source in ('csv', 'xml'):
            src_name = ("IBKR Standard-Bericht" if fx_source == 'csv'
                        else "XML FxTransactions")
            filtered = (fx_source == 'xml' and correction_enabled
                        and fx_opt_a_meta.get('debt_repayments', 0))
            value_kind = ("Buchungsgenaue IBKR-FIFO-Werte"
                          if fx_source == 'xml'
                          else "Aggregierte IBKR-FIFO-Rohwerte")
            st.success(
                f"{value_kind} aus {src_name} (alle Währungen)."
                + (" Buchungen, die eine Fremdwährungs-Schuld tilgen, sind "
                   "herausgerechnet (siehe Hinweis oben)." if filtered else "")
                + (" Einzelne Schuldtilgungen sind im CSV nicht erkennbar."
                   if fx_source == 'csv' else "")
            )
        else:
            no_xml_fx = not d.get('xml_has_fx_data', True)
            fx_prior = d.get('fx_has_prior_data', False)
            extra = (" Die Flex Query enthält keine FxTransactions; "
                     "Kursgenauigkeit der Approximation ist eingeschränkt."
                     if no_xml_fx else "")
            prefix = ("FIFO-Approximation aus Flex Query "
                      "(Tagesraten-Substitution)." if fx_prior
                      else "**Nur Steuerjahr geladen.** FIFO-Approximation.")
            st.warning(
                f"{prefix}{extra} Der IBKR-Standardbericht (CSV) kann über "
                "\"Daten ändern\" als aggregierter Rohwert hochgeladen werden "
                "(nur bei einem einzelnen Konto); eine buchungsweise "
                "Schuldtilgungsprüfung ist damit nicht möglich."
            )
        st.info(
            "**Rechtsgrundlage:** BMF-Schreiben Rn. 131, verzinsliches "
            "Fremdwährungsguthaben, §20 Abs. 2 S. 1 Nr. 7 i.V.m. Abs. 4 S. 1 "
            "EStG (Anlage KAP, Topf 2). FIFO-Methode (§20 Abs. 4 S. 7). "
            "Erfasst wird die Veräußerung von Guthaben; die Tilgung einer "
            "Fremdwährungs-Schuld zählt nicht. In Topf 2 enthalten."
        )


def _render_kap_sonderprodukte():
    if not no_invstg_summary:
        return
    section_title("Topf 2 · Sonderprodukte außerhalb InvStG")
    st.caption(
        "Einzelnachweis für Schuldverschreibungen/ETNs und sonstige ETPs "
        "außerhalb des InvStG. Die Beträge sind bereits in Topf 2 enthalten; "
        "negative QSt-Werte kennzeichnen Erstattungen."
    )
    with st.expander("Sonderprodukte nach ISIN", expanded=False):
        special_table = (
            "| Ticker | ISIN | Realisiertes G/V | Tageskurs | "
            "Ausschüttungen | QSt | Summe Topf 2 |\n"
            "|--------|------|----------------:|----------:|---------------:|"
            "----:|-------------:|\n")
        for isin, info in sorted(no_invstg_summary.items(),
                                 key=lambda x: x[1].get('ticker', '')):
            realized = info.get('gain', 0) + info.get('loss', 0)
            special_table += (
                f"| {info.get('ticker', isin)} | {isin} | "
                f"{fmt_de(realized)} | {fmt_de(info.get('tageskurs', 0))} | "
                f"{fmt_de(info.get('div', 0))} | "
                f"{fmt_de(info.get('wht_reported', 0))} | "
                f"{fmt_de(info.get('total', 0))} |\n")
        st.markdown(special_table)


def _render_kap_variante_b():
    z7_raw = d.get(
        'zeile_7_kapitalertraege_mit_inlaendischem_steuerabzug_eur', 0)
    if abs(z7_raw) <= 0.01:
        return
    z37_raw = d.get('zeile_37_kapitalertragsteuer_eur', 0)
    z38_raw = d.get('zeile_38_solidaritaetszuschlag_eur', 0)
    section_title("Deutsche Dividenden (Zeilen 7/37/38)")
    st.markdown(notice_html({
        'class': 'transparenz', 'severity': 'normal',
        'title': 'Deutsche Dividenden erkannt',
        'body': (
            f"{fmt_de(z7_raw)} EUR Bruttodividende und "
            f"{fmt_de(z37_raw + z38_raw)} EUR DE-KESt/Soli auf DE-ISINs. Die "
            "deutsche Verwahrstelle hat 26,375 % an der Quelle einbehalten "
            "(§43 EStG), auch wenn IBKR keine Steuerbescheinigung ausstellt. "
            "Variante A (Default, präzise): Eintragung in Z. 7/37/38. "
            "Variante B (technische Ersatzdarstellung): Bruttodividende nach "
            "Z. 19, KESt+Soli nach Z. 41, falls das Steuerprogramm Z. 7/37/38 "
            "ohne Bescheinigung nach §45a EStG sperrt. Variante B ist kein "
            "amtlich belegter Ersatz; vor Abgabe fachlich abstimmen."
        ),
        'target': None,
    }, show_target=False), unsafe_allow_html=True)
    _bind_toggle(
        _dom, 'variante_b', f"_ui_tg_variante_b_{_dataset_id[:12]}",
        "Variante B: DE-KESt nach Zeile 19/41 verschieben",
        "Nur aktivieren, falls das Steuerprogramm Zeile 7/37/38 mangels "
        "Steuerbescheinigung nicht freischaltet.",
    )


def _render_kap_multi_account():
    if n_accounts <= 1:
        return
    section_title(f"Aufschlüsselung nach Konten ({n_accounts})")
    with st.expander("Konten im Detail", expanded=False):
        acct_table = (
            "| Konto | Topf 1 | Topf 2 | Z. 7 | Z. 19 | Z. 37 | Z. 38 | "
            "Z. 41 |\n"
            "|-------|-------:|-------:|-----:|------:|------:|------:|"
            "------:|\n")
        for idx, (name, acct_final) in enumerate(
                zip(account_names, per_account_finals)):
            acct_table += (
                f"| Konto {idx + 1} ({name}) | "
                f"{fmt_de(acct_final['topf_1'])} | "
                f"{fmt_de(acct_final['topf_2'])} | "
                f"{fmt_de(acct_final['zeile_7'])} | "
                f"{fmt_de(acct_final['zeile_19'])} | "
                f"{fmt_de(acct_final['zeile_37'])} | "
                f"{fmt_de(acct_final['zeile_38'])} | "
                f"{fmt_de(acct_final['quellensteuer'])} |\n")
        acct_table += (
            f"| **Gesamt** | **{fmt_de(final['topf_1'])}** | "
            f"**{fmt_de(final['topf_2'])}** | **{fmt_de(final['zeile_7'])}** | "
            f"**{fmt_de(final['zeile_19'])}** | **{fmt_de(final['zeile_37'])}** | "
            f"**{fmt_de(final['zeile_38'])}** | "
            f"**{fmt_de(final['quellensteuer'])}** |\n")
        st.markdown(acct_table)
        st.caption(
            "Jedes Konto wurde vollständig separat berechnet (Trades, "
            "Dividenden, FX, Stillhalter); die Einzelergebnisse wurden "
            "anschließend addiert. Konto- und Gesamtzeile enthalten dieselben "
            "aktiven Methoden (Zufluss, InvStG, Tageskurs und gegebenenfalls "
            "Variante B). Abweichungen von einem Cent können ausschließlich "
            "durch die Darstellung gerundeter Einzelwerte entstehen."
        )


# ── Renderer: Anlage KAP-INV ─────────────────────────────────────────────────

_ETF_CLS_OPTIONS = {
    "Sonstiger Fonds (0% TFS)": 0.0,
    "Aktienfonds (30% TFS)": 0.30,
    "Mischfonds (15% TFS)": 0.15,
    "Immobilienfonds (60% TFS)": 0.60,
    "Auslands-Immobilienfonds (80% TFS)": 0.80,
}


def _render_etf_confirmation_widgets():
    """Manuelle Fondsart-Bestaetigung fuer unbekannte ETFs.

    Quelle der Wahrheit ist _dom['etf_overrides'] (dataset-genamespaced);
    die Widgets tragen _ui_*-Keys und synchronisieren per Callback. Damit
    ueberleben Bestaetigungen jeden Bereichswechsel (Streamlit loescht den
    State nicht gerenderter Widgets).
    """
    raw_kap_inv = d.get('kap_inv', {}) or {}
    etf_unknown = raw_kap_inv.get('etf_unknown_isins', []) or []
    if not etf_unknown:
        return
    st.markdown(notice_html({
        'class': 'prueffall', 'severity': 'normal',
        'title': 'Fondsart bestätigen',
        'body': (
            f"{len(etf_unknown)} Fondsprodukt(e) sind unbekannt oder wegen "
            "offener Rechtsform-/InvStG-Fragen nicht automatisch "
            "klassifiziert. Ohne Bestätigung entsteht keine "
            "KAP-INV-Formularzeile. Fondsart wählen und ausdrücklich "
            "bestätigen: Aktienfonds (mind. 51% Aktienquote, 30% TFS), "
            "Mischfonds (mind. 25%, 15% TFS), Immobilienfonds (60% TFS), "
            "Auslands-Immobilienfonds (80% TFS)."
        ),
        'target': None,
    }, show_target=False), unsafe_allow_html=True)

    tfs_to_label = {v: k for k, v in _ETF_CLS_OPTIONS.items()}
    option_labels = list(_ETF_CLS_OPTIONS.keys())
    for isin in etf_unknown:
        info = etf_by_isin.get(isin, {})
        ticker = info.get('ticker', isin[:12])
        name = info.get('name', '')
        label = f"{ticker} ({isin})" + (f": {name}" if name else "")
        if info.get('review_reason'):
            st.caption(f"Prüfgrund: {info['review_reason']}")

        sel_key = f"_ui_etf_cls_{_dataset_id[:12]}_{isin}"
        conf_key = f"_ui_etf_conf_{_dataset_id[:12]}_{isin}"
        stored_tfs = _dom['etf_overrides'].get(isin)
        default_label = tfs_to_label.get(stored_tfs, option_labels[0])

        def _sync(isin=isin, sel_key=sel_key, conf_key=conf_key):
            tfs = _ETF_CLS_OPTIONS[st.session_state[sel_key]]
            if st.session_state[conf_key]:
                _dom['etf_overrides'][isin] = tfs
            else:
                _dom['etf_overrides'].pop(isin, None)

        st.selectbox(
            label, option_labels,
            index=option_labels.index(default_label),
            key=sel_key, on_change=_sync,
        )
        st.checkbox(
            "Fondsart für die Formularzuordnung bestätigen",
            value=isin in _dom['etf_overrides'],
            key=conf_key, on_change=_sync,
        )


def render_kap_inv():
    st.markdown('<p class="page-title">Anlage KAP-INV</p>',
                unsafe_allow_html=True)
    if not (has_etf_data and invstg_aktiv):
        st.info(
            "Für die aktuelle Berechnung wurden keine Formularzeilen auf "
            "Anlage KAP-INV erzeugt. Produktzuordnungen und mögliche "
            "Prüffälle stehen in den Bereichen Rechenwege und Prüffälle."
        )
        return
    _inline_marker('kap_inv')

    etf_gain_raw = kap_inv.get('etf_gain_raw_eur', 0)
    etf_loss_raw = kap_inv.get('etf_loss_raw_eur', 0)
    etf_gain_taxable = sum(
        info.get('gain_taxable', info.get('gain', 0))
        for info in etf_by_isin.values()
    )
    etf_loss_taxable = sum(
        info.get('loss_taxable', info.get('loss', 0))
        for info in etf_by_isin.values()
    )
    etf_wht_raw = kap_inv.get('etf_wht_eur', 0)
    etf_wht = final['etf_wht']

    _render_etf_confirmation_widgets()

    kap_inv_lines = kap_inv_form.get('lines', [])
    kap_inv_distribution_control = sum(
        line.get('taxable_control_eur', 0)
        for line in kap_inv_lines if line.get('kind') == 'distribution'
    )
    kap_inv_sale_control = sum(
        line.get('taxable_control_eur', 0)
        for line in kap_inv_lines if line.get('kind') == 'sale'
    )
    st.caption(
        f"{len(kap_inv_lines)} KAP-INV-Formularzeilen werden aus den "
        "Fondsbuchungen abgeleitet; die Eintragungswerte stehen in der "
        "Übersicht. Ausländische Steuer steht ausschließlich in Anlage KAP "
        "Zeile 41."
    )

    general_warnings = [
        warning for warning in kap_inv_form.get('warnings', [])
        # Vorabpauschale und gezahlte Ausschuettungen haben direkt bei den
        # Formularwerten bzw. Prueffaellen eigene Hinweise.
        if ('Vorabpauschale' not in str(warning)
            and 'Gezahlte Dividenden' not in str(warning))
    ]

    section_title("Formularwerte")
    if kap_inv_lines:
        form_table = (
            "| KAP-INV | Fondsart | Eintragungswert vor TFS | "
            "Steuerpflichtiger Kontrollwert* |\n"
            "|---------|----------|-----------------------:|"
            "-------------------------------:|\n")
        for line in kap_inv_lines:
            form_table += (
                f"| Zeile {line['line']} | {line['fund_type']} | "
                f"{fmt_de(line['amount_raw_eur'])} | "
                f"{fmt_de(line['taxable_control_eur'])} |\n")
        st.markdown(form_table)
        st.caption(
            "* Kontrollrechnung nach Teilfreistellung, kein Eintragungswert. "
            "Veräußerungszeilen sind bis zur Berücksichtigung bereits "
            "angesetzter Vorabpauschalen noch nicht final."
        )

    has_kap_inv_review = bool(
        general_warnings
        or kap_inv_form.get('blocked_details')
        or kap_inv_form.get('negative_distribution_details')
    )
    if has_kap_inv_review:
        section_title("Noch zu prüfen")
    for warning in general_warnings:
        st.warning(warning)

    if kap_inv_form.get('blocked_details'):
        blocked_table = (
            "| Nicht zugeordnet | ISIN | Prüfgrund | Ausschüttung roh | "
            "G/V roh |\n"
            "|-----------------|------|-----------|------------------:|"
            "---------:|\n")
        for item in kap_inv_form['blocked_details']:
            blocked_table += (
                f"| {item.get('ticker', '')} | {item['isin']} | "
                f"{item.get('review_reason', 'Fondsart nicht bestätigt')} | "
                f"{fmt_de(item.get('distribution_raw_eur', 0))} | "
                f"{fmt_de(item.get('sale_raw_eur', 0))} |\n")
        st.markdown(blocked_table)

    if kap_inv_form.get('negative_distribution_details'):
        paid_table = (
            "| Prüffall: gezahlte Ausschüttungen | ISIN | Gezahlt | "
            "Erhalten |\n"
            "|----------------------------------|------|--------:|--------:|\n")
        for item in kap_inv_form['negative_distribution_details']:
            paid_table += (
                f"| {item.get('ticker', '')} ({item.get('fund_type', '')}) | "
                f"{item['isin']} | "
                f"{fmt_de(item.get('paid_distribution_eur', 0))} | "
                f"{fmt_de(item.get('received_distribution_eur', 0))} |\n")
        st.markdown(paid_table)
        st.caption(
            "Gezahlte Dividenden/Ersatzzahlungen (Short-Positionen) sind "
            "nicht in den Ausschüttungszeilen enthalten; steuerliche "
            "Behandlung manuell prüfen."
        )

    section_title("Quellensteuer")
    if dba_wht_beta_enabled:
        st.markdown('<span class="badge-beta">DBA-Prüfung · Beta</span>',
                    unsafe_allow_html=True)
        st.warning(
            f"DBA-Beta aktiv: anrechenbare Fonds-Quellensteuer "
            f"{fmt_de(etf_wht)} EUR. Ereignis-Matching und DBA-Caps sind "
            "experimentell; der Betrag ist bereits in Anlage KAP Zeile 41 "
            "enthalten."
        )
    else:
        st.info(
            f"Standardberechnung Fonds-Quellensteuer: {fmt_de(etf_wht)} EUR "
            "(Rohsteuer × (1 − Teilfreistellung)); bereits in Anlage KAP "
            "Zeile 41 enthalten. Die optionale DBA-Beta ist deaktiviert."
        )

    # Modus-Vergleich Standard vs. Beta — KONTOWEISE auf Kopien (die Beta ist
    # durch den Refund-Offset nichtlinear; nie auf dem gemergten Pool rechnen).
    has_wht_activity = any(
        abs(info.get('wht', 0)) > 0.005 or info.get('wht_events')
        for info in etf_by_isin.values()
    )
    if has_wht_activity:
        compare_pools = []
        for pool in per_account_wht_pools or [etf_by_isin]:
            pool_copy = copy.deepcopy(pool)
            for isin, entry in pool_copy.items():
                merged_entry = etf_by_isin.get(isin)
                if merged_entry:
                    entry['tfs_rate'] = merged_entry.get(
                        'tfs_rate', entry.get('tfs_rate'))
                    entry['classification'] = merged_entry.get(
                        'classification', entry.get('classification'))
            compare_pools.append(pool_copy)
        mode_compare = calculate_tax_report.compare_kap_inv_wht_modes(
            compare_pools)
        wht_mode_diff = mode_compare['difference_eur']
        compare_table = (
            "| Fonds-QSt in Zeile 41 | Betrag |\n"
            "|-----------------------|-------:|\n"
            f"| Standardberechnung | {fmt_de(mode_compare['standard_eur'])} |\n"
            f"| DBA-Beta | {fmt_de(mode_compare['beta_eur'])} |\n"
            f"| Differenz | {fmt_de(wht_mode_diff)} |\n")
        st.markdown(compare_table)
        if abs(wht_mode_diff) <= 0.005:
            st.caption(
                "Die DBA-Beta ändert an diesen Daten nichts; beide Modi "
                "liefern denselben anrechenbaren Betrag."
            )
        else:
            richtung = "erhöht" if wht_mode_diff > 0 else "verringert"
            st.caption(
                f"Die DBA-Beta {richtung} die anrechenbare "
                f"Fonds-Quellensteuer um {fmt_de(abs(wht_mode_diff))} EUR "
                "gegenüber der Standardberechnung. Aktivieren in der "
                "Sidebar unter Berechnung."
            )

    wht_metric_html = metric_card(
        "Fonds-Quellensteuer → KAP Z. 41", etf_wht, "info")
    if abs(etf_wht_raw - etf_wht) > 0.01:
        wht_metric_html = (
            metric_card("Fonds-Quellensteuer roh", etf_wht_raw, "info")
            + wht_metric_html)
    tk_metric_html = ""
    if tageskurs_aktiv and abs(tageskurs_kapinv_corr_raw) > 0.01:
        tk_metric_html += metric_card(
            "Tageskurs-Anpassung vor TFS", tageskurs_kapinv_corr_raw, "info")
        tk_metric_html += metric_card(
            "Tageskurs-Anpassung nach TFS", tageskurs_kapinv_corr, "info")

    with st.expander("Summen und Kontrollwerte", expanded=False):
        st.caption(
            "Werte zur rechnerischen Abstimmung; die zu übertragenden "
            "Steuerzeilen stehen in der Eintragungsübersicht."
        )
        st.markdown(metric_grid(
            metric_card("Fonds-Gewinne vor Tageskurs", etf_gain_raw, "info"),
            metric_card("Fonds-Verluste vor Tageskurs", etf_loss_raw, "info"),
            metric_card(
                "Teilfreistellungseffekt vor Tageskurs",
                etf_gain_raw - etf_gain_taxable
                + etf_loss_raw - etf_loss_taxable,
                "info"),
            metric_card("Veräußerungs-Kontrollwert nach TFS",
                        kap_inv_sale_control, "info"),
            metric_card("Ausschüttungs-Kontrollwert nach TFS",
                        kap_inv_distribution_control, "info"),
            tk_metric_html,
            wht_metric_html,
        ), unsafe_allow_html=True)

    wht_review_items = vm['wht_review_items'] or []
    if dba_wht_beta_enabled and wht_review_items:
        unverified_sum = sum(
            e.get('creditable_tax_eur', 0) for e in wht_review_items
            if e.get('status') == 'dba_unverified')
        if abs(unverified_sum) > 0.01:
            st.warning(
                f"Zeile 41 enthält {fmt_de(unverified_sum)} EUR "
                "Fonds-Quellensteuer ohne hinterlegten DBA-Höchstsatz. Der "
                "Betrag ist nur auf den deutschen 25%-Höchstbetrag begrenzt; "
                "besteht im Quellenstaat ein Erstattungsanspruch, ist er "
                "entsprechend zu kürzen."
            )
        with st.expander(
                f"Quellensteuer-Prüffälle ({len(wht_review_items)})",
                expanded=False):
            st.caption(
                "DBA-Sätze werden nicht aus dem ISIN-Länderpräfix geraten. "
                "Unbelegte DBA-Fälle und zeitversetzte Erstattungen bleiben "
                "sichtbar. Buchung = Buchungsdatum (bestimmt das "
                "Steuerjahr); Bezugsdatum = Datum der zugrunde liegenden "
                "Ausschüttung."
            )
            review_rows = calculate_tax_report.build_wht_review_rows(
                wht_review_items, etf_by_isin)
            review_groups = {}
            for row in review_rows:
                group_key = (row['product'], row['booking_date'],
                             row['status_label'])
                group = review_groups.setdefault(group_key, {
                    'count': 0, 'sum': 0.0, 'name': row['name'],
                })
                group['count'] += 1
                group['sum'] += row['net_foreign_tax_eur']
            for (product, booking, status_label), group in sorted(
                    review_groups.items()):
                name_part = f" ({group['name']})" if group['name'] else ""
                plural = "Ereignisse" if group['count'] != 1 else "Ereignis"
                st.markdown(
                    f"**{product}**{name_part}: {group['count']} {plural}, "
                    f"gebucht am {booking}, Netto-QSt gesamt "
                    f"{fmt_de(group['sum'])} EUR. Status: {status_label}."
                )
            review_df_rows = []
            for row in review_rows:
                treaty_cap = row['treaty_cap_eur']
                review_df_rows.append({
                    'Produkt': row['product'],
                    'Buchung': row['booking_date'],
                    'Bezugsdatum': row['entitlement_date'],
                    'Netto-QSt (+ Einbehalt / − Erstattung)':
                        fmt_de(row['net_foreign_tax_eur']),
                    'DE-Höchstbetrag': fmt_de(row['german_cap_eur']),
                    'DBA-Limit': (fmt_de(treaty_cap)
                                  if treaty_cap is not None else 'prüfen'),
                    'Anrechenbar': fmt_de(row['creditable_tax_eur']),
                    'Status': row['status_label'],
                })
            st.dataframe(review_df_rows, use_container_width=True,
                         hide_index=True)
            st.caption("Alle Beträge in EUR.")

    with st.expander("Fondsdetails nach ISIN · KAP-INV-Abstimmung"):
        form_tab, calculation_tab = st.tabs(["Formularwerte", "Prüfrechnung"])
        sorted_fund_details = sorted(
            kap_inv_form.get('details', []),
            key=lambda item: item.get('ticker', ''),
        )
        with form_tab:
            form_fund_table = (
                "| Produkt | ISIN | Fondsart / TFS | Ausschüttungen | "
                "Veräußerung |\n"
                "|---------|------|----------------|---------------:|"
                "------------:|\n")
            for detail in sorted_fund_details:
                fund_type = detail.get(
                    'fund_type', detail.get('classification', '?'))
                if fund_type == 'Sonstige Investmentfonds':
                    fund_type = 'Sonstiger Fonds'
                form_fund_table += (
                    f"| {detail.get('ticker', detail['isin'])} | "
                    f"{detail['isin']} | {fund_type} "
                    f"({detail.get('tfs_rate', 0) * 100:.0f}%) | "
                    f"{fmt_de(detail.get('distribution_raw_eur', 0))} | "
                    f"{fmt_de(detail.get('sale_raw_eur', 0))} |\n")
            st.markdown(form_fund_table)
            st.caption(
                "Alle Beträge in EUR vor Teilfreistellung. Ausschüttungen "
                "enthalten nur erhaltene Zahlungen; gezahlte Ersatzzahlungen "
                "auf Short-Positionen stehen separat bei den Prüffällen."
            )
        with calculation_tab:
            calculation_table = (
                "| Produkt | Vor Tageskurs | Tageskurs-Anpassung | "
                "KAP-INV-Wert | Nach TFS* |\n"
                "|---------|---------------:|--------------------:|"
                "-------------:|----------:|\n")
            for detail in sorted_fund_details:
                sale_raw = detail.get('sale_raw_eur', 0)
                tageskurs_raw = detail.get('tageskurs_raw_eur', 0)
                calculation_table += (
                    f"| {detail.get('ticker', detail['isin'])} | "
                    f"{fmt_de(sale_raw - tageskurs_raw)} | "
                    f"{fmt_de(tageskurs_raw)} | {fmt_de(sale_raw)} | "
                    f"{fmt_de(detail.get('sale_taxable_control_eur', 0))} |\n")
            st.markdown(calculation_table)
            st.caption(
                "* Steuerpflichtiger Kontrollwert nach Teilfreistellung, "
                "kein zusätzlicher Formularwert. Der KAP-INV-Wert enthält "
                "die Tageskurs-Anpassung. Durch Rundung je ISIN kann die "
                "sichtbare Summe um einen Cent vom Formularwert abweichen."
            )


# ── Renderer: Anlage SO ──────────────────────────────────────────────────────

def render_anlage_so():
    st.markdown('<p class="page-title">Anlage SO</p>', unsafe_allow_html=True)
    _inline_marker('anlage_so')
    if not has_so_data:
        st.info(
            "Aktuell führt kein Produkt zu Anlage-SO-Werten. Die manuelle "
            "Zuordnung von Gold-ETCs steht im Bereich Prüffälle."
        )
        return

    so_by_isin = anlage_so.get('by_isin', {})
    so_unknown = (abs(anlage_so.get('unknown_gain', 0))
                  + abs(anlage_so.get('unknown_loss', 0)))
    history_hint = ""
    if so_unknown > 0.01:
        history_hint = (
            " Haltedauer nicht ermittelbar: ohne Vorjahres-XMLs oder "
            f"CLOSED_LOT-Daten werden {fmt(so_unknown)} konservativ als "
            "steuerpflichtig behandelt; XML des Kaufjahres als Historie "
            "hochladen."
        )
    st.markdown(notice_html({
        'class': 'transparenz', 'severity': 'normal',
        'title': 'Private Veräußerungsgeschäfte (§23 EStG)',
        'body': (
            "Physische Edelmetall-ETCs mit Lieferanspruch werden nach §23 "
            "Abs. 1 S. 1 Nr. 2 EStG behandelt (BFH VIII R 35/14, "
            "VIII R 4/15). Spekulationsfrist 1 Jahr: Gewinne nach Ablauf "
            "steuerfrei, innerhalb der Frist auf Anlage SO zu erklären "
            "(nicht auf Anlage KAP)." + history_hint
        ),
        'target': None,
    }, show_target=False), unsafe_allow_html=True)

    st.markdown(
        '<div class="card">'
        '<div class="eyebrow" style="margin-top:0;">Anlage SO</div>'
        + kap_row("SO", "Steuerpflichtiger Gewinn/Verlust (bis 1 Jahr)",
                  so_taxable, highlight=True)
        + kap_row("SO", "Steuerfrei (über 1 Jahr Haltedauer)", so_free)
        + kap_row("SO", "Gesamtergebnis", so_total)
        + '</div>',
        unsafe_allow_html=True,
    )

    with st.expander("ETC-Details nach ISIN"):
        so_table = (
            "| Ticker | ISIN | Gesamt | Steuerfrei (>1J) | "
            "Steuerpflichtig (≤1J) |\n"
            "|--------|------|-------:|------------------:|"
            "-----------------------:|\n")
        for isin, info in sorted(so_by_isin.items(),
                                 key=lambda x: abs(x[1]['total']),
                                 reverse=True):
            so_table += (
                f"| {info.get('ticker', isin)} | {isin} | "
                f"{fmt_de(info['total'])} | {fmt_de(info.get('tax_free', 0))} | "
                f"{fmt_de(info.get('taxable', 0))} |\n")
        st.markdown(so_table)
        so_details = anlage_so.get('details', [])
        if so_details:
            st.markdown("**Lot-Details (FIFO-Zuordnung):**")
            lot_table = (
                "| Ticker | Kauf | Verkauf | Stk. | G/V (EUR) | Status |\n"
                "|--------|------|---------|-----:|----------:|--------|\n")
            for lot in so_details:
                status = ("steuerfrei" if lot.get('is_tax_free')
                          else "steuerpflichtig")
                qty = abs(lot.get('quantity', 0))
                lot_table += (
                    f"| {lot['ticker']} | {lot.get('open_date', '?')} | "
                    f"{lot.get('close_date', '?')} | {qty:.0f} | "
                    f"{fmt_de(lot['pnl_eur'])} | {status} |\n")
            st.markdown(lot_table)


# ── Renderer: Prüffälle ──────────────────────────────────────────────────────

def _render_so_override_picker():
    """Manuelle Anlage-SO-Zuordnung (Issue #51). Lebt bewusst im immer
    sichtbaren Prueffall-Bereich: ohne bestehendes SO-Ergebnis waere der
    Anlage-SO-Bereich sonst nie erreichbar."""
    all_traded_etf_isins = d.get('all_traded_etf_isins', [])
    so_lookup = anlage_so.get('by_isin', {})

    def _so_label(isin):
        info = get_etf_info(isin)
        if info:
            name = info.get('name', '')
            return (f"{info['ticker']} - {name} ({isin})" if name
                    else f"{info['ticker']} ({isin})")
        fb = etf_by_isin.get(isin) or so_lookup.get(isin) or {}
        ticker = fb.get('ticker', isin[:12])
        name = fb.get('name', '')
        return f"{ticker} - {name} ({isin})" if name else f"{ticker} ({isin})"

    selectable = sorted(
        isin for isin in all_traded_etf_isins
        if isin
        and not is_anlage_so(isin)
        # Nur ETCs (no_invstg) und unklassifizierte ETFs: InvStG-Fonds
        # gehoeren nach §20 EStG, nicht §23 EStG.
        and get_classification(isin) in ('no_invstg', None)
    )
    # Overrides bereinigen, deren ISIN im aktuellen Datensatz fehlt.
    cleaned = [i for i in _dom.get('anlage_so_overrides', [])
               if i in selectable]
    if cleaned != list(_dom.get('anlage_so_overrides', [])):
        _dom['anlage_so_overrides'] = cleaned
    if not selectable:
        return

    section_title("Anlage SO · manuelle Zuordnung")
    st.caption(
        "Nur für physische Edelmetall-ETCs mit Lieferanspruch (BFH VIII R "
        "35/14, VIII R 4/15). Ausgewählte Produkte werden aus KAP-INV "
        "entfernt und auf Anlage SO (§23 EStG) mit 1-Jahres-Frist "
        "berechnet. Die Auswahl ändert die Berechnung und löst einen "
        "Neulauf aus."
    )
    so_key = f"_ui_so_overrides_{_dataset_id[:12]}"

    def _sync_so():
        _dom['anlage_so_overrides'] = list(st.session_state[so_key])

    st.multiselect(
        "Produkte als Anlage SO (§23 EStG) behandeln",
        options=selectable,
        default=cleaned,
        format_func=_so_label,
        key=so_key,
        on_change=_sync_so,
        placeholder="Produkte auswählen",
    )


def render_prueffaelle():
    st.markdown('<p class="page-title">Prüffälle</p>', unsafe_allow_html=True)
    n = vm['notice_counts']['prueffaelle']
    if n == 0:
        st.markdown(notice_html({
            'class': 'transparenz', 'severity': 'normal',
            'title': 'Keine offenen Prüffälle',
            'body': (
                "Alle Buchungen und Instrumente wurden automatisch "
                "zugeordnet. Transparenzhinweise (Kapitalmaßnahmen, "
                "Symbol-Aliasse, Methodik) stehen im Bereich Rechenwege."
            ),
            'target': None,
        }, show_target=False), unsafe_allow_html=True)
    else:
        noun = "Prüffall" if n == 1 else "Prüffälle"
        st.caption(
            f"{n} {noun}. Kritische Fälle zuerst; jeder Eintrag nennt den "
            "betroffenen Bereich."
        )
        ordered = kritisch_notices + [
            notice for notice in prueffall_notices
            if notice['severity'] != 'kritisch'
        ]
        render_notices(ordered, show_target=True)

    _render_so_override_picker()


# ── Renderer: Rechenwege ─────────────────────────────────────────────────────

def _render_transparency_details():
    """Transparenzhinweise mit Detail-Expandern (Kenntnisnahme, kein
    Handlungsbedarf)."""
    if transparenz_notices:
        render_notices(transparenz_notices, show_target=False)

    occ_rename_matches = audit.get('occ_rename_matches', [])
    if occ_rename_matches:
        split_count = sum(
            m.get('match_type') == 'split' for m in occ_rename_matches)
        adjustment_count = sum(
            m.get('match_type') == 'contract_adjustment'
            for m in occ_rename_matches)
        rename_count = len(occ_rename_matches) - split_count - adjustment_count
        action_parts = []
        if split_count:
            action_parts.append(f"{split_count} Split-Zuordnung(en)")
        if adjustment_count:
            action_parts.append(f"{adjustment_count} Kontraktanpassung(en)")
        if rename_count:
            action_parts.append(f"{rename_count} Serien-Umbenennung(en)")
        with st.expander(
                f"Kapitalmaßnahmen · {' und '.join(action_parts)}",
                expanded=False):
            st.caption(
                "Die veränderte Optionsserie wurde über die stabile "
                "IBKR-Kontraktidentität und ihre FIFO-Kostenbasis dem "
                "ursprünglichen Verkauf zugeordnet, damit die "
                "Stillhalterprämie nur einmal versteuert wird. Falls es "
                "sich wider Erwarten um zwei verschiedene Kontrakte "
                "handelt, die Positionen in den Trade-Details prüfen."
            )
            occ_table = ("| Verkauft | am | Geschlossen als | am | Menge |\n"
                         "|----------|----|-----------------|----|-------|\n")
            for m in occ_rename_matches:
                if m.get('match_type') == 'split':
                    qty_text = (f"{m.get('quantity', 0):g} alte → "
                                f"{m.get('close_quantity', 0):g} neue")
                else:
                    qty_text = f"{m.get('quantity', 0):g} Kontrakt(e)"
                occ_table += (
                    f"| {esc(m['sell_symbol'])} | {esc(m['sell_date'])} | "
                    f"{esc(m['close_symbol'])} | {esc(m['close_date'])} | "
                    f"{esc(qty_text)} |\n")
            st.markdown(occ_table)

    underlying_symbol_aliases = audit.get('underlying_symbol_aliases', {})
    if underlying_symbol_aliases:
        with st.expander("Symbol-Aliasse und Ticker-Umbenennungen",
                         expanded=False):
            st.caption(
                "IBKR führt dieselbe Aktie unter verschiedenen Symbolen "
                "(Handelsplatz-Suffix oder Ticker-Umbenennung); die "
                "Zuordnung von Optionsprämien zu Aktien-Trades läuft über "
                "die stabile Kontraktidentität (conid/ISIN)."
            )
            alias_table = "| Schreibweisen | Kanonisch |\n|---|---|\n"
            for canon, members in sorted(underlying_symbol_aliases.items()):
                alias_table += (
                    f"| {esc(', '.join(sorted(str(m) for m in members)))} | "
                    f"{esc(canon)} |\n")
            st.markdown(alias_table)

    open_short = audit.get('stillhalter_open_short', [])
    if open_short:
        with st.expander(
                f"Offene Short-Positionen aus Andienungen ({len(open_short)})",
                expanded=False):
            st.caption(
                "PnL unrealisiert, keine Korrektur nötig. Beim "
                "Folgejahr-Lauf dieses XML als Historie mitladen."
            )
            for item in open_short:
                st.markdown(f"- {esc(item)}")


def _render_stillhalter_zufluss():
    if not (cross_year_details or zufluss_details or prior_zufluss_details):
        return
    section_title("Stillhalter & Zuflussprinzip")
    st.caption(
        "Alle berücksichtigten Summen sofort sichtbar; die einzelnen "
        "Positionen sind zur Prüfung eingeklappt."
    )
    summary_cards = []
    if zuflussprinzip_aktiv and cross_year_details:
        summary_cards.append(metric_card(
            "Vorjahres-Prämien herausgerechnet", -cross_year_premium, "info"))
    if zufluss_details:
        summary_cards.append(metric_card(
            "Offene Stillhalter · Zufluss", zufluss_premium, "info"))
    if prior_zufluss_details:
        summary_cards.append(metric_card(
            "Vorjahres-Glattstellungen korrigiert",
            -prior_zufluss_correction, "info"))
    if summary_cards:
        st.markdown(metric_grid(*summary_cards), unsafe_allow_html=True)

    if zuflussprinzip_aktiv and cross_year_details:
        with st.expander(
                f"Vorjahres-Prämien · {len(cross_year_details)} Position(en)",
                expanded=False):
            st.caption(
                "Diese Prämien gehören in die Steuererklärung des "
                "jeweiligen Vorjahres und wurden aus dem aktuellen "
                "Steuerjahr herausgerechnet."
            )
            detail_table = (
                "| Symbol | Strike | Verkauf (Zufluss) | Assignment | "
                "Prämie (EUR) |\n"
                "|--------|--------|-------------------|------------|"
                "-------------:|\n")
            for det in cross_year_details:
                detail_table += (
                    f"| {det['symbol']} | {det['strike']} | "
                    f"{det['orig_sell_date']} | {det['assignment_date']} | "
                    f"{fmt_de(det['premium_eur'])} |\n")
            st.markdown(detail_table)
            st.markdown("**Zusammenfassung nach Zuflussjahr:**")
            year_table = ("| Steuerjahr | Prämien-Summe (EUR) | Hinweis |\n"
                          "|:----------:|--------------------:|--------|\n")
            for year in sorted(cross_year_by_year):
                year_table += (
                    f"| {year} | {fmt_de(cross_year_by_year[year])} | "
                    f"In Steuererklärung {year} eintragen |\n")
            st.markdown(year_table)
            st.info(
                f"Zeile 19 wurde im aktuellen Jahr um "
                f"{fmt_de(cross_year_premium)} EUR reduziert."
            )

    if zufluss_details:
        with st.expander(
                f"Offene Stillhalterpositionen · {len(zufluss_details)} Position(en)",
                expanded=False):
            st.caption(
                f"{fmt_de(zufluss_premium)} EUR Prämien wurden im Steuerjahr "
                "vereinnahmt und bereits zu Topf 2 addiert."
            )
            zt = ("| Symbol | Verkaufsdatum | Stk. | Prämie (EUR) |\n"
                  "|--------|--------------|-----:|-------------:|\n")
            for det in zufluss_details:
                zt += (
                    f"| {det['symbol']} | {det['sell_date'][:10]} | "
                    f"{det['quantity']} | {fmt_de(det['premium_eur'])} |\n")
            st.markdown(zt)

    if prior_zufluss_details:
        with st.expander(
                f"Vorjahres-Glattstellungen · {len(prior_zufluss_details)} Korrektur(en)",
                expanded=False):
            st.caption(
                f"{fmt_de(prior_zufluss_correction)} EUR waren bereits im "
                "Verkaufsjahr steuerpflichtig und wurden vom aktuellen PnL "
                "abgezogen."
            )
            pt = ("| Symbol | Verkaufsjahr | Stk. | Korrektur (EUR) |\n"
                  "|--------|:-----------:|-----:|----------------:|\n")
            for det in prior_zufluss_details:
                pt += (
                    f"| {det['symbol']} | {det['sell_year']} | "
                    f"{det['quantity']} | -{fmt_de(det['premium_eur'])} |\n")
            st.markdown(pt)


def _render_toggle_explainers():
    """Rechtsgrundlagen und Wirkung der aktiven Methoden."""
    method_blocks = []
    if cross_year_details or zufluss_details or prior_zufluss_details:
        zufluss_parts = []
        if cross_year_details:
            zufluss_parts.append(
                f"{len(cross_year_details)} Assignment-Prämienanteil(e) aus "
                f"Vorjahren ({fmt_de(cross_year_premium)} EUR)")
        if zufluss_details:
            zufluss_parts.append(
                f"{len(zufluss_details)} offene Stillhalter-Position(en) mit "
                f"Zufluss im Steuerjahr ({fmt_de(zufluss_premium)} EUR, "
                "bereits enthalten)")
        if prior_zufluss_details:
            zufluss_parts.append(
                f"{len(prior_zufluss_details)} Vorjahres-Prämie(n) aus "
                f"Glattstellungen korrigiert "
                f"(-{fmt_de(prior_zufluss_correction)} EUR, bereits "
                "enthalten)")
        method_blocks.append((
            "Zuflussprinzip",
            "**BMF Rn. 25, 33:** " + "; ".join(zufluss_parts) + ".",
            None,
        ))

    if has_etf_data:
        cls_labels = {
            'aktienfonds': 'Aktienfonds (30% TFS)',
            'mischfonds': 'Mischfonds (15% TFS)',
            'immobilienfonds': 'Immobilienfonds (60% TFS)',
            'auslands_immobilienfonds': 'Auslands-Immobilienfonds (80% TFS)',
        }
        cls_counts = {}
        for v in etf_by_isin.values():
            classification = v.get('classification')
            label = ('Fondsart nicht bestätigt' if classification is None
                     else cls_labels.get(classification,
                                         'sonstige Fonds (0% TFS)'))
            cls_counts[label] = cls_counts.get(label, 0) + 1
        cls_summary = ", ".join(
            f"{n} {label}" for label, n in sorted(cls_counts.items()))
        etf_tickers = ", ".join(sorted(
            v.get('ticker', '?') for v in etf_by_isin.values()))
        method_blocks.append((
            "InvStG-Klassifizierung",
            f"**§2 InvStG:** {len(etf_by_isin)} Produkte laufen im "
            "Investmentfondspfad. Verifiziert klassifizierte Investmentfonds "
            "werden auf Anlage KAP-INV gemeldet; bei unbekannten ISINs "
            "entsteht erst nach ausdrücklicher Fondsart-Bestätigung eine "
            f"Formularzeile. Davon {cls_summary}.",
            f"Betroffene Fondsprodukte: {esc(etf_tickers)}",
        ))

    if abs(fx_corr_total) > 0.01:
        method_blocks.append((
            "Tageskurs-Methode",
            "**§20 Abs. 4 S. 1 EStG:** Einnahmen werden zum Verkaufskurs und "
            "Anschaffungskosten zum Kaufkurs in Euro umgerechnet. IBKR rechnet "
            "den gesamten Netto-PnL zum Schlusskurs um. Abweichung für "
            f"{steuerjahr}: **{'+' if fx_corr_total >= 0 else ''}"
            f"{fmt_de(fx_corr_total)} EUR** (CLOSED_LOT-Analyse, ohne Futures).",
            "Futures bleiben ausgeschlossen: Ihre Kostenbasis ist der volle "
            "Kontraktwert, nicht die gezahlte Margin; eine FX-Korrektur auf "
            "den Notional würde Phantom-Gewinne oder -Verluste erzeugen.",
        ))

    if not method_blocks:
        return

    section_title("Aktive Methoden im Report")
    for title, body, caption in method_blocks:
        st.markdown(f"### {title}")
        st.markdown(body)
        if caption:
            st.caption(caption)


def _render_catalogs():
    section_title("Produktzuordnungen")
    st.caption(
        "Nachvollziehbare Steuerpfade für alle hinterlegten ETF-, Fonds- "
        "und ETP-Zuordnungen."
    )
    traded_product_isins = d.get('all_traded_etf_isins', []) or []
    if traded_product_isins:
        report_catalog_rows = get_classification_catalog(traded_product_isins)
        report_kap_inv_count = sum(
            row.get('tax_route') == 'Anlage KAP-INV'
            for row in report_catalog_rows)
        st.markdown("### Produkte aus dem Upload · inklusive Vorjahreshistorie")
        st.write(
            f"{len(report_catalog_rows)} im Upload erkannte ETF-, "
            "Fonds- und ETP-ISINs, einschließlich Produkten aus der "
            "Vorjahreshistorie. Zuordnungs- und Transparenzkatalog, "
            "keine Zählung der im Steuerjahr betroffenen Positionen. "
            f"Davon laufen {report_kap_inv_count} über Anlage KAP-INV "
            f"und {len(report_catalog_rows) - report_kap_inv_count} "
            "über einen anderen Steuerpfad. Unbekannte ISINs bleiben "
            "ausdrücklich unklassifiziert."
        )
        render_classification_catalog(
            report_catalog_rows,
            key_prefix="report_classification_catalog",
            show_filters=False,
            offer_download=False,
        )
    full_catalog = get_classification_catalog()
    with st.expander(
            f"Gesamtkatalog · alle hinterlegten Zuordnungen "
            f"({len(full_catalog)})",
            expanded=False):
        st.write(
            "Der Gesamtkatalog trennt produktspezifisch geprüfte "
            "Entscheidungen von festen, aktiv berechneten "
            "Katalogzuordnungen. Beide werden angewandt; nur unbekannte ISINs "
            "bleiben bis zur Bestätigung unklassifiziert."
        )
        render_classification_catalog(
            full_catalog,
            key_prefix="global_classification_catalog",
            show_filters=True,
            offer_download=True,
        )


def _render_plausibility():
    if not plaus:
        return
    section_title("Plausibilitätscheck (IBKR-Bericht vs. Berechnung)")
    check_table = (
        "| Kategorie | IBKR-Bericht | Unsere Berechnung | Differenz |\n"
        "|-----------|-------------|-------------------|----------|\n")
    for row in plaus['rows']:
        if row['match']:
            icon = ""
        elif row['label'] == "Zinsen":
            icon = " **(FX)**"
        elif row['is_fx_saldo']:
            icon = " **(FX-Saldo)**"
        else:
            icon = " **(!)**"
        check_table += (
            f"| {row['label']} | {fmt_de(row['ibkr'])} | "
            f"{fmt_de(row['ours'])} | {fmt_de(row['diff'])}{icon} |\n")
    st.markdown(check_table)
    if plaus['all_match'] and not (plaus['zinsen_fx_diff']
                                   or plaus['fx_saldo_diff']):
        st.success("Alle Kategorien stimmen mit dem IBKR-Bericht überein.")
    elif plaus['all_match']:
        explanations = []
        if plaus['zinsen_fx_diff']:
            explanations.append(
                "Zinsen-Differenz ist eine bekannte FX-Konvertierungs"
                "differenz (IBKR konvertiert Fremdwährungs-Anleiheposten im "
                "CSV mit anderen Kursen als in der XML-BaseCurrency-Ansicht)")
        if plaus['fx_saldo_diff']:
            explanations.append(
                "FX-Differenz ist die aktivierte Saldo-Korrektur (§20 Abs. 2 "
                "S. 1 Nr. 7 EStG, BMF Rn. 131; IBKR kennt keine "
                "Margin-Schuld-Unterscheidung)")
        st.success("Alle Kategorien stimmen überein. "
                   + "; ".join(explanations) + ".")
    else:
        st.info(
            "Kleine Abweichungen sind normal (FX-Rundung, Steuerkorrekturen "
            "aus Vorjahren)."
        )
    if has_etf_data and invstg_aktiv:
        st.caption(
            "InvStG aktiv: ETF-Werte wurden für den Vergleich zurückaddiert, "
            "da der IBKR-Bericht keine InvStG-Trennung kennt."
        )
    if has_so_data:
        st.caption(
            "Anlage SO aktiv: Gold-ETC-Werte wurden für den Vergleich "
            "zurückaddiert, da IBKR sie als Aktien zählt."
        )
    if tageskurs_aktiv or plaus['zufluss_adj'] != 0 or plaus['fx_margin_relevant']:
        notes = []
        if tageskurs_aktiv:
            corr_sign = "+" if fx_corr_total >= 0 else ""
            notes.append(
                f"Tageskurs-Korrektur ({corr_sign}{fmt_de(fx_corr_total)} EUR)")
        if plaus['zufluss_adj'] != 0:
            notes.append(
                f"Stillhalter-Zufluss "
                f"({'+' if plaus['zufluss_adj'] >= 0 else ''}"
                f"{fmt_de(plaus['zufluss_adj'])} EUR)")
        if plaus['fx_margin_relevant']:
            notes.append(
                f"FX-Saldo-Korrektur "
                f"({'+' if plaus['fx_margin_diff'] >= 0 else ''}"
                f"{fmt_de(plaus['fx_margin_diff'])} EUR)")
        st.caption(
            "Der Plausibilitätscheck vergleicht die Berechnung 1:1 gegen "
            "IBKRs eigene Summen. Steuerliche Korrekturen über IBKRs Zahlen "
            "hinaus werden dabei herausgerechnet: " + " und ".join(notes)
            + ". So lässt sich prüfen, ob die Basisdaten korrekt verarbeitet "
            "wurden, bevor die steuerlichen Anpassungen aufsetzen."
        )


def _render_diagnostics():
    section_title("Berechnungs-Diagnose")
    st.caption(
        "Nicht-sensitive Metadaten des Berechnungslaufs; Cache-Zustand "
        "gilt für den aktuellen Seitenaufbau."
    )
    diag_table = (
        "| Merkmal | Wert |\n|---|---|\n"
        f"| Datensatz | {esc(_dataset_id[:16])}… |\n"
        f"| Berechnungslauf | {esc(_snapshot['computed_at'])}, "
        f"{_snapshot['duration_s']} s |\n"
        f"| Schema-Version | {_snapshot['schema_version']} |\n"
        f"| Generation | {_snapshot['generation']} |\n"
        f"| Cache | {'Treffer (keine Neuberechnung)' if _cache_hit else 'Neuberechnung in diesem Lauf'} |\n"
        f"| Unterdrückte Log-Zeilen | {_snapshot.get('suppressed_log_lines', 0)} |\n"
        f"| FX-Saldo-Korrektur | {'aktiv' if _dom['toggles'].get('fx_margin', True) else 'deaktiviert'} |\n"
        f"| DBA-Beta | {'aktiv' if _dom['toggles'].get('dba_beta') else 'deaktiviert'} |\n"
    )
    st.markdown(diag_table)


def _render_legal():
    section_title("Rechtliche Hinweise")
    with st.container():
        st.markdown(f"""
**Eigenverantwortliche Nutzung.** Dieses Tool dient ausschließlich zur Unterstützung
bei der Erstellung der Einkommensteuererklärung. Die berechneten Werte sind
unverbindlich und ohne Gewähr für Richtigkeit, Vollständigkeit oder
Aktualität. Alle Ergebnisse und Angaben sind vor der Übernahme in die
Steuererklärung eigenverantwortlich zu prüfen.

**Keine Steuerberatung.** Dieses Tool stellt keine Steuerberatung im Sinne
des Steuerberatungsgesetzes (StBerG) dar und ersetzt nicht die Beratung
durch einen Steuerberater, Wirtschaftsprüfer oder eine andere zur
Steuerberatung befugte Person. Bei Unsicherheiten oder komplexen
Sachverhalten ist eine steuerliche Beratung hinzuzuziehen.

**Haftungsbeschränkung.** Soweit gesetzlich zulässig, ist die Haftung für
Schäden aus der Nutzung oder Nichtnutzung des Tools sowie aus fehlerhaften,
unvollständigen oder nicht aktuellen Berechnungsergebnissen ausgeschlossen.
Die Haftungsbeschränkung gilt nicht bei Vorsatz, grober Fahrlässigkeit, bei
Schäden aus der Verletzung von Leben, Körper oder Gesundheit sowie in
sonstigen Fällen zwingender gesetzlicher Haftung. Bei einer leicht
fahrlässigen Verletzung wesentlicher Pflichten ist die Haftung auf den
typischerweise vorhersehbaren Schaden begrenzt.

**Datenschutz und Datenverarbeitung.** Sämtliche Berechnungen erfolgen in
der lokal gestarteten Anwendung auf dem eigenen Rechner. Hochgeladene Dateien
werden nur an den lokalen Streamlit-Prozess (`localhost`) übertragen und
dort temporär verarbeitet; das Tool sendet sie nicht an externe Server oder
Dritte und speichert sie nicht dauerhaft. Im Tool findet kein Tracking und
keine Nutzungsanalyse statt.

**Rechtsstand und Aktualität.** Dieser Bericht wurde für das Steuerjahr
{int(steuerjahr)} berechnet. Berücksichtigter Rechtsstand: §20 EStG, das
BMF-Schreiben vom 14.05.2025 (Einzelfragen zur Abgeltungsteuer) und das
Jahressteuergesetz 2024, jeweils soweit für das ausgewählte Steuerjahr
anwendbar und einschließlich rückwirkender Änderungen. Spätere Änderungen
der Rechtslage, Verwaltungsauffassung oder Rechtsprechung werden nicht
automatisch berücksichtigt.

**Open Source.** Dieses Projekt ist unter der MIT-Lizenz veröffentlicht.
Der Quellcode ist frei einsehbar und prüfbar unter
[github.com/KonvexInvestment/ibkr-steuer](https://github.com/KonvexInvestment/ibkr-steuer).
""")


def render_rechenwege():
    st.markdown('<p class="page-title">Rechenwege</p>', unsafe_allow_html=True)
    st.caption(
        "Methoden, Produktzuordnungen und Nachweise sind in vier Themen "
        "gegliedert. Alles hier ist Kenntnisnahme; Handlungsbedarf steht im "
        "Bereich Prüffälle."
    )
    methods_tab, products_tab, calculation_tab, legal_tab = st.tabs([
        "Methoden im Report",
        "Produktzuordnungen",
        "Berechnung & Diagnose",
        "Rechtliches",
    ])
    with methods_tab:
        _render_toggle_explainers()
        _render_stillhalter_zufluss()
        _render_transparency_details()
    with products_tab:
        _render_catalogs()
    with calculation_tab:
        tax_tab, processing_tab, diagnostics_tab = st.tabs([
            "Steuerlogik",
            "XML & Verarbeitung",
            "Diagnose",
        ])
        with tax_tab:
            _render_tax_rules()
        with processing_tab:
            _render_processing_method()
        with diagnostics_tab:
            _render_plausibility()
            _render_diagnostics()
    with legal_tab:
        _render_legal()


def _render_tax_rules():
    section_title("Steuerliche Regeln & Berechnungsmethodik")
    sh_count = audit.get("stillhalter_count", 0)
    sh_eur = audit.get("stillhalter_premium_eur", 0)
    base_curr = d.get("base_currency", "USD")
    with st.container():
        st.markdown(f"""
### Zwei-Töpfe-Struktur (§20 Abs. 6 EStG)

Das deutsche Steuerrecht unterscheidet zwei getrennte Verrechnungstöpfe:

| Topf | Inhalt | Verlustverrechnung |
|------|--------|-------------------|
| **Topf 1 - Aktien** | Gewinne und Verluste aus Aktienveräußerungen (STK) | Aktienverluste dürfen **nur** mit Aktiengewinnen verrechnet werden. Überschüsse werden als Verlustvortrag ins nächste Jahr übertragen. |
| **Topf 2 - Sonstiges** | Optionen, Futures, Anleihen, T-Bills, Dividenden, Zinsen, Stillhalterprämien | Alle Verluste frei mit allen Gewinnen in Topf 2 verrechenbar. |

Das Finanzamt wendet die Verlustverrechnungsbeschränkung anhand der Zeilen 20 und 23 an - der Steuerpflichtige meldet die Bruttowerte.

---

### Instrument-Klassifikation

| IBKR-Kategorie | Steuerliche Einordnung | Topf |
|----------------|----------------------|------|
| **STK** (Aktien) | Aktienveräußerung (§20 Abs. 2 Nr. 1) | Topf 1 |
| **OPT** (Optionen) | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **FUT** (Futures) | Termingeschäft/Festgeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **FOP** (Futures-Optionen) | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **FSFOP** (Future-Style-Optionen) | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **BILL** (T-Bills) | Kapitalforderung (§20 Abs. 2 Nr. 7) | Topf 2 |
| **BOND** (Anleihen) | Kapitalforderung (§20 Abs. 2 Nr. 7) | Topf 2 |
| **WAR** (Optionsscheine) | Verbriefte Kapitalforderung (§20 Abs. 2 Nr. 7, BMF Rn. 8 f.), **kein** Termingeschäft | Topf 2 |
| **CFD** | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **DIV/PIL** (Dividenden) | Laufende Erträge (§20 Abs. 1 Nr. 1) | Topf 2 |
| **INTR/CINT** (Zinsen) | Zinserträge (§20 Abs. 1 Nr. 7) | Topf 2 |
| **CASH/FOREX** (Fremdwährung) | Verzinsl. Fremdwährungs**guthaben** (§20 Abs. 2 Nr. 7, BMF Rn. 131). Tilgung einer Fremdwährungs-Schuld zählt nicht. | Topf 2 |
| **DINT** (Sollzinsen) | Werbungskosten, durch den Sparer-Pauschbetrag abgegolten (§20 Abs. 9) | nachrichtlich, keine Zeile |
| **OFEE/STAX** (Gebühren, Umsatzsteuer) | Nicht abziehbar (§20 Abs. 9) | nachrichtlich, keine Zeile |

---

### Fremdwährungs-Gewinne/Verluste (BMF Rn. 131)

Beim Halten von Fremdwährungsguthaben (z.B. USD) auf einem verzinslichen Konto (IBKR zahlt Zinsen) entstehen bei Kursänderungen steuerlich relevante Gewinne oder Verluste:

- **Anschaffung** = jeder Zufluss von Fremdwährung (Kauf, Dividende, Verkaufserlös)
- **Veräußerung** = jeder Abfluss, der ein Guthaben auflöst (Rücktausch, Aktienkauf, Gebühren)
- **Auslegung des Tools für Margin-Schulden:** Ein Abfluss bei bereits negativem Saldo vertieft die Schuld; ein Zufluss kann sie tilgen. Beides löst kein Guthaben auf. BMF Rn. 131 knüpft an ein Fremdwährungs**guthaben** bzw. eine Kapital**forderung** an und regelt die Verbindlichkeit nicht ausdrücklich. Diese konservative Auslegung ist über die Checkbox "FX-Saldo-Korrektur" in der Sidebar unter Berechnung abschaltbar.
- **FIFO-Methode**: die zuerst erworbenen Beträge werden zuerst veräußert
- **Rechtsgrundlage**: §20 Abs. 2 S. 1 Nr. 7 EStG, Anlage KAP, Topf 2

**Datenquellen:** Enthält die Flex Query `<FxTransactions>`, verwendet das Tool IBKRs FIFO-Ergebnis pro Buchung und kann Schuldtilgungen einzeln prüfen. Bei EUR-Basiskonten dient andernfalls ein hochgeladener IBKR-Standardbericht als aggregierter FIFO-Rohwert (CSV-Upload beim Start oder über "Daten ändern" in der Sidebar; nur bei einem einzelnen Konto aktiv); einzelne Schuldtilgungen sind darin nicht erkennbar. Fehlen beide Quellen, rechnet das Tool bei EUR-Basiskonten selbst eine FIFO-Näherung aus den Kontobewegungen. Vorjahres-XMLs vervollständigen dabei die Lot-Historie, beseitigen aber nicht die Kursnäherung. Ohne Vorjahres-XMLs wird zusätzlich der Jahresanfangsbestand vereinfachend zum 01.01.-Kurs angesetzt. Für USD-Basiskonten ist ohne `<FxTransactions>` weder der CSV- noch der FIFO-Fallback verfügbar.

---

### Stillhalterprämien bei Assignments (BMF Rn. 25–35)

Wird eine verkaufte Option (Stillhalterposition) ausgeübt (Assignment), muss die Prämie steuerlich korrekt zugeordnet werden:

- **Prämie** = laufende Einnahmen nach §20 Abs. 1 Nr. 11 → gehört in **Topf 2**
- **Aktientransaktion** = Veräußerung (Call, Rn. 26) bzw. Anschaffung (Put, Rn. 33) nach §20 Abs. 2 → gehört in **Topf 1**

Bei **beiden** Assignment-Typen gilt laut BMF: „Die vereinnahmte Optionsprämie wird bei der Ermittlung des Veräußerungsgewinns **nicht berücksichtigt**." IBKR bündelt die Prämie jedoch im Aktien-Trade (Call: im Verkaufserlös, Put: in den reduzierten Anschaffungskosten). Dieses Tool erkennt Assignments automatisch und trennt die Prämie heraus.

{"**In diesem Report:** " + str(sh_count) + " Assignments erkannt (Call + Put), " + fmt_de(sh_eur) + " EUR Stillhalterprämien von Topf 1 nach Topf 2 verschoben." if sh_count > 0 else "**In diesem Report:** Keine Assignments erkannt."}

---

### Dividenden & Payment in Lieu (PIL)

- **Dividenden** (DIV): Laufende Erträge in Topf 2
- **Deutsche Dividenden mit `- DE Steuer`**: werden separat als Kapitalerträge mit inländischem Steuerabzug behandelt (Zeile 7), nicht als ausländische Kapitalerträge in Zeile 19. Liegt die Buchung auf einem deutschen Investmentfonds, wird die einbehaltene Steuer als Prüffall gemeldet statt automatisch zugeordnet (siehe Quellensteuer-Abschnitt)
- **Payment in Lieu** (PIL): Ersatzzahlung wenn Aktien verliehen sind, wird steuerlich wie eine Dividende behandelt und mit diesen zusammen verrechnet

---

### Zinsen & Stückzinsen

- **INTR/CINT**: Zins- und Couponerträge aus Anleihen → Topf 2
- **INTP** (Stückzinsen): Beim Kauf einer Anleihe gezahlte aufgelaufene Zinsen sind **negative Einnahmen** (BMF Rn. 51). Sie reduzieren den Zinsertrag und können diesen insgesamt negativ werden lassen.
- **DINT** (Sollzinsen, Leihgebühren, CFD-Finanzierung): werden **nicht** von den Zinserträgen abgezogen. Es sind Werbungskosten, und die sind nach §20 Abs. 9 EStG durch den Sparer-Pauschbetrag abgegolten. Der Betrag wird nur nachrichtlich ausgewiesen.

---

### Quellensteuer (Zeile 41)

Ausländische Quellensteuern auf Dividenden und Zinsen (z.B. 15% US-Quellensteuer) werden in Zeile 41 als **anrechenbare ausländische Steuern** gemeldet. Zeile 41 setzt sich aus zwei Teilen zusammen: der Quellensteuer außerhalb der Fonds und der anrechenbaren Fonds-Quellensteuer aus KAP-INV. Damit steht die Fonds-Quellensteuer genau einmal im Formular; KAP-INV hat keine eigene Quellensteuer-Zeile.

Deutsche Dividendensteuer aus Buchungen mit `- DE Steuer` wird dagegen in Kapitalertragsteuer (Zeile 37) und Solidaritätszuschlag (Zeile 38) aufgeteilt. Wenn das Steuerprogramm diese Zeilen ohne Steuerbescheinigung nach §45a EStG sperrt, bietet "Variante B" eine technische Ersatzdarstellung über Zeile 19 bzw. 41 (Checkbox im Bereich Anlage KAP). Sie ist kein amtlich belegter Ersatz für die Steuerbescheinigung und sollte vor der Abgabe mit Finanzamt oder Steuerberatung abgestimmt werden.

Sonderfall deutscher Investmentfonds: Behält IBKR deutsche Kapitalertragsteuer auf einem DE-Fonds ein, wird sie weder in Zeile 41 angerechnet noch automatisch in Zeile 37/38 eingetragen. §32d Abs. 5 EStG erfasst nur ausländische Steuern, und die auszahlende Stelle berücksichtigt die Teilfreistellung bereits beim Steuerabzug (§43a Abs. 2 EStG). Der Betrag erscheint als Prüffall ("DE-Steuer auf Fonds") und muss anhand der IBKR-Abrechnung manuell zugeordnet werden.

---

### Zeilen-Zuordnung Anlage KAP

Da Interactive Brokers ein **ausländischer Broker ohne inländischen Steuerabzug** ist, werden die meisten Einkünfte in der Sektion "Kapitalerträge, die **nicht** dem inländischen Steuerabzug unterlegen haben" (Zeilen 18-23) eingetragen. Ausnahme: deutsche Dividenden mit separat gebuchter `- DE Steuer`.

| Zeile | Bedeutung | Berechnung |
|-------|-----------|------------|
| **7** | Kapitalerträge mit inländischem Steuerabzug | DE-Dividenden mit passender `- DE Steuer`-Buchung |
| **19** | Ausländische Kapitalerträge (Netto) | Topf 1 + Topf 2 (Summe aller Erträge und Verluste) |
| **20** | Davon: Aktiengewinne | Brutto-Aktiengewinne (ohne Verluste) |
| **22** | Verluste ohne Aktien | Verluste aus Optionen, Futures, Anleihen etc. (positiver Betrag) |
| **23** | Aktienverluste | Verluste aus Aktienveräußerungen (positiver Betrag) |
| **37/38** | Kapitalertragsteuer / Soli | Aufteilung deutscher Dividendensteuer (25% + 5,5% Soli) |
| **41** | Anrechenbare Quellensteuer | Quellensteuer außerhalb der Fonds + anrechenbare Fonds-Quellensteuer aus KAP-INV |

---

### Währungsumrechnung

{"**Das Konto hat EUR als Basiswährung.** Alle Beträge in der IBKR-Abrechnung sind bereits in EUR umgerechnet. Bei USD-Trades nutzt IBKR den Tageskurs (`fxRateToBase`), der direkt in EUR umrechnet, kein zusätzlicher FX-Lookup erforderlich." if base_curr == "EUR" else "**Das Konto hat USD als Basiswährung.** Beträge werden in zwei Schritten umgerechnet: (1) Trade-Währung → USD über `fxRateToBase`, (2) USD → EUR über den Tageskurs des vorherigen Geschäftstags. Die täglichen USD/EUR-Kurse werden aus den IBKR-Daten extrahiert."}

---

### Rechtsgrundlagen

- **§20 EStG** - Einkünfte aus Kapitalvermögen
- **BMF-Schreiben vom 14.05.2025** - "Einzelfragen zur Abgeltungsteuer" (IV C 1 - S 2252/00075/016/070)
- **Jahressteuergesetz 2024** - Abschaffung des €20.000-Caps für Termingeschäfteverluste (§20 Abs. 6 Satz 5 EStG), rückwirkend für alle offenen Fälle
- **Anlage KAP** - Zeilen 9/14 (Termingeschäfte) existieren nur in der Sektion mit inländischem Steuerabzug und sind für IBKR nicht relevant
""")

def _render_processing_method():
    base_curr = d.get("base_currency", "USD")
    section_title("XML-Verarbeitung und Rechenlogik")
    with st.container():
        st.markdown(f"""
### Schritt 1: XML-Extraktion

Die IBKR Flex Query XML wird in einzelne CSV-Dateien zerlegt. Jede XML-Sektion enthält spezifische Daten:

| XML-Sektion | Inhalt | Filter |
|---|---|---|
| `<Trades>` | Alle Trades. Felder: `assetCategory`, `fifoPnlRealized`, `fxRateToBase`, `reportDate`, `buySell`, `transactionType` | `EXECUTION` → trades.csv, `CLOSED_LOT` → closed_lots.csv (für Tageskurs-Korrektur) |
| `<StmtFunds>` | Dividenden, Zinsen, Steuern, Gebühren. Felder: `activityCode`, `amount`, `fxRateToBase`, `reportDate`, `transactionID` | Bei einer einzelnen XML-Datei vollständig übernommen; Split-/Quartals-XMLs werden bereits beim Merge dedupliziert. Die Berechnung dedupliziert anschließend nochmals defensiv (Schritt 2) |
| `<FIFOPerformanceSummaryInBase>` | Aggregierter PnL pro Instrument. Felder: `assetCategory`, `isin`, `totalRealizedPnl` | Fallback für fehlende Trades (z.B. T-Bill Maturity) |
| `<FxTransactions>` | FX-Gewinne/-Verluste. Felder: `fxCurrency`, `realizedPL`, `reportDate` | Nur `levelOfDetail=TRANSACTION` |
| `<AccountInformation>` | Basiswährung (`currency`), Kontotyp | Einzelner Eintrag |
| `<FlexStatement>` | Berichtszeitraum → Steuerjahr aus `toDate` | Automatisch erkannt |

**Multi-XML (Vorjahre):** Trades aus allen XMLs werden in eine gemeinsame `trades.csv` zusammengeführt (für Stillhalter-Matching über Jahresgrenzen). FX-Kontobewegungen werden chronologisch gemergt; als Duplikat gilt dabei eine Zeile, bei der alle sechs Schlüsselfelder übereinstimmen: Währung, Datum, `transactionID`, Buchungstext, Betrag und Saldo. Ein Schlüssel allein aus `transactionID` wäre falsch: IBKR vergibt dieselbe ID für jede Folgebuchung derselben Position, etwa für alle täglichen Abrechnungen eines Futures.

---

### Schritt 2: Deduplizierung

IBKR liefert in einigen Sektionen Duplikate:

| Quelle | Duplikat-Ursache | Deduplizierungs-Schlüssel |
|---|---|---|
| **Trades** | Erweiterte Flex Queries enthalten ORDER + EXECUTION für denselben Trade | `tradeID` (wenn vorhanden) oder `(dateTime, isin, buySell, quantity, closePrice, fifoPnlRealized)` |
| **StmtFunds** | Mehrfachansichten und überlappende Exporte können dieselbe Buchung wiederholen | `(transactionID, activityDescription)`. Der Buchungstext gehört zum Schlüssel, weil IBKR unter einer `transactionID` mehrere verschiedene Vorgänge bündeln kann. Ohne `transactionID` wird die vollständige CSV-Zeile verglichen |

---

### Schritt 3: Kapitalgewinne berechnen

Für jeden Trade im Steuerjahr (`reportDate.year == Steuerjahr`):

```
IBKR-Methode:     PnL (EUR) = fifoPnlRealized × fxRateToBase
Tageskurs-Methode: PnL (EUR) = Erlös × FX_Verkaufstag − AK × FX_Kauftag
```

**IBKR-Methode (Standard):** Rechnet den Netto-PnL komplett zum Schlusskurs um.

**Tageskurs-Methode (§20 Abs. 4 S. 1 EStG, optional):** *"Bei nicht in Euro getätigten Geschäften sind die Einnahmen im Zeitpunkt der Veräußerung und die Anschaffungskosten im Zeitpunkt der Anschaffung in Euro umzurechnen."* Verwendet CLOSED_LOT Daten aus Extended Flex Queries. Futures werden ausgeschlossen (Kostenbasis = Notional, kein realer Cashflow). Korrektur: `|AK| × (FX_Schlusskurs - FX_Kaufkurs)` pro Lot. IBKR vergibt pro Tag zwei `fxRateToBase`-Kurse: einen Intraday-Kurs (ExchTrades) und einen Settlement-Kurs (BookTrades, 16:20). Für den Kaufkurs wird der ExchTrade-Kurs bevorzugt; an reinen Verfall-/Andienungstagen der BookTrade-Kurs als Fallback.

| Feld | Bedeutung |
|---|---|
| `fifoPnlRealized` | IBKR's FIFO-basierter realisierter Gewinn/Verlust in **Trade-Währung** |
| `fxRateToBase` | Umrechnungskurs Trade-Währung → Basiswährung (EUR) am **Schlusstag** |
| `reportDate` | Buchungsdatum (bestimmt das Steuerjahr, Zuflussprinzip) |
| `assetCategory` | Ausgangspunkt der Routingtabelle unten: `STK` wird weiter nach Produktart unterschieden; bekannte Derivate und Kapitalforderungen gehen in Topf 2, `CASH` in die separate FX-Rechnung. Unbekannte Kategorien werden als Prüffall gemeldet |
| `subCategory` | ETF-Erkennung: `ETF` → InvStG-Prüfung, `COMMON` → Einzelaktie |

**Topf-Zuordnung:**

| `assetCategory` | `subCategory` | Steuerliche Einordnung | Topf |
|---|---|---|---|
| `STK` | `COMMON` / `REIT` / `ADR` | Aktienveräußerung (§20 Abs. 2 Nr. 1) | **Topf 1** |
| `STK` | `ETF` (InvStG-Fonds) | Investmentfonds (InvStG §2) | **KAP-INV** (optional) |
| `STK` | `ETF` (no\\_invstg, z.B. VXX/FNGU-ETNs) | Schuldverschreibung, kein Investmentfonds i.S.d. InvStG | **Topf 2** |
| `STK` | `ETF` (Personengesellschaft, z.B. USO/UNG) | §1 Abs. 3 Nr. 2 InvStG; Besteuerung nach anteiliger Jahresallokation | **blockiert**, bis K-1/K-3 bzw. äquivalenter Nachweis vorliegt |
| `OPT` | | Termingeschäft, Option (§20 Abs. 2 Nr. 3) | Topf 2 |
| `FUT` | | Termingeschäft, Future (§20 Abs. 2 Nr. 3) | Topf 2 |
| `FOP` / `FSFOP` | | Termingeschäft, Future-Option (§20 Abs. 2 Nr. 3) | Topf 2 |
| `BILL` | | Kapitalforderung, T-Bill (§20 Abs. 2 Nr. 7) | Topf 2 |
| `BOND` | | Kapitalforderung, Anleihe (§20 Abs. 2 Nr. 7) | Topf 2 |
| `WAR` | | Optionsschein, verbriefte Kapitalforderung (§20 Abs. 2 Nr. 7) | Topf 2 |
| `CFD` | | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| `CASH` | | Devisenumsatz; wird von der FX-Rechnung erfasst, hier keine zweite Zuordnung | keine |

Kategorien außerhalb dieser Tabelle werden nicht stillschweigend verworfen: Taucht eine unbekannte `assetCategory` mit einem Ergebnis auf, meldet das Tool sie als Prüffall.

**InvStG-Klassifizierung (optional):** ETFs mit `subCategory="ETF"` werden gegen die belegte Produkttabelle geprüft. Maßgeblich ist §1 Abs. 2 InvStG i.V.m. dem Investmentvermögensbegriff des §1 Abs. 1 KAGB; das geltende Recht verlangt keine Risikomischung. Passive Single-Asset-/Grantor-Trusts wie GLD oder IBIT und registrierte Closed-End-Funds werden deshalb nach der hier vertretenen Auffassung als Investmentfonds behandelt; die Einordnung ist nicht höchstrichterlich geklärt (Gegenauffassung: transparente Behandlung als anteiliges Wirtschaftsgut nach §23 EStG). Ohne verbindlich belegte Kapitalbeteiligungsquote gilt 0% Teilfreistellung; eine fortlaufende Quote über 50% führt zum Aktienfonds mit 30%, über 25% zum Mischfonds mit 15% (§2 Abs. 6, 7 InvStG). Produkte ohne bestätigte Fondsart bleiben aus den KAP-INV-Formularzeilen ausgeschlossen, bis die Fondsart im Bereich Anlage KAP-INV ausdrücklich bestätigt wird; bis dahin erscheinen sie dort als Prüffall. Ausdrückliche Limited Partnerships wie USO/UNG sind dagegen nach §1 Abs. 3 Nr. 2 InvStG ausgeschlossen. Ihre Broker-PnL und Ausschüttungen werden nicht ersatzweise in Topf 2 geschoben, sondern bis zur Jahresallokation blockiert. Optionen auf ETFs bleiben in Topf 2.

**Jahresfilter:** Es wird `reportDate` verwendet, nicht `dateTime`. Grund: Trades am Jahresende (z.B. `dateTime=2024-12-29`, Settlement `reportDate=2025-01-02`) gehören steuerlich zum Settlement-Jahr (Zuflussprinzip §11 EStG).

---

### Schritt 4: Stillhalterprämien separieren (BMF Rn. 26, 33)

Bei Optionsassignments bündelt IBKR die Prämie in den Aktien-PnL. Das BMF verlangt eine Trennung:

**Erkennung eines Assignments:**
- `assetCategory` ∈ (OPT, FOP, FSFOP)
- `transactionType` = `BookTrade` (keine Börsentransaktion, sondern Ausbuchung)
- `buySell` = `BUY` (Short-Position wird geschlossen)
- `putCall` ∈ (C, P), sowohl Calls als auch Puts
- `fifoPnlRealized` ≈ 0 (IBKR zeigt keinen PnL auf der Option)

**Original-Verkauf finden:**
- Vorrangig über IBKRs Kontraktnummer (`conid`), die auch eine Umbenennung oder einen Split der Optionsserie übersteht. Ohne `conid` über `strike`, `expiry`, `putCall`.
- Es können mehrere Teilfüllungen sein; verbraucht wird nach FIFO, älteste zuerst.

**Prämien-Berechnung:**
```
Prämie (Trade-Währung) = tradePrice × multiplier × quantity
Prämie (EUR)           = Σ über die verbrauchten Teilfüllungen:
                         Betrag je Füllung × Kurs am Verkaufstag dieser Füllung
```
Jede Teilfüllung wird einzeln umgerechnet. Ein gewichteter Durchschnittskurs wäre falsch, sobald sich Preise und Wechselkurse der Füllungen unterscheiden.

**Topf-Umbuchung:**
- `stocks_gain -= Prämie` (aus Topf 1 entfernen)
- `options_gain += Prämie` (in Topf 2 als §20 Abs. 1 Nr. 11)

**Cross-Year:** Wenn die Option in einem Vorjahr verkauft wurde und im Steuerjahr assigned wird, gehört die Prämie ins Vorjahr (Zuflussprinzip). Vorjahres-XMLs müssen mit hochgeladen werden (beim Start oder über "Daten ändern" in der Sidebar), damit der Original-SELL gefunden wird. Findet das Tool zu einer Andienung im Steuerjahr keinen Original-Verkauf (fehlendes oder lückenhaftes Vorjahres-XML), erscheint eine Stillhalter-Warnung mit den betroffenen Serien; die Prämie bleibt dann unkorrigiert im Aktien-Ergebnis. Fehlt der Original-Verkauf bei einer noch älteren Put-Andienung, wird nur dann ein separater Prüffall angezeigt, wenn das daraus entstandene Aktien-Lot tatsächlich im Steuerjahr veräußert wurde.

**Cross-Year Put-Korrektur:** Wenn Aktien aus Put-Assignments früherer Jahre im Steuerjahr verkauft werden, wird IBKR's PnL korrigiert. Die Prämie war bereits im Assignment-Jahr versteuert und darf die Anschaffungskosten nicht mindern. Das Matching läuft über FIFO-Lots; Schreibweisen desselben Basiswerts (Handelsplatz-Suffix, Ticker-Wechsel im Jahresverlauf) werden dabei über die Kontraktnummer oder ISIN zusammengeführt.

---

### Schritt 5: Dividenden, Zinsen & Quellensteuer

Aus `statement_of_funds.csv` werden Cash-Positionen nach `activityCode` zugeordnet:

| `activityCode` | Bedeutung | Zuordnung |
|---|---|---|
| `DIV` | Dividenden | Topf 2 (§20 Abs. 1 Nr. 1) |
| `PIL` | Payment in Lieu (Ersatzzahlung bei Wertpapierleihe) | Wie Dividende, Topf 2 |
| `INTR` | Anleihekupon / Zinserträge | Topf 2 (§20 Abs. 1 Nr. 7) |
| `CINT` | Credit Interest (Guthabenzinsen) | Topf 2 |
| `INTP` | Stückzinsen (beim Kauf gezahlt) | Negative Einnahmen, Topf 2 (BMF Rn. 51) |
| `DINT` | Debit Interest (Sollzinsen, Leihgebühren, CFD-Finanzierung) | **Nicht** in Topf 2. Werbungskosten, nach §20 Abs. 9 EStG durch den Sparer-Pauschbetrag abgegolten; nur nachrichtlich |
| `CFD` | CFD-Zinsen und -Gebühren | Habenzinsen in Topf 2, Finanzierungskosten wie `DINT` nur nachrichtlich |
| `FRTAX` / `WHT` | Quellensteuer (Withholding Tax) | Zeile 41 (anrechenbar). Ausnahmen: deutsche Kapitalertragsteuer auf DE-Wertpapieren geht nach Zeile 37/38; liegt sie auf einem DE-Fonds, wird sie als Prüffall gemeldet, da §32d Abs. 5 EStG nur ausländische Steuern erfasst und die Formularzuordnung nicht automatisierbar ist |
| `OFEE` / `STAX` | Gebühren, Umsatzsteuer | Nicht abziehbar (§20 Abs. 9), nur nachrichtlich |
| `TTAX` | Transaktionssteuer | Nach §20 Abs. 4 EStG ergebniswirksam. Bei eindeutigem Match wird die Verkaufssteuer sofort und die Kaufsteuer über das geschlossene Lot anteilig im realisierten Ergebnis berücksichtigt. Bereits in `Trade.taxes` enthaltene Beträge werden nicht doppelt abgezogen. Prüffall bleiben nicht eindeutige Zuordnungen, Steuern auf Stillhalter-Eröffnungen (Zufluss im Eröffnungsjahr, §11 EStG) sowie Instrumente mit eigenem Rechenweg (Anlage SO, Personengesellschaften) |
| `BUY` / `SELL` / `ADJ` / `ASSIGN` / `EXE` | Trade- und Settlement-Buchungen | Übersprungen; das realisierte Ergebnis kommt aus den Trade-Daten |
| `DEP` / `WITH` | Ein- und Auszahlungen | Übersprungen; kein eigener Kapitalertrag |
| `FOREX` | Devisenumsatz | Übersprungen; das Ergebnis kommt aus der separaten FX-Rechnung |
| `CORP` | Kapitalmaßnahme | Wird derzeit übersprungen. Bei T-Bill-Maturities kommt das Ergebnis aus dem BILL-Fallback; Return-of-Capital-Buchungen sind noch nicht automatisch verarbeitet und müssen geprüft werden |

Buchungscodes außerhalb dieser Tabelle werden nicht stillschweigend übergangen, sondern als Prüffall gemeldet.

**Währungsumrechnung (EUR-Basis):** `amount` ist bereits in EUR (BaseCurrency-Ansicht). Keine weitere Umrechnung nötig.

**Jahresfilter:** `reportDate.year == Steuerjahr`. Steuer-Rückforderungen (Tax Reclaims) aus Vorjahren, die im Steuerjahr gebucht werden, sind korrekt dem Buchungsjahr zugeordnet.

---

### Schritt 6: Währungsumrechnung

{"**Konto: EUR-Basis.** Alle Beträge in `statement_of_funds` und `fifoPnlRealized × fxRateToBase` sind direkt in EUR. Es wird kein separater Tageskurs-Lookup benötigt." if base_curr == "EUR" else "**Konto: USD-Basis.** Zweistufige Umrechnung: (1) `fifoPnlRealized × fxRateToBase` → USD, (2) USD → EUR über täglichen Wechselkurs. Kurse werden primär aus EUR-Einträgen in Trades/Funds extrahiert; Lücken werden automatisch mit **EZB-Referenzkursen** gefüllt."}

| Szenario | Formel |
|---|---|
| **EUR-Base, EUR-Trade** | `PnL_EUR = fifoPnlRealized × fxRateToBase` (fxRate ≈ 1.0) |
| **EUR-Base, USD-Trade** | `PnL_EUR = fifoPnlRealized × fxRateToBase` (fxRate ≈ 0.86–0.92) |
| **USD-Base, USD-Trade** | `PnL_EUR = fifoPnlRealized × fxRateToBase × daily_usd_eur_rate` |
| **USD-Base, EUR-Trade** | `PnL_EUR = amount_eur` (direkt, da Trade in EUR) |

**Plausibilitätsprüfung:** USD→EUR-Kurse außerhalb [0.70, 1.30] werden verworfen. `fxRateToBase=1.0` auf EUR-Währungseinträgen wird nicht als USD/EUR-Kursquelle verwendet; die zugrunde liegende Buchung bleibt in der Verarbeitung erhalten.

---

### Schritt 7: FX-Gewinne/-Verluste

Fremdwährungsgewinne/-verluste entstehen durch Kursänderungen auf verzinslichen Fremdwährungskonten (BMF Rn. 131, §20 Abs. 2 S. 1 Nr. 7 EStG).

**Datenquellen (Priorität):**

| Priorität | Quelle | Genauigkeit | Wann verfügbar |
|---|---|---|---|
| 1. | **XML `<FxTransactions>`** | Exakt (IBKR-internes FIFO, `realizedPL` pro Transaktion) | Wenn in Flex Query aktiviert |
| 2. | **IBKR Standard-Bericht (CSV)** | Aggregierter IBKR-FIFO-Rohwert; keine Prüfung einzelner Schuldtilgungen möglich | Manuell erstellt und über den CSV-Uploader geladen; nur für EUR-Basiskonten und nur bei einem einzelnen Konto als FX-Quelle |
| 3. | **FIFO-Approximation** | Näherung aus den Kontobewegungen | Bei EUR-Basiskonten, wenn keine vorrangige Quelle greift |

Nur Quelle 1 enthält einzelne Buchungen: Dort filtert das Tool bei aktiver Saldo-Korrektur Schuldtilgungen heraus; Abflüsse bleiben ungekürzt, weil IBKR bei nur teilweise gedeckten Buchungen bereits allein das Ergebnis des gedeckten Teils ausweist. Quelle 2 ist aggregiert und kann diese Prüfung nicht leisten. Bei negativem Währungssaldo und aktiver Korrektur wird der CSV-Wert deshalb nicht verwendet, sondern auf Quelle 3 zurückgefallen. Ist die Korrektur deaktiviert, wird der CSV-Rohwert bewusst unverändert übernommen.

FX-Gewinne/-Verluste fließen in **Topf 2**.

---

### Schritt 8: Anlage KAP + KAP-INV Berechnung

```
Topf 1 = Aktiengewinne + Aktienverluste (nach Stillhalter-Separation)
         (ohne InvStG-ETFs, wenn aktiviert)
Topf 2 = Dividenden + Zinsen
         + realisierte G/V aller Topf-2-Instrumente
           (Optionen, Futures, Anleihen, T-Bills und Produkte außerhalb InvStG)
         + Stillhalter-, Zufluss-, FX- und Tageskurs-Anpassungen

Zeile 19 = Topf 1 + Topf 2 (Nettobetrag)
Zeile 20 = Aktiengewinne (brutto, ohne Verluste)
Zeile 22 = |Verluste ohne Aktien| (positiver Betrag)
Zeile 23 = |Aktienverluste| (positiver Betrag)
Zeile 41 = Quellensteuer außerhalb der Fonds
           + anrechenbare Fonds-Quellensteuer aus KAP-INV
           (Erstattungsüberschüsse bleiben als negative Korrektur erhalten)
```

**Anlage KAP-INV (wenn InvStG aktiviert):**

```
ETF-Gewinne/-Verluste und ETF-Dividenden werden vor Teilfreistellung und
einschließlich ausländischen Steuerabzugs nach Fondsart auf KAP-INV gemeldet.
Die Teilfreistellung wird pro ISIN nur als steuerlicher Kontrollwert berechnet:
  Aktienfonds (≥51% Aktienquote): 30% steuerfrei
  Sonstiger Fonds:                 0% steuerfrei

KAP-INV-Zeilen = Rohbeträge vor TFS nach Fondsart (Z. 4–8 und 14/17/20/23/26)
Kontrollwert = (ETF-G/V bzw. erhaltene ETF-Ausschüttung) × (1 − TFS);
                gezahlte Ersatzzahlungen werden nicht gegengerechnet;
                kein Eintragungswert
ETF-Quellensteuer wird im Standardmodus wie vor dem DBA-Update proportional
zur Teilfreistellung gekürzt. Optional kann die ausdrücklich als Beta markierte
ereignisbezogene DBA-/Erstattungsprüfung aktiviert werden. Der Ergebnisbetrag
wird in Anlage KAP Zeile 41 ausgewiesen.
Veräußerungswerte sind bis zur Berücksichtigung bereits angesetzter
Vorabpauschalen ausdrücklich noch nicht final.
```

**Tageskurs-Korrektur (wenn aktiviert):**

```
Korrektur = Σ |Anschaffungskosten| × (FX_Verkauf − FX_Kauf) pro CLOSED_LOT
Futures ausgeschlossen (Kostenbasis = Notional, kein realer Cashflow).
Wird auf Topf 1, Topf 2 und KAP-INV aufgeteilt. In die KAP-INV-Formularzeile
fließt das rohe Delta; das teilfreigestellte Delta dient nur der Kontrollrechnung.
```
""")


# ── Export: Excel-Builder (aus dem Bestands-Code gehoben) ────────────────────

topf_readable = {
    'Topf1': 'Topf 1 - Aktien (§20 Abs. 2 Nr. 1 EStG)',
    'Topf2': 'Topf 2 - Sonstiges (Termingeschäfte, Stillhalter, FX)',
    'KAP-INV': 'Anlage KAP-INV (InvStG)',
    'Anlage SO': 'Anlage SO (§23 EStG)',
    'Personengesellschaft': (
        'Personengesellschaft - beobachtete Brokerwerte; '
        'Jahresallokation fehlt'
    ),
    'Nicht zugeordnet': 'Nicht zugeordnet (bitte manuell prüfen)',
}
EXPORT_TOPF_ORDER = (
    'Topf1', 'Topf2', 'KAP-INV', 'Anlage SO', 'Personengesellschaft',
    'Nicht zugeordnet',
)
cat_labels = {
    'STK': 'Aktie', 'OPT': 'Option', 'FUT': 'Future',
    'FOP': 'Futures-Option', 'FSFOP': 'Flex-Option',
    'BILL': 'T-Bill', 'BOND': 'Anleihe',
    'WAR': 'Optionsschein', 'CFD': 'CFD',
}

def _format_instrument(row):
    sym = row.get('symbol', '') or ''; desc = row.get('description', '') or ''
    pc = row.get('putCall', '') or ''; strike = row.get('strike', '') or ''; expiry = row.get('expiry', '') or ''
    if pc and strike and expiry:
        pc_label = 'Call' if pc == 'C' else 'Put'
        exp_fmt = f"{expiry[:4]}-{expiry[4:6]}-{expiry[6:]}" if len(expiry) == 8 else expiry[:10]
        return f"{sym} ({pc_label} {strike} exp. {exp_fmt})"
    if desc and sym: return f"{sym} ({desc})"
    return sym or desc or ''

def _get_group_key(row):
    us = (row.get('underlyingSymbol', '') or '').strip()
    if us: return us.split()[0]
    sym = (row.get('symbol', '') or '').strip()
    return sym.split()[0] if sym else '?'

def _build_excel(trade_details, trades_by_topf, export_context):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    wb = Workbook()
    ws_details = wb.active
    ws_details.title = f"Trade-Details {steuerjahr}"
    ws = wb.create_sheet("Zusammenfassung")
    hdr_font = Font(bold=True, color="FFFFFF", size=11); hdr_fill = PatternFill("solid", fgColor="1e3a5f")
    grp_font = Font(bold=True, size=10); grp_fill = PatternFill("solid", fgColor="d6e4f0")
    gain_font = Font(color="006100", size=9); gain_fill = PatternFill("solid", fgColor="e2efda")
    loss_font = Font(color="9c0006", size=9); loss_fill = PatternFill("solid", fgColor="fce4ec")
    korr_font = Font(italic=True, size=9); korr_fill = PatternFill("solid", fgColor="fff9c4")
    sub_font = Font(bold=True, size=9); sub_fill = PatternFill("solid", fgColor="f2f2f2")
    total_font = Font(bold=True, size=10, color="FFFFFF"); total_fill = PatternFill("solid", fgColor="4a4a4a")
    normal_font = Font(size=9); thin_border = Border(bottom=Side(style='thin', color='cccccc'))
    num_fmt_eur = '#,##0.00'; num_fmt_4d = '#,##0.0000'

    f = export_context['final']
    has_etf = export_context['has_etf_data'] and export_context['invstg_aktiv']
    kap_inv_form_export = export_context.get('kap_inv_form', {})
    has_so = export_context['has_so_data']
    special_products = export_context['no_invstg_summary']
    partnership_tax_items = export_context.get('partnership_tax_items', {})
    so_taxable = export_context['so_taxable']
    so_free = export_context['so_free']
    trade_sums = {
        key: sum(float(r.get('pnl_eur') or 0) for r in rows)
        for key, rows in trades_by_topf.items()
    }
    trade_topf1_reconciled = trade_sums.get('Topf1', 0)
    trade_topf2_reconciled = trade_sums.get('Topf2', 0)
    kap_inv_trade_total = trade_sums.get('KAP-INV', 0)
    kap_inv_reintegration_note = ""
    if export_context['has_etf_data'] and not export_context['invstg_aktiv'] and abs(kap_inv_trade_total) > 0.005:
        trade_topf1_reconciled += kap_inv_trade_total
        kap_inv_reintegration_note = "KAP-INV-Detailwerte wurden wegen deaktivierter InvStG-Klassifizierung in Topf 1 reintegriert."

    summary_cols = ["Bereich", "Position", "Wert EUR / Anzahl", "Hinweis"]
    for i, width in enumerate([22, 44, 16, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_cols))
    cell = ws.cell(row=1, column=1, value=f"ANLAGE KAP {steuerjahr} - Steuerbericht")
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='left')
    row_num = 2
    meta_rows = [
        ("Erstellt", export_context['created_at']),
        ("Basiswährung", export_context['base_currency']),
        ("Quelle", "final (GUI/Textreport Single Source of Truth)"),
    ]
    for label, value in meta_rows:
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=value)
        for ci in range(1, len(summary_cols) + 1):
            cell = ws.cell(row=row_num, column=ci)
            cell.font = normal_font
            cell.border = thin_border
        ws.cell(row=row_num, column=1).font = sub_font
        row_num += 1

    summary_rows = [
        ("Topf 1", "Aktiengewinne", f['stocks_gain'], ""),
        ("Topf 1", "Aktienverluste", f['stocks_loss'], ""),
        ("Topf 1", "Saldo Aktien", f['topf_1'], ""),
        ("Topf 2", "Dividenden", f['dividends'], ""),
        ("Topf 2", "Zinsen netto", f['interest'], ""),
        ("Topf 2", "Sonstige Gewinne", f['options_gain'], ""),
        ("Topf 2", "Sonstige Verluste", f['options_loss'], ""),
        ("Topf 2", "Saldo Sonstiges", f['topf_2'], ""),
        ("Anlage KAP", "Zeile 7 - inländischer Steuerabzug", f['zeile_7'], ""),
        ("Anlage KAP", "Zeile 19 - ausländische Kapitalerträge netto", f['zeile_19'], ""),
        ("Anlage KAP", "Zeile 20 - Aktiengewinne", f['zeile_20'], ""),
        ("Anlage KAP", "Zeile 22 - Verluste ohne Aktien", f['zeile_22'], "positiver Eintrag"),
        ("Anlage KAP", "Zeile 23 - Aktienverluste", f['zeile_23'], "positiver Eintrag"),
        ("Anlage KAP", "Zeile 37 - Kapitalertragsteuer", f['zeile_37'], ""),
        ("Anlage KAP", "Zeile 38 - Solidaritätszuschlag", f['zeile_38'], ""),
        ("Anlage KAP", "Zeile 41 - ausl. Quellensteuer", f['quellensteuer'], ""),
    ]
    if has_etf:
        for line in kap_inv_form_export.get('lines', []):
            note = (
                "Bruttowert vor TFS; steuerpflichtiger Kontrollwert "
                f"{fmt_de(line.get('taxable_control_eur', 0))} EUR"
            )
            if line.get('kind') == 'sale':
                note += "; vor Abzug bereits angesetzter Vorabpauschalen"
            summary_rows.append((
                "Anlage KAP-INV",
                f"Zeile {line['line']} - {line['fund_type']}",
                line['amount_raw_eur'],
                note,
            ))
        for item in kap_inv_form_export.get('blocked_details', []):
            summary_rows.append((
                "Anlage KAP-INV Prüffall",
                f"{item.get('ticker', '')} ({item['isin']}) - Ausschüttung roh",
                item.get('distribution_raw_eur', 0),
                "keine Formularzeile bis zur bestätigten Fondsart; "
                f"G/V roh {fmt_de(item.get('sale_raw_eur', 0))} EUR",
            ))
        for item in kap_inv_form_export.get('negative_distribution_details', []):
            summary_rows.append((
                "Anlage KAP-INV Prüffall",
                f"{item.get('ticker', '')} ({item['isin']}) - gezahlte Ausschüttungen",
                item.get('paid_distribution_eur', 0),
                "gezahlte Dividenden/Ersatzzahlungen (Short-Position); nicht in "
                "den Ausschüttungszeilen enthalten; steuerliche Behandlung "
                "manuell prüfen",
            ))
        summary_rows.append((
            "Anlage KAP-INV",
            "Fonds-QSt: enthalten in Anlage KAP Zeile 41",
            f['etf_wht'],
            "keine eigene KAP-INV-Zeile",
        ))
    for isin, info in sorted(special_products.items(), key=lambda x: x[1].get('ticker', '')):
        summary_rows.append((
            "Topf 2 Sonderprodukte",
            f"{info.get('ticker', isin)} ({isin})",
            info.get('total', 0),
            "no_invstg; realisiertes G/V + Tageskurs + Ausschüttungen; "
            f"QSt {fmt_de(info.get('wht_reported', 0))} EUR; negativ = Erstattung",
        ))
    for isin, item in sorted(
            partnership_tax_items.items(),
            key=lambda x: x[1].get('ticker', '')):
        observed_total = (
            item.get('observed_trade_pnl_eur', 0)
            + item.get('observed_tageskurs_delta_eur', 0)
            + item.get('observed_distributions_eur', 0)
            + item.get('observed_other_cash_eur', 0)
        )
        summary_rows.append((
            "Personengesellschaft · blockiert",
            f"{item.get('ticker', isin)} ({isin}) · beobachtete Brokerwerte",
            observed_total,
            "Kein Steuerwert: K-1/K-3 bzw. äquivalente Jahresallokation fehlt; "
            f"QSt beobachtet {fmt_de(item.get('observed_withholding_tax_eur', 0))} EUR",
        ))
    if has_so:
        summary_rows.extend([
            ("Anlage SO", "Steuerpflichtig <= 1 Jahr", so_taxable, ""),
            ("Anlage SO", "Steuerfrei > 1 Jahr", so_free, "nicht in Anlage KAP"),
        ])

    row_num += 1
    for ci, cn in enumerate(summary_cols, 1):
        cell = ws.cell(row=row_num, column=ci, value=cn)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="e8e8e8")
        cell.border = thin_border
    row_num += 1
    current_area = None
    for area, label, value, note in summary_rows:
        if area != current_area:
            current_area = area
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(summary_cols))
            cell = ws.cell(row=row_num, column=1, value=area)
            cell.font = grp_font
            cell.fill = grp_fill
            row_num += 1
        values = [area, label, value, note]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if ci == 3 and isinstance(val, (int, float)):
                cell.number_format = num_fmt_eur
                if val > 0.005:
                    cell.fill = gain_fill
                    cell.font = gain_font
                elif val < -0.005:
                    cell.fill = loss_fill
                    cell.font = loss_font
        if label.startswith("Saldo ") or label.startswith("Zeile 19") or label.startswith("Erträge "):
            for ci in range(1, len(summary_cols) + 1):
                ws.cell(row=row_num, column=ci).font = sub_font
                ws.cell(row=row_num, column=ci).fill = sub_fill
        row_num += 1

    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(summary_cols))
    cell = ws.cell(row=row_num, column=1, value="Abstimmung")
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='left')
    row_num += 1
    reconciliation_rows = [
        ("Topf 1", "Finaler Wert", f['topf_1'], "GUI/TXT/Excel-Summary"),
        ("Topf 1", "Trade-Details", trade_topf1_reconciled, kap_inv_reintegration_note or "Nachweis-Sheet"),
        ("Topf 1", "Differenz", f['topf_1'] - trade_topf1_reconciled, "Kontrollwert; Abweichungen können aus PnL-Summary-Fallbacks, Rundung oder Toggle-Reintegration stammen"),
        ("Topf 2", "Finaler Wert", f['topf_2'], "inkl. Dividenden, Zinsen, FX, Stillhalter"),
        ("Topf 2", "Trade-Details", trade_topf2_reconciled, "Trades und Korrekturen"),
        ("Topf 2", "Nicht-Trade-Anteil", f['topf_2'] - trade_topf2_reconciled, "Cash-Erträge/Korrekturen, z.B. Dividenden und Zinsen"),
        ("Anlage KAP", "Zeile 19 minus Topf 1 minus Topf 2", f['zeile_19'] - (f['topf_1'] + f['topf_2']), "sollte 0 sein"),
    ]
    for ci, cn in enumerate(summary_cols, 1):
        cell = ws.cell(row=row_num, column=ci, value=cn)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="e8e8e8")
        cell.border = thin_border
    row_num += 1
    current_area = None
    for area, label, value, note in reconciliation_rows:
        if area != current_area:
            current_area = area
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(summary_cols))
            cell = ws.cell(row=row_num, column=1, value=area)
            cell.font = grp_font
            cell.fill = grp_fill
            row_num += 1
        for ci, val in enumerate([area, label, value, note], 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if ci == 3 and isinstance(val, (int, float)):
                cell.number_format = num_fmt_eur
        if label in ("Differenz", "Nicht-Trade-Anteil", "Zeile 19 minus Topf 1 minus Topf 2"):
            for ci in range(1, len(summary_cols) + 1):
                ws.cell(row=row_num, column=ci).font = sub_font
                ws.cell(row=row_num, column=ci).fill = sub_fill
        row_num += 1
    ws.freeze_panes = 'A6'

    ws = ws_details
    cols = ['Datum', 'Handelsdatum', 'Wertpapier', 'ISIN', 'Kategorie',
            'K/V', 'Stk.', 'Kurs', 'Kostenbasis', 'Erloese',
            'G/V (Orig.)', 'Kommission', 'Waehrung', 'Wechselkurs', 'G/V (EUR)', 'Anmerkung']
    col_widths = [12, 12, 42, 15, 10, 6, 8, 11, 13, 13, 13, 11, 6, 11, 14, 40]
    eur_col = 15
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    notice = ws.cell(
        row=1,
        column=1,
        value=(
            "Hinweis: IBKR-/OCC-Bezeichnungen können vom geläufigen "
            "Basiswert abweichen, z. B. BRKB/BRK B und ODAX für DAX."
        ),
    )
    notice.font = Font(italic=True, size=9, color="7f6000")
    notice.fill = korr_fill
    row_num = 3
    for topf_key in EXPORT_TOPF_ORDER:
        topf_rows = trades_by_topf.get(topf_key, [])
        if not topf_rows: continue
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(cols))
        cell = ws.cell(row=row_num, column=1, value=topf_readable.get(topf_key, topf_key))
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal='left')
        row_num += 1
        for ci, cn in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=ci, value=cn)
            cell.font = Font(bold=True, size=9); cell.fill = PatternFill("solid", fgColor="e8e8e8"); cell.border = thin_border
        row_num += 1
        groups = defaultdict(list)
        for r in topf_rows: groups[_get_group_key(r)].append(r)
        topf_total = 0.0
        for grp_key in sorted(groups.keys()):
            grp_rows = groups[grp_key]
            grp_rows.sort(key=lambda r: r.get('dateTime', '') or r.get('reportDate', '') or '')
            grp_desc = ''; grp_isin = ''
            for r in grp_rows:
                if r.get('description') and r.get('source') != 'stillhalter_korrektur': grp_desc = r['description']
                if r.get('isin'): grp_isin = r['isin']
                if grp_desc and grp_isin: break
            grp_label = grp_key
            if grp_desc: grp_label += f" - {grp_desc}"
            if grp_isin: grp_label += f" ({grp_isin})"
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(cols))
            cell = ws.cell(row=row_num, column=1, value=grp_label); cell.font = grp_font; cell.fill = grp_fill
            row_num += 1; grp_total = 0.0
            for r in grp_rows:
                source = r.get('source', ''); pnl_eur = r.get('pnl_eur', 0)
                pnl_orig = r.get('fifoPnlRealized', 0); fx = r.get('fxRateToBase', 0)
                cost = r.get('cost', 0); proceeds = r.get('proceeds', 0); price = r.get('tradePrice', 0)
                commission = r.get('ibCommission', 0) or 0
                anmerkung = ''
                if source == 'pnl_summary': anmerkung = 'Aus IBKR PnL-Summary'
                elif source == 'stillhalter_korrektur': anmerkung = r.get('description', 'Korrektur')
                elif source == 'zufluss': anmerkung = r.get('description', 'Zufluss §11 EStG')
                elif source == 'zufluss_korrektur': anmerkung = r.get('description', 'Vorjahres-Korrektur')
                elif source == 'tageskurs_korrektur': anmerkung = r.get('description', 'Tageskurs')
                elif source == 'cross_year_put_korrektur': anmerkung = r.get('description', 'Cross-Year Put-Korrektur')
                elif source == 'trades' and r.get('stillhalter_adjusted'):
                    anmerkung = 'Korrigiert: Prämie separiert (s. Stillhalterprämie Topf 2)'
                    if r.get('invstg_basis_adjustment_raw'):
                        anmerkung += '; KAP-INV-AK auf Ausübungspreis normalisiert'
                bs = r.get('buySell', ''); oc = r.get('openClose', '')
                if bs == 'SELL' and oc == 'O': bs_label = 'STO'
                elif bs == 'BUY' and oc == 'C': bs_label = 'BTC'
                elif bs == 'BUY' and oc == 'O': bs_label = 'BTO'
                elif bs == 'SELL' and oc == 'C': bs_label = 'STC'
                else: bs_label = bs
                values = [
                    (r.get('reportDate', '') or '')[:10], (r.get('dateTime', '') or '')[:10],
                    _format_instrument(r), r.get('isin', ''),
                    cat_labels.get(r.get('assetCategory', ''), r.get('assetCategory', '')),
                    bs_label, r.get('quantity', ''),
                    price if price else None, cost if cost else None, proceeds if proceeds else None,
                    pnl_orig if pnl_orig else None, commission if commission else None, r.get('currency', ''),
                    fx if fx else None, pnl_eur, anmerkung,
                ]
                for ci, val in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=ci, value=val); cell.font = normal_font
                    if ci in (8, 9, 10, 11, 12, 15) and isinstance(val, (int, float)): cell.number_format = num_fmt_eur
                    elif ci == 14 and isinstance(val, (int, float)): cell.number_format = num_fmt_4d
                if source in ('stillhalter_korrektur', 'zufluss', 'zufluss_korrektur', 'tageskurs_korrektur', 'cross_year_put_korrektur'):
                    for ci in range(1, len(cols) + 1): ws.cell(row=row_num, column=ci).fill = korr_fill; ws.cell(row=row_num, column=ci).font = korr_font
                elif pnl_eur > 0.005:
                    for ci in range(1, len(cols) + 1): ws.cell(row=row_num, column=ci).fill = gain_fill; ws.cell(row=row_num, column=ci).font = gain_font
                elif pnl_eur < -0.005:
                    for ci in range(1, len(cols) + 1): ws.cell(row=row_num, column=ci).fill = loss_fill; ws.cell(row=row_num, column=ci).font = loss_font
                grp_total += pnl_eur; row_num += 1
            ws.cell(row=row_num, column=1, value=f"Zwischensumme {grp_key}")
            for ci in range(1, len(cols) + 1): ws.cell(row=row_num, column=ci).font = sub_font; ws.cell(row=row_num, column=ci).fill = sub_fill; ws.cell(row=row_num, column=ci).border = thin_border
            cell = ws.cell(row=row_num, column=eur_col, value=grp_total); cell.number_format = num_fmt_eur; cell.font = sub_font; cell.fill = sub_fill
            topf_total += grp_total; row_num += 1
        topf_label = topf_readable.get(topf_key, topf_key).split(' - ')[0]
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=eur_col - 1)
        cell = ws.cell(row=row_num, column=1, value=f"SUMME {topf_label}"); cell.font = total_font; cell.fill = total_fill; cell.alignment = Alignment(horizontal='right')
        for ci in range(1, len(cols) + 1): ws.cell(row=row_num, column=ci).fill = total_fill
        cell = ws.cell(row=row_num, column=eur_col, value=topf_total); cell.number_format = num_fmt_eur; cell.font = total_font; cell.fill = total_fill
        row_num += 2
    ws.freeze_panes = 'A4'
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── Export: Textreport ───────────────────────────────────────────────────────

def _build_text_report():
    etf_wht = final["etf_wht"]
    sh_count = audit.get("stillhalter_count", 0)
    sh_eur = audit.get("stillhalter_premium_eur", 0)
    zeile_7 = d.get("zeile_7_kapitalertraege_mit_inlaendischem_steuerabzug_eur", 0)
    zeile_37 = d.get("zeile_37_kapitalertragsteuer_eur", 0)
    zeile_38 = d.get("zeile_38_solidaritaetszuschlag_eur", 0)
    classification_review_items = d.get("classification_review_items", []) or []
    partnership_tax_items = d.get("partnership_tax_items", {}) or {}
    fx_export = ""
    if fx_results:
        fx_export = "\nFREMDWÄHRUNGS-GEWINNE/VERLUSTE (FIFO)\n"
        for curr, data in sorted(fx_results.items()):
            fx_export += f"  {curr}: Gewinn {fmt_de(data['gain']):>10}  Verlust {fmt_de(data['loss']):>10}  Netto {fmt_de(data['net']):>10} EUR\n"
        fx_net = fx_total_gain + fx_total_loss
        fx_export += f"  ─────────────────────────────────────────────────\n"
        fx_export += f"  FX Gesamt Gewinn:      {fmt_de(fx_total_gain):>14} EUR\n"
        fx_export += f"  FX Gesamt Verlust:     {fmt_de(fx_total_loss):>14} EUR\n"
        fx_export += f"  FX Netto:              {fmt_de(fx_net):>14} EUR\n"
        fx_export += "  (In Topf 2 enthalten, BMF Rn. 131)\n"
        # Toggle-Stand dokumentieren, damit der Bericht auch ohne GUI nachvollziehbar bleibt
        _fx_corr_active = d.get('fx_margin_correction_enabled', True)
        _fx_neg = d.get('fx_has_negative_balance', False)
        _fx_meta = d.get('fx_option_a_meta', {}) or {}
        _fx_debt_n = _fx_meta.get('debt_repayments', 0)
        _fx_debt_pnl = _fx_meta.get('debt_repayment_pnl', 0.0)
        if _fx_corr_active:
            fx_export += "  Saldo-Korrektur (§20 Abs. 2 S. 1 Nr. 7 EStG): AKTIV (buchungs- und saldogeprüft)\n"
            if _fx_debt_n:
                fx_export += (
                    f"    → {_fx_debt_n} Buchungen tilgen eine Fremdwährungs-Schuld "
                    f"({fmt_de(_fx_debt_pnl)} EUR) und bleiben unberücksichtigt.\n"
                )
        else:
            fx_export += "  Saldo-Korrektur (§20 Abs. 2 S. 1 Nr. 7 EStG): DEAKTIVIERT (Opt-out)\n"
            if _fx_debt_n:
                fx_export += (
                    f"    → {_fx_debt_n} Buchungen aus Schuldtilgung ({fmt_de(_fx_debt_pnl)} EUR) "
                    f"sind mit IBKR-Rohwert enthalten.\n"
                )
            elif _fx_neg:
                fx_export += "    → IBKR-/Rohwerte übernommen trotz negativem Fremdwährungssaldo.\n"
        _fx_open_anom = _fx_meta.get('open_rows_with_pnl', []) or []
        if _fx_open_anom:
            fx_export += (
                f"  PRUEFFALL: {len(_fx_open_anom)} Buchungen tragen ein Ergebnis, obwohl IBKR\n"
                f"    sie als Eroeffnung ausweist (code != 'C'). Als steuerbar behandelt.\n"
            )

    sh_export = ""
    if sh_count > 0:
        sh_export = f"\nSTILLHALTERPRÄMIEN (BMF Rn. 25-35)\n"
        sh_export += f"  {sh_count} Assignment(s) erkannt\n"
        sh_export += f"  Prämien umgebucht:     {fmt_de(sh_eur):>14} EUR\n"
        sh_export += f"  (Von Topf 1 nach Topf 2 verschoben)\n"

    inv_export = ""
    kap_inv_entries_export = ""
    if has_etf_data and invstg_aktiv:
        inv_export = f"\nANLAGE KAP-INV: INVESTMENTFONDS (InvStG)\n"
        if dba_wht_beta_enabled:
            inv_export += (
                "  Fonds-QSt-Modus: DBA-BETA AKTIV "
                "(Ereignis-Matching/DBA-Caps; manuell prüfen)\n"
            )
        else:
            inv_export += (
                "  Fonds-QSt-Modus: STANDARD "
                "(Rohsteuer × (1 - Teilfreistellung); DBA-Beta aus)\n"
            )
        kap_inv_entries_export = "\nANLAGE KAP-INV EINTRAGUNGEN\n"
        for line in kap_inv_form.get('lines', []):
            suffix = (
                " (vor Abzug bereits angesetzter Vorabpauschalen)"
                if line.get('kind') == 'sale' else ""
            )
            form_row = (
                f"  Zeile {line['line']:>2}: {fmt_de(line['amount_raw_eur']):>12} EUR  "
                f"{line['fund_type']} · vor TFS{suffix}\n"
            )
            inv_export += form_row
            kap_inv_entries_export += form_row
            inv_export += (
                f"             Kontrollwert nach TFS: "
                f"{fmt_de(line['taxable_control_eur']):>10} EUR "
                "(kein Eintragungswert)\n"
            )
        inv_export += (
            f"  Fonds-QSt anrechenbar: {fmt_de(etf_wht):>12} EUR  "
            "→ bereits in Anlage KAP Zeile 41 enthalten\n"
        )
        kap_inv_entries_export += (
            f"  Fonds-QSt: {fmt_de(etf_wht):>12} EUR  "
            "→ Anlage KAP Zeile 41, keine KAP-INV-Zeile\n"
        )
        for warning in kap_inv_form.get('warnings', []):
            inv_export += f"  ACHTUNG: {warning}\n"
            kap_inv_entries_export += f"  ACHTUNG: {warning}\n"
        for item in kap_inv_form.get('blocked_details', []):
            blocked_row = (
                f"  PRUEFFALL {item.get('ticker', '')} ({item['isin']}): "
                f"Aussch. roh {fmt_de(item.get('distribution_raw_eur', 0))} EUR, "
                f"G/V roh {fmt_de(item.get('sale_raw_eur', 0))} EUR; "
                "keine Formularzeile ohne bestaetigte Fondsart\n"
            )
            inv_export += blocked_row
            kap_inv_entries_export += blocked_row
        for item in kap_inv_form.get('negative_distribution_details', []):
            paid_row = (
                f"  PRUEFFALL {item.get('ticker', '')} ({item['isin']}): "
                f"gezahlte Ausschüttungen {fmt_de(item.get('paid_distribution_eur', 0))} EUR "
                "(Short-Position); nicht in den Ausschüttungszeilen enthalten\n"
            )
            inv_export += paid_row
            kap_inv_entries_export += paid_row
        inv_export += "  Details je ISIN:\n"
        for detail in kap_inv_form.get('details', []):
            inv_export += (
                f"    {detail.get('ticker', detail['isin']):8s} "
                f"Aussch. Formular {fmt_de(detail['distribution_raw_eur']):>10} EUR  "
                f"G/V KAP-INV {fmt_de(detail['sale_raw_eur']):>10} EUR  "
                f"TFS {detail['tfs_rate']*100:.0f}%\n"
            )

    topf2_detail_export = ""
    if topf2_breakdown:
        topf2_detail_export = "\nAUFSCHLÜSSELUNG TOPF 2\n"
        for row in topf2_breakdown['rows']:
            topf2_detail_export += (
                f"  {row['label']:24s} G {fmt_de(row['gain']):>10} "
                f"V {fmt_de(row['loss']):>10} N {fmt_de(row['net']):>10} EUR\n"
            )
        topf2_detail_export += (
            f"  {'Saldo Topf 2':24s} G {fmt_de(topf2_breakdown['total_gain']):>10} "
            f"V {fmt_de(topf2_breakdown['total_loss']):>10} "
            f"N {fmt_de(topf2_breakdown['net']):>10} EUR\n"
            "  Anpassungszeilen sind mit Vorzeichen in G/V addierbar.\n"
        )

    special_products_export = ""
    if no_invstg_summary:
        special_products_export = "\nSONDERPRODUKTE AUSSERHALB INVSTG (IN TOPF 2 ENTHALTEN)\n"
        for isin, info in sorted(no_invstg_summary.items(), key=lambda x: x[1].get('ticker', '')):
            realized = info.get('gain', 0) + info.get('loss', 0)
            special_products_export += (
                f"  {info.get('ticker', isin):8s} {isin}  "
                f"G/V {fmt_de(realized):>10}  TK {fmt_de(info.get('tageskurs', 0)):>10}  "
                f"Aussch. {fmt_de(info.get('div', 0)):>10}  "
                f"Summe {fmt_de(info.get('total', 0)):>10} EUR\n"
            )
            wht_reported = info.get('wht_reported', 0)
            if abs(wht_reported) > 0.005:
                wht_label = "Quellensteuer" if wht_reported > 0 else "QSt-Erstattung"
                special_products_export += (
                    f"           {wht_label}: {fmt_de(wht_reported):>10} EUR\n"
                )

    de_kest_export = ""
    if abs(zeile_7) > 0.01:
        z_kest_total = zeile_37 + zeile_38
        if de_kest_variante_b:
            de_kest_export = (
                "\nHINWEIS DEUTSCHE DIVIDENDEN (Variante B aktiv)\n"
                f"  Bruttodividende {fmt_de(zeile_7)} EUR wurde nach Zeile 19 verschoben.\n"
                f"  DE-KESt+Soli {fmt_de(z_kest_total)} EUR wurde nach Zeile 41 verschoben.\n"
                "  Variante A (Z. 7/37/38) ist tax-legally präziser und im GUI umschaltbar.\n"
            )
        else:
            de_kest_export = (
                "\nHINWEIS DEUTSCHE DIVIDENDEN (Variante A aktiv)\n"
                "  Falls das Steuerprogramm Z. 7/37/38 nicht freischaltet, alternativ:\n"
                f"    Zeile 19: +{fmt_de(zeile_7)} EUR  (Bruttodividende)\n"
                f"    Zeile 41: +{fmt_de(z_kest_total)} EUR  (DE-KESt+Soli als anrechenbare Steuer)\n"
                "  Variante B ist eine technische Ersatzdarstellung und kein amtlich belegter\n"
                "  Ersatz fuer die Steuerbescheinigung; vor Abgabe fachlich abstimmen.\n"
            )

    multi_acct_export = ""
    if n_accounts > 1:
        multi_acct_export = f"Konten: {n_accounts} (separat berechnet, Ergebnisse addiert)\n"

    classification_review_export = ""
    if classification_review_items:
        classification_review_export = "\nOFFENE PRODUKTKLASSIFIKATIONEN\n"
        classification_review_export += (
            "  Kein Altpfad-Fallback; Übernahme bis zum Nachweis blockiert.\n"
        )
        for item in classification_review_items:
            classification_review_export += (
                f"  {item.get('ticker', ''):8s} {item.get('isin', '')}  "
                f"Pfad {item.get('routing_classification', '')}: "
                f"{item.get('review_reason', '')}\n"
            )

    partnership_export = ""
    if partnership_tax_items:
        partnership_export = "\nPERSONENGESELLSCHAFTEN - BERECHNUNG BLOCKIERT\n"
        partnership_export += (
            "  Nicht in KAP/KAP-INV enthalten. K-1/K-3 bzw. aequivalente "
            "Jahresallokation und deutsche Ueberleitung erforderlich.\n"
        )
        for isin, item in sorted(partnership_tax_items.items()):
            partnership_export += (
                f"  {item.get('ticker', isin):8s} {isin}  "
                f"Broker-PnL {fmt_de(item.get('observed_trade_pnl_eur', 0))} EUR, "
                f"Tageskurs {fmt_de(item.get('observed_tageskurs_delta_eur', 0))} EUR, "
                f"Ausschuettungen {fmt_de(item.get('observed_distributions_eur', 0))} EUR, "
                f"sonstige Cashwerte {fmt_de(item.get('observed_other_cash_eur', 0))} EUR, "
                f"QSt {fmt_de(item.get('observed_withholding_tax_eur', 0))} EUR\n"
            )

    report_text = f"""ANLAGE KAP {steuerjahr} - Steuerbericht
Erstellt: {created_at}
Basiswährung: {d.get('base_currency', 'USD')}
{multi_acct_export}
{classification_review_export}
{partnership_export}
═══════════════════════════════════════════════════
TOPF 1: AKTIEN (ohne ETF-Fonds)
  Aktiengewinne:         {fmt_de(final['stocks_gain']):>14} EUR
  Aktienverluste:        {fmt_de(final['stocks_loss']):>14} EUR
  ─────────────────────────────────────────────────
  Saldo Aktien:          {fmt_de(final['topf_1']):>14} EUR

TOPF 2: SONSTIGES (inkl. Termingeschäfte)
  Dividenden:            {fmt_de(final['dividends']):>14} EUR
  Zinsen (netto):        {fmt_de(final['interest']):>14} EUR
  Sonstige Gewinne:     {fmt_de(final['options_gain']):>14} EUR
  Sonstige Verluste:    {fmt_de(final['options_loss']):>14} EUR
  ─────────────────────────────────────────────────
  Saldo Sonstiges:       {fmt_de(final['topf_2']):>14} EUR
{topf2_detail_export}{special_products_export}{fx_export}{sh_export}{inv_export}
═══════════════════════════════════════════════════
ANLAGE KAP EINTRAGUNGEN
{"" if abs(final['zeile_7']) <= 0.01 else f"  Zeile 7 (inländischer Steuerabzug): {fmt_de(final['zeile_7']):>7} EUR" + chr(10) + f"  Zeile 37 (Kapitalertragsteuer): {fmt_de(final['zeile_37']):>10} EUR" + chr(10) + f"  Zeile 38 (Solidaritätszuschlag): {fmt_de(final['zeile_38']):>9} EUR" + chr(10)}
  Zeile 19 (Netto):      {fmt_de(final['zeile_19']):>14} EUR
  Zeile 20 (Aktiengewinne): {fmt_de(final['zeile_20']):>11} EUR
  Zeile 22 (Verluste o. Aktien): {fmt_de(final['zeile_22']):>8} EUR
  Zeile 23 (Aktienverluste): {fmt_de(final['zeile_23']):>11} EUR
  Zeile 41 (ausl. Quellensteuer): {fmt_de(final['quellensteuer']):>8} EUR
{de_kest_export}{kap_inv_entries_export}{"" if not has_so_data else chr(10) + "ANLAGE SO (§23 EStG): PRIVATE VERÄUSSERUNGSGESCHÄFTE" + chr(10) + f"  Physische Gold-ETCs (BFH VIII R 35/14, VIII R 4/15)" + chr(10) + f"  Steuerpflichtig (≤ 1J): {fmt_de(so_taxable):>12} EUR  → Anlage SO" + chr(10) + f"  Steuerfrei (> 1J):      {fmt_de(so_free):>12} EUR" + chr(10)}═══════════════════════════════════════════════════
"""
    return report_text


# ── Renderer: Export ─────────────────────────────────────────────────────────

def _build_export_trade_details():
    """Trade-Details inkl. Tageskurs-Korrekturzeilen (bei aktivem Toggle)."""
    trade_details = list(d.get('trade_details', []))
    if trade_details and tageskurs_aktiv:
        for lot in d.get('fx_correction_details', []):
            if abs(lot.get('delta_eur', 0)) < 0.005:
                continue
            underlying = (lot.get('underlyingSymbol', '')
                          or lot.get('symbol', '') or '').split()[0]
            open_dt = (lot.get('openDateTime', '') or '')[:10]
            close_dt = lot.get('reportDate', '')
            delta_eur = lot['delta_eur']
            note = (f'Tageskurs-Korrektur (Kauf {open_dt}, Kurs '
                    f'{lot["fx_open"]:.5f} → {lot["fx_close"]:.5f})')
            if lot.get('invstg_basis_adjustment_raw', 0) > 0:
                note += (' · KAP-INV-AK inkl. zusätzlicher ausländischer '
                         'Basisreduktion (z. B. ROC) auf Put-Strike '
                         'normalisiert')
            if lot.get('topf') == 'KAP-INV' and invstg_aktiv:
                isin = lot.get('isin', '')
                tfs_rate = etf_by_isin.get(isin, {}).get(
                    'tfs_rate', lot.get('tfs_rate', 0))
                taxable_delta = delta_eur * (1 - tfs_rate)
                if abs(taxable_delta - delta_eur) > 0.005:
                    note += (f' · roh {fmt_de(delta_eur)} EUR, stpfl. nach '
                             f'TFS {fmt_de(taxable_delta)} EUR')
                delta_eur = taxable_delta
            trade_details.append({
                'dateTime': close_dt, 'reportDate': close_dt,
                'symbol': lot.get('symbol', ''),
                'description': note,
                'isin': lot.get('isin', ''),
                'assetCategory': lot.get('assetCategory', ''),
                'subCategory': lot.get('subCategory', ''),
                'buySell': '', 'openClose': '',
                'quantity': lot.get('quantity', ''),
                'transactionType': 'FX-Korrektur',
                'currency': lot.get('currency', ''), 'tradePrice': 0,
                'cost': lot.get('cost', 0),
                'proceeds': 0, 'fifoPnlRealized': 0, 'fxRateToBase': 0,
                'pnl_eur': delta_eur, 'topf': lot.get('topf', 'Topf2'),
                'strike': '', 'expiry': '', 'putCall': '', 'multiplier': '',
                'underlyingSymbol': underlying,
                'source': 'tageskurs_korrektur',
            })
        trade_details.sort(
            key=lambda r: r.get('dateTime', '') or r.get('reportDate', '')
            or 'zzzz')
    return trade_details


def _build_exports():
    """XLSX + TXT genau einmal pro view_key erzeugen (lazy, nur im aktiven
    Export-Renderer). Nach jeder Fondsbestätigung oder Toggle-Änderung
    ändert sich der view_key; ein alter Export wird nie ausgeliefert."""
    cache = st.session_state.get('export_cache')
    if cache and cache.get('view_key') == vm['view_key']:
        return cache

    trade_details = _build_export_trade_details()
    trades_by_topf = defaultdict(list)
    for row in trade_details:
        trades_by_topf[row.get('topf', 'Topf2')].append(row)

    export_context = {
        'final': final,
        'base_currency': d.get('base_currency', 'USD'),
        'created_at': created_at,
        'has_etf_data': has_etf_data,
        'invstg_aktiv': invstg_aktiv,
        'kap_inv_form': kap_inv_form if (has_etf_data and invstg_aktiv) else {},
        'no_invstg_summary': no_invstg_summary,
        'partnership_tax_items': d.get('partnership_tax_items', {}) or {},
        'has_so_data': has_so_data,
        'so_taxable': so_taxable,
        'so_free': so_free,
    }
    xlsx_data = None
    xlsx_error = None
    try:
        xlsx_data = _build_excel(trade_details, trades_by_topf, export_context)
    except ModuleNotFoundError as exc:
        if exc.name != 'openpyxl':
            raise
        xlsx_error = ("Excel-Export nicht verfügbar: openpyxl installieren "
                      "(pip install openpyxl).")

    cache = {
        'view_key': vm['view_key'],
        'xlsx': xlsx_data,
        'xlsx_error': xlsx_error,
        'txt': _build_text_report(),
        'n_details': len(trade_details),
        'trade_sums': {
            key: sum(float(r.get('pnl_eur') or 0) for r in rows)
            for key, rows in trades_by_topf.items()
        },
        'n_trades': sum(
            1 for r in trade_details if r.get('source') == 'trades'),
        'n_underlyings': len(set(
            (r.get('underlyingSymbol', '') or r.get('symbol', '') or '?')
            .split()[0]
            for r in trade_details if r.get('source') == 'trades')),
    }
    st.session_state['export_cache'] = cache
    return cache


def render_export():
    st.markdown('<p class="page-title">Export</p>', unsafe_allow_html=True)
    st.caption(
        "Excel mit Einzelnachweisen je Topf, Textreport zum Kopieren. Beide "
        "enthalten dieselben finalen Werte wie die Übersicht."
    )
    exports = _build_exports()

    st.markdown(
        '<div class="notice transparenz">'
        '<div class="notice-title">Formularwerte und Detailwerte sind zwei '
        'verschiedene Ebenen</div>'
        '<strong>Für die Steuererklärung</strong> gelten die Werte aus der '
        'Übersicht, dem Excel-Blatt „Zusammenfassung“ und dem Textreport. '
        '<strong>Für die Kontrolle</strong> folgen darunter die Summen der '
        'einzelnen Trades, Korrekturen und Zuflüsse. Diese Detail-Summen '
        'dürfen von den Formularwerten abweichen.'
        '</div>',
        unsafe_allow_html=True,
    )

    n_trades = exports['n_trades']
    n_korr = exports['n_details'] - n_trades
    if n_trades > 0:
        header = (f"**Detailabstimmung: {n_trades} Trades, "
                  f"{exports['n_underlyings']} Wertpapiere")
        if n_korr > 0:
            header += f" (+ {n_korr} Korrekturen/Zuflüsse)"
        header += "**"
    elif n_korr > 0:
        header = (f"**Detailabstimmung: keine Trades, {n_korr} "
                  "Korrektur-/Zuflusspositionen**")
    else:
        header = "**Keine Trade-Details für die Detailabstimmung verfügbar**"
    summary_lines = [header]
    for topf_key in EXPORT_TOPF_ORDER:
        if topf_key in exports['trade_sums']:
            label = topf_readable.get(topf_key, topf_key).split(' - ')[0]
            summary_lines.append(
                f"Details {label}: {fmt_de(exports['trade_sums'][topf_key])} EUR")
    st.markdown(" | ".join(summary_lines))
    st.caption(
        "Warum können die Summen abweichen? Die Detailpositionen bilden "
        "Trades und einzelne Korrekturen ab. Die Formularwerte berücksichtigen "
        "zusätzlich Cash-Buchungen, steuerliche Zuordnungen, "
        "Teilfreistellungen und die Verlustverrechnung. Die "
        "Detailabstimmung ist deshalb ein Kontrollnachweis und kein zweiter "
        "Satz Formularwerte."
    )

    section_title("Excel-Steuerreport")
    if exports['xlsx_error']:
        st.warning(exports['xlsx_error'])
    elif exports['xlsx'] is not None:
        st.download_button(
            label=(f"Steuerbericht als Excel herunterladen "
                   f"({exports['n_details']} Detailpositionen)"),
            data=exports['xlsx'],
            file_name=f"steuerbericht_{steuerjahr}.xlsx",
            mime=("application/vnd.openxmlformats-officedocument."
                  "spreadsheetml.sheet"),
            use_container_width=True,
        )

    section_title("Textreport")
    st.download_button(
        label="Textreport herunterladen",
        data=exports['txt'],
        file_name=f"steuerbericht_{steuerjahr}.txt",
        mime="text/plain",
        use_container_width=True,
    )
    with st.expander("Report als Text anzeigen (zum Kopieren)"):
        st.code(exports['txt'], language=None)


# ── Dispatch: genau ein aktiver Renderer ─────────────────────────────────────

_PAGE_RENDERERS = {
    'overview': render_overview,
    'kap': render_kap,
    'kap_inv': render_kap_inv,
    'anlage_so': render_anlage_so,
    'prueffaelle': render_prueffaelle,
    'rechenwege': render_rechenwege,
    'export': render_export,
}

_PAGE_RENDERERS[_nav_current]()
